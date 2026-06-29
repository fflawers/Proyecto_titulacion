# =========================================
# main.py — FastAPI Backend
# =========================================

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional, List
from urllib.parse import quote

import auth
import database
import ai_engine
import pdf_manager
import excel_manager
import gemini_vision
from config import CORS_ORIGINS, ADMIN_HISTORIAL_PASSWORD
import threading
import base64
import json
import io

# =========================================
# APP
# =========================================

app = FastAPI(
    title="LUXO API",
    description="API del asistente inteligente de Sunglass Hut",
    version="2.0.0",
)

# Límite de tamaño de archivos subidos: 50 MB
# (suficiente para cualquier Excel/PDF operativo de Sunglass Hut)
MAX_UPLOAD_SIZE = 50 * 1024 * 1024  # 50 MB en bytes


app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS if CORS_ORIGINS != ["*"] else [],
    allow_origin_regex=".*" if CORS_ORIGINS == ["*"] else None,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================
# SEGURIDAD — JWT
# =========================================

security = HTTPBearer()


def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Dependencia que extrae y valida el JWT del header Authorization."""
    user = auth.verificar_token(credentials.credentials)
    if not user:
        raise HTTPException(status_code=401, detail="Token inválido o expirado")
    return user


def require_admin(user=Depends(get_current_user)):
    """Dependencia que verifica que el usuario es admin."""
    if user["rol"] != "Admin":
        raise HTTPException(status_code=403, detail="Acceso solo para administradores")
    return user


# =========================================
# MODELOS (Request/Response)
# =========================================

class LoginRequest(BaseModel):
    usuario: str
    contrasena: str


class LoginResponse(BaseModel):
    token: str
    nombre: str
    rol: str


class ChatRequest(BaseModel):
    pregunta: str
    idioma: Optional[str] = "es"


class ChatResponse(BaseModel):
    respuesta: str
    intencion: str
    id_manual: Optional[int] = None
    nombre_pdf: Optional[str] = None
    id_conversacion: Optional[int] = None
    es_abierto: Optional[bool] = True
    sugiere_ticket: Optional[bool] = False


class FeedbackRequest(BaseModel):
    id_conversacion: int
    es_positivo: bool
    comentario: Optional[str] = None


class SugerenciaRequest(BaseModel):
    sugerencia: str


class ManualResponse(BaseModel):
    id: int
    nombre_archivo: Optional[str] = None
    titulo: Optional[str] = None
    version: Optional[str] = None
    Abierto: Optional[bool] = True


class MessageResponse(BaseModel):
    message: str
    success: bool


# =========================================
# ENDPOINTS — AUTENTICACIÓN
# =========================================

@app.post("/api/auth/login", response_model=LoginResponse)
def login_endpoint(req: LoginRequest):
    """Login — verifica credenciales y retorna JWT."""
    user_data = auth.login(req.usuario, req.contrasena)

    if not user_data:
        # Mensaje específico
        if database.usuario_existe(req.usuario):
            raise HTTPException(status_code=401, detail="Contraseña incorrecta")
        else:
            raise HTTPException(status_code=401, detail="Usuario no registrado")

    token = auth.crear_token(user_data)

    # Registrar sesión (fire-and-forget)
    try:
        database.registrar_sesion(user_data["id"], ip=None, ciudad=None, pais=None)
    except Exception:
        pass

    return LoginResponse(
        token=token,
        nombre=user_data["nombre"],
        rol=user_data["rol"],
    )


@app.get("/api/auth/me")
def me_endpoint(user=Depends(get_current_user)):
    """Retorna la info del usuario autenticado, incluida su tienda."""
    return user


# =========================================
# ENDPOINTS — CHAT
# =========================================

@app.post("/api/chat", response_model=ChatResponse)
def chat_endpoint(
    pregunta: str = Form(...),
    idioma: str = Form("es"),
    archivo: UploadFile = File(None),
    user=Depends(get_current_user)
):
    """Envía una pregunta y recibe respuesta con RAG."""
    if not pregunta.strip():
        raise HTTPException(status_code=400, detail="La pregunta no puede estar vacía")

    archivo_bytes = None
    archivo_tipo = None
    if archivo and archivo.filename:
        archivo_bytes = archivo.file.read()
        archivo_tipo = archivo.content_type

    # Llamar al motor pasándole la pregunta, el usuario y el idioma
    resultado = ai_engine.generar_respuesta(
        pregunta, user["id"], idioma=idioma,
        archivo_bytes=archivo_bytes, archivo_tipo=archivo_tipo
    )

    return ChatResponse(
        respuesta=resultado["respuesta"],
        intencion=resultado["intencion"],
        id_manual=resultado.get("id_manual"),
        nombre_pdf=resultado.get("nombre_pdf"),
        id_conversacion=resultado.get("id_conversacion"),
        sugiere_ticket=resultado.get("sugiere_ticket", False),
    )


# =========================================
# ENDPOINTS — FEEDBACK
# =========================================

@app.post("/api/feedback", response_model=MessageResponse)
def feedback_endpoint(req: FeedbackRequest, user=Depends(get_current_user)):
    """Guarda feedback (👍/👎) y opcionalmente un comentario de falla."""
    database.guardar_feedback(req.id_conversacion, req.es_positivo, req.comentario)
    return MessageResponse(
        message="Feedback registrado",
        success=True,
    )


# =========================================
# ENDPOINTS — MANUALES
# =========================================

@app.get("/api/manuales", response_model=list[ManualResponse])
def listar_manuales(user=Depends(get_current_user)):
    """Lista todos los manuales disponibles."""
    manuales = database.obtener_manuales_listado()
    return [
        ManualResponse(
            id=m["ID_Manual"],
            nombre_archivo=m.get("Nombre_Archivo"),
            titulo=m.get("Titulo"),
            version=m.get("Version"),
            Abierto=bool(m.get("Abierto", 1))
        )
        for m in manuales
    ]


@app.post("/api/manuales/upload", response_model=MessageResponse)
async def upload_manual(
    archivo: UploadFile = File(...),
    user=Depends(require_admin),
):
    """Sube un nuevo manual PDF, Excel o imagen (solo admin)."""
    nombre = archivo.filename.lower()
    es_pdf   = nombre.endswith(".pdf")
    es_excel = nombre.endswith(".xlsx") or nombre.endswith(".xls")
    es_imagen = nombre.endswith(".jpg") or nombre.endswith(".jpeg") or nombre.endswith(".png")

    if not es_pdf and not es_excel and not es_imagen:
        raise HTTPException(
            status_code=400,
            detail="Solo se permiten archivos PDF (.pdf), Excel (.xlsx, .xls) o imagen (.jpg, .png)",
        )

    contenido = await archivo.read()

    if len(contenido) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"El archivo supera el límite de {MAX_UPLOAD_SIZE // (1024*1024)} MB.",
        )

    if es_pdf:
        exito, mensaje = pdf_manager.cargar_pdf(archivo.filename, contenido)
    elif es_excel:
        exito, mensaje = excel_manager.cargar_excel(archivo.filename, contenido)
    else:  # imagen
        exito, mensaje = pdf_manager.cargar_imagen(archivo.filename, contenido)

    if not exito:
        raise HTTPException(status_code=400, detail=mensaje)

    return MessageResponse(message=mensaje, success=True)


@app.put("/api/manuales/update", response_model=MessageResponse)
async def update_manual(
    archivo: UploadFile = File(...),
    user=Depends(require_admin),
):
    """Actualiza un manual PDF, Excel o imagen existente (solo admin)."""
    nombre = archivo.filename.lower()
    es_pdf   = nombre.endswith(".pdf")
    es_excel = nombre.endswith(".xlsx") or nombre.endswith(".xls")
    es_imagen = nombre.endswith(".jpg") or nombre.endswith(".jpeg") or nombre.endswith(".png")

    if not es_pdf and not es_excel and not es_imagen:
        raise HTTPException(
            status_code=400,
            detail="Solo se permiten archivos PDF (.pdf), Excel (.xlsx, .xls) o imagen (.jpg, .png)",
        )

    contenido = await archivo.read()

    if len(contenido) > MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"El archivo supera el límite de {MAX_UPLOAD_SIZE // (1024*1024)} MB.",
        )

    if es_pdf:
        exito, mensaje = pdf_manager.actualizar_pdf(archivo.filename, contenido)
    elif es_excel:
        exito, mensaje = excel_manager.actualizar_excel(archivo.filename, contenido)
    else:  # imagen
        exito, mensaje = pdf_manager.cargar_imagen(archivo.filename, contenido)

    if not exito:
        raise HTTPException(status_code=400, detail=mensaje)

    return MessageResponse(message=mensaje, success=True)


# =========================================
# ENDPOINTS — ADMIN: RE-INDEXAR
# =========================================

@app.post("/api/admin/reindexar")
def reindexar_manuales(user=Depends(require_admin)):
    """
    Re-indexa todos los manuales existentes con el pipeline mejorado.
    Útil después de actualizar los parámetros de chunking o el modelo de embeddings.
    Solo admin.
    """
    import vector_store as vs
    try:
        vs.reindexar_todos()
        coleccion = vs.obtener_coleccion()
        total_chunks = coleccion.count()
        manuales = database.obtener_manuales()
        return {
            "success": True,
            "message": f"Re-indexación completada: {len(manuales)} manuales, {total_chunks} chunks.",
            "total_manuales": len(manuales),
            "total_chunks": total_chunks,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error en re-indexación: {e}")


@app.delete("/api/manuales/{id_manual}", response_model=MessageResponse)
def delete_manual(id_manual: int, user=Depends(require_admin)):
    """Borra un manual (solo admin)."""
    exito, mensaje = pdf_manager.borrar_manual(id_manual)

    if not exito:
        raise HTTPException(status_code=400, detail=mensaje)

    return MessageResponse(message=mensaje, success=True)


@app.get("/api/manuales/{id_manual}/download")
def download_manual(id_manual: int, user=Depends(get_current_user)):
    """Descarga el PDF de un manual."""
    pdf_data = pdf_manager.obtener_pdf_para_descarga(id_manual)

    if not pdf_data:
        raise HTTPException(status_code=404, detail="PDF no encontrado")

    nombre = pdf_data["nombre"]
    nombre_encoded = quote(nombre)
    return Response(
        content=pdf_data["contenido_bytes"],
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{nombre_encoded}"
        },
    )


@app.get("/api/manuales/{id_manual}/download-excel")
def download_excel(id_manual: int, user=Depends(get_current_user)):
    """Descarga el Excel de un manual."""
    excel_data = excel_manager.obtener_excel_para_descarga(id_manual)

    datos = excel_manager.obtener_excel_para_descarga(id_manual)
    if datos:
        return Response(content=datos["Archivo_Excel"], media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    raise HTTPException(status_code=404, detail="Excel no encontrado")

@app.put("/api/admin/manuales/{id_manual}/toggle-abierto")
def toggle_manual_abierto_endpoint(id_manual: int, user=Depends(require_admin)):
    if database.toggle_manual_abierto(id_manual):
        return {"success": True, "message": "Estado del manual actualizado correctamente"}
    raise HTTPException(status_code=500, detail="Error al actualizar el estado del manual")


# =========================================
# ENDPOINTS — HISTORIAL & PENDIENTES
# =========================================

@app.get("/api/historial/me")
def historial_propio(limite: int = 50, user=Depends(get_current_user)):
    """Retorna el historial de consultas del usuario autenticado."""
    historial = database.obtener_historial_usuario(user["id"], limite)
    return historial


# =========================================
# ENDPOINTS — ADMIN: HISTORIAL DE CONSULTAS
# =========================================

class VerifyPasswordRequest(BaseModel):
    password: str


@app.post("/api/admin/verify-historial")
def verify_historial_password(req: VerifyPasswordRequest, user=Depends(require_admin)):
    """Verifica la contraseña de dev para acceder al historial."""
    if req.password != ADMIN_HISTORIAL_PASSWORD:
        raise HTTPException(status_code=403, detail="Contraseña incorrecta")
    return {"success": True}


@app.get("/api/admin/historial")
def admin_historial(
    limite: int = 100,
    user=Depends(require_admin),
    x_historial_password: Optional[str] = None,
):
    """Retorna el historial completo de consultas de todos los usuarios (solo admin + contraseña)."""
    if x_historial_password != ADMIN_HISTORIAL_PASSWORD:
        raise HTTPException(status_code=403, detail="Contraseña de historial requerida")
    historial = database.obtener_historial_admin(limite)
    return historial


# =========================================
# ENDPOINTS — ADMIN: PENDIENTES (sin respuesta)
# =========================================

@app.get("/api/admin/pendientes")
def admin_pendientes(
    limite: int = 200,
    user=Depends(require_admin),
):
    """Retorna preguntas que LUXO no pudo responder (solo admin)."""
    pendientes = database.obtener_pendientes(limite)
    return pendientes


# =========================================
# ENDPOINTS — ADMIN: ESTADÍSTICAS
# =========================================

@app.get("/api/admin/estadisticas")
def admin_estadisticas(user=Depends(require_admin)):
    """Retorna estadísticas de uso del sistema (solo admin)."""
    stats = database.obtener_estadisticas()
    return stats


# =========================================
# ENDPOINTS — ADMIN: USUARIOS / TIENDAS
# =========================================

class ActualizarTiendaRequest(BaseModel):
    tienda: str


@app.get("/api/admin/usuarios")
def admin_usuarios(user=Depends(require_admin)):
    """Lista todos los usuarios con su tienda asignada (solo admin)."""
    usuarios = database.obtener_usuarios_admin()
    return usuarios


@app.put("/api/admin/usuarios/{id_usuario}/tienda")
def actualizar_tienda(
    id_usuario: int,
    req: ActualizarTiendaRequest,
    user=Depends(require_admin),
):
    """Actualiza la tienda asignada a un usuario (solo admin)."""
    exito = database.actualizar_tienda_usuario(id_usuario, req.tienda)
    if not exito:
        raise HTTPException(status_code=500, detail="No se pudo actualizar la tienda")
    return {"message": "Tienda actualizada correctamente", "success": True}


# =========================================
# HEALTH CHECK
# =========================================

@app.get("/api/health")
def health():
    """Health check del servidor."""
    return {"status": "ok", "service": "LUXO API"}


# =========================================
# ENDPOINTS — NOTIFICACIONES
# =========================================

@app.get("/api/notificaciones")
def get_notificaciones(user=Depends(get_current_user)):
    """Retorna las notificaciones del usuario autenticado."""
    return database.obtener_notificaciones(user["id"])


@app.get("/api/notificaciones/no-leidas-count")
def count_no_leidas(user=Depends(get_current_user)):
    """Retorna el conteo de notificaciones no leídas."""
    return {"count": database.contar_no_leidas(user["id"])}


@app.post("/api/notificaciones/marcar-leidas")
def marcar_leidas(user=Depends(get_current_user)):
    """Marca todas las notificaciones del usuario como leídas."""
    database.marcar_notificaciones_leidas(user["id"])
    return {"success": True}


class NotificacionRequest(BaseModel):
    id_usuario: int
    titulo: str
    cuerpo: str
    tipo: Optional[str] = "general"


class NotificacionRolRequest(BaseModel):
    rol: str
    titulo: str
    cuerpo: str
    tipo: Optional[str] = "general"


@app.post("/api/admin/notificaciones")
def enviar_notificacion(req: NotificacionRequest, user=Depends(require_admin)):
    """Envía una notificación a un usuario específico (solo admin)."""
    database.crear_notificacion(req.id_usuario, req.titulo, req.cuerpo, req.tipo)
    return {"success": True}


@app.post("/api/admin/notificaciones/rol")
def enviar_notificacion_rol(req: NotificacionRolRequest, user=Depends(require_admin)):
    """Envía una notificación a todos los usuarios de un rol (solo admin)."""
    database.crear_notificacion_rol(req.rol, req.titulo, req.cuerpo, req.tipo)
    return {"success": True}


# =========================================
# ENDPOINTS — SESIONES
# =========================================

@app.get("/api/admin/sesiones")
def admin_sesiones(limite: int = 50, user=Depends(require_admin)):
    """Historial de inicios de sesión (solo admin)."""
    return database.obtener_sesiones_admin(limite)


# =========================================
# ENDPOINTS — SUGERENCIAS
# =========================================

@app.post("/api/sugerencias")
def post_sugerencia(req: SugerenciaRequest, user=Depends(get_current_user)):
    """Envía una nueva sugerencia."""
    if not req.sugerencia.strip():
        raise HTTPException(status_code=400, detail="La sugerencia no puede estar vacía")
    if not database.guardar_sugerencia(user["id"], req.sugerencia):
        raise HTTPException(status_code=500, detail="Error al guardar la sugerencia")
    return {"message": "Sugerencia enviada correctamente", "success": True}

@app.get("/api/admin/sugerencias")
def get_sugerencias(user=Depends(require_admin)):
    """Obtiene la lista de sugerencias de los usuarios."""
    return database.obtener_sugerencias_admin()


# =========================================
# ENDPOINTS — TICKETS DE SOPORTE
# =========================================

class TicketRequest(BaseModel):
    detalle: str


class ResolverTicketRequest(BaseModel):
    respuesta: str


@app.post("/api/tickets")
def crear_ticket(req: TicketRequest, user=Depends(get_current_user)):
    """Crea un ticket de soporte."""
    id_ticket = database.crear_ticket(user["id"], req.detalle)
    if not id_ticket:
        raise HTTPException(status_code=500, detail="Error al crear el ticket")
    # Notificar a admins
    database.crear_notificacion_rol(
        "Admin",
        "🎫 Nuevo Ticket de Soporte",
        f"{user['nombre']} ha reportado un problema: {req.detalle[:100]}...",
        "ticket",
    )
    return {"id_ticket": id_ticket, "success": True}


@app.get("/api/tickets/mis-tickets")
def mis_tickets(user=Depends(get_current_user)):
    """Retorna los tickets de soporte del usuario autenticado."""
    return database.obtener_tickets_usuario(user["id"])


@app.get("/api/admin/tickets")
def admin_tickets(user=Depends(require_admin)):
    """Lista todos los tickets de soporte (solo admin)."""
    return database.obtener_tickets_admin()


@app.put("/api/admin/tickets/{id_ticket}/resolver")
def resolver_ticket(
    id_ticket: int,
    req: ResolverTicketRequest,
    user=Depends(require_admin),
):
    """Marca un ticket como resuelto (solo admin)."""
    exito = database.resolver_ticket(id_ticket, req.respuesta)
    if not exito:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    return {"success": True}


class PrioridadTicketRequest(BaseModel):
    prioridad: str  # 'Normal' | 'Alta' | 'Urgente'


@app.put("/api/admin/tickets/{id_ticket}/prioridad")
def cambiar_prioridad_ticket(
    id_ticket: int,
    req: PrioridadTicketRequest,
    user=Depends(require_admin),
):
    """Cambia la prioridad de un ticket (solo admin)."""
    if req.prioridad not in ('Normal', 'Alta', 'Urgente'):
        raise HTTPException(status_code=400, detail="Prioridad inválida. Usa: Normal, Alta o Urgente")
    exito = database.marcar_prioridad_ticket(id_ticket, req.prioridad)
    if not exito:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    return {"success": True, "prioridad": req.prioridad}


# =========================================
# ENDPOINTS — CHECKLISTS OPERATIVOS
# =========================================

class ToggleChecklistRequest(BaseModel):
    id_plantilla: int
    completado: bool


class AgregarTareaRequest(BaseModel):
    categoria: int  # 1=Apertura, 2=Cierre, 3=Venta
    descripcion: str
    prioridad: str = "Normal"
    notas: str = None


@app.get("/api/checklists")
def get_checklists(user=Depends(get_current_user)):
    """Retorna todas las tareas de checklist con el estado del usuario para hoy."""
    plantillas = database.obtener_plantillas_checklist()
    completadas = database.obtener_completadas_hoy(user["id"])
    result = []
    for p in plantillas:
        result.append({
            **p,
            "completado": p["ID_Plantilla"] in completadas,
        })
    return result


@app.post("/api/checklists/toggle")
def toggle_checklist(req: ToggleChecklistRequest, user=Depends(get_current_user)):
    """Marca o desmarca una tarea del checklist diario."""
    database.toggle_checklist(user["id"], req.id_plantilla, req.completado)
    return {"success": True}


@app.post("/api/admin/checklists/tarea")
def agregar_tarea_checklist(req: AgregarTareaRequest, user=Depends(require_admin)):
    """Agrega una nueva tarea al checklist (solo admin)."""
    id_p = database.agregar_tarea_checklist(
        req.categoria, 
        req.descripcion,
        req.prioridad,
        req.notas
    )
    if not id_p:
        raise HTTPException(status_code=500, detail="Error al agregar la tarea")
    return {"id_plantilla": id_p, "success": True}


@app.delete("/api/admin/checklists/tarea/{id_plantilla}")
def eliminar_tarea_checklist(id_plantilla: int, user=Depends(require_admin)):
    """Elimina una tarea del checklist (solo admin)."""
    exito = database.eliminar_tarea_checklist(id_plantilla)
    if not exito:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    return {"success": True}


# =========================================
# ENDPOINTS — TAREAS CONSOLIDADAS
# =========================================

class CrearTareaRequest(BaseModel):
    titulo: str
    descripcion: Optional[str] = ""
    fecha_limite: Optional[str] = None  # YYYY-MM-DD HH:MM


class ResponderTareaRequest(BaseModel):
    id_tarea: int
    respuestas: dict  # {"columna": "valor", ...}


@app.get("/api/tareas")
def get_tareas(user=Depends(get_current_user)):
    """Retorna todas las tareas disponibles."""
    database.cerrar_tareas_vencidas()
    return database.obtener_tareas_activas()


@app.post("/api/tareas/responder")
def responder_tarea(req: ResponderTareaRequest, user=Depends(get_current_user)):
    """Guarda la respuesta de un usuario a una tarea."""
    exito = database.guardar_respuesta_tarea(
        req.id_tarea,
        user["id"],
        user.get("tienda", ""),
        json.dumps(req.respuestas, ensure_ascii=False),
    )
    if not exito:
        raise HTTPException(status_code=500, detail="Error al guardar la respuesta")
    return {"success": True}


@app.post("/api/admin/tareas/crear")
async def crear_tarea(
    titulo: str,
    descripcion: Optional[str] = "",
    fecha_limite: Optional[str] = None,
    plantilla: UploadFile = File(None),
    user=Depends(require_admin),
):
    """Crea una nueva tarea con plantilla Excel opcional (solo admin)."""
    plantilla_bytes = None
    nombre_plantilla = None
    columnas_json = "[]"

    if plantilla:
        contenido = await plantilla.read()
        if len(contenido) > MAX_UPLOAD_SIZE:
            raise HTTPException(status_code=413, detail="Archivo demasiado grande")
        plantilla_bytes = contenido
        nombre_plantilla = plantilla.filename

        # Extraer columnas del Excel
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
            ws = wb.active
            omit_kw = {"tienda", "sucursal", "gerente", "usuario", "fecha"}
            cols = []
            for row in ws.iter_rows(values_only=True):
                if any(c is not None for c in row):
                    cols = [
                        str(c).strip() for c in row
                        if c is not None and not any(k in str(c).lower() for k in omit_kw)
                    ]
                    break
            columnas_json = json.dumps(cols, ensure_ascii=False)
        except Exception as e:
            print("Error leyendo columnas Excel:", e)

    id_tarea = database.crear_tarea(
        titulo, descripcion, plantilla_bytes, nombre_plantilla, columnas_json, fecha_limite
    )
    if not id_tarea:
        raise HTTPException(status_code=500, detail="Error al crear la tarea")

    # Notificar a gerentes
    database.crear_notificacion_rol(
        "Gerente",
        f"📋 Nueva Tarea: {titulo}",
        descripcion or "Revisa el portal para ver los detalles.",
        "tarea",
    )
    return {"id_tarea": id_tarea, "success": True}


@app.get("/api/admin/tareas/{id_tarea}/respuestas")
def get_respuestas_tarea(id_tarea: int, user=Depends(require_admin)):
    """Obtiene todas las respuestas de una tarea (solo admin)."""
    return database.obtener_respuestas_tarea(id_tarea)


@app.get("/api/admin/tareas/{id_tarea}/consolidado")
def descargar_consolidado(
    id_tarea: int,
    user=Depends(require_admin),
):
    """Genera y descarga el Excel consolidado de respuestas de una tarea."""
    tarea_info = database.obtener_plantilla_tarea(id_tarea)
    if not tarea_info or not tarea_info.get("Plantilla_Bytes"):
        raise HTTPException(status_code=404, detail="No hay plantilla para esta tarea")

    respuestas = database.obtener_respuestas_tarea(id_tarea)

    try:
        import openpyxl
        template_bytes = tarea_info["Plantilla_Bytes"]
        wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
        ws = wb.active

        # Leer fila de cabeceras
        headers = {}
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if any(c is not None for c in row):
                for c_idx, cell in enumerate(row, start=1):
                    if cell is not None:
                        headers[str(cell).strip()] = c_idx
                break

        next_row = ws.max_row + 1
        for r_item in respuestas:
            vals = json.loads(r_item["Respuestas_JSON"])
            for h_name, col_idx in headers.items():
                h_lower = h_name.lower()
                if any(k in h_lower for k in ["tienda", "sucursal"]):
                    ws.cell(row=next_row, column=col_idx, value=r_item["Tienda"])
                elif any(k in h_lower for k in ["gerente", "usuario"]):
                    ws.cell(row=next_row, column=col_idx, value=r_item["Gerente"])
                elif "fecha" in h_lower:
                    ws.cell(row=next_row, column=col_idx, value=r_item["Fecha_Envio"])
                else:
                    ws.cell(row=next_row, column=col_idx, value=vals.get(h_name, ""))
            next_row += 1

        out_buf = io.BytesIO()
        wb.save(out_buf)
        xlsx_bytes = out_buf.getvalue()
        nombre = tarea_info["Nombre_Plantilla"] or "consolidado.xlsx"
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{nombre}"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generando consolidado: {e}")


# =========================================
# ENDPOINTS — CAMPAÑAS DE EXHIBICIÓN
# =========================================

@app.get("/api/campanas")
def get_campanas(user=Depends(get_current_user)):
    """Retorna la campaña activa (todos los roles)."""
    campana = database.obtener_campana_activa()
    return campana or {}


@app.get("/api/campanas/todas")
def get_todas_campanas(user=Depends(require_admin)):
    """Lista todas las campañas (solo admin)."""
    return database.obtener_todas_campanas()


@app.get("/api/campanas/{id_campana}/fotos-guia")
async def get_fotos_guia(
    id_campana: int,
    segmento: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Retorna las fotos guía de una campaña (con imagen en base64)."""
    fotos = database.obtener_fotos_guia(id_campana, segmento)
    result = []
    for f in fotos:
        img_b64 = base64.b64encode(f["Imagen_Bytes"]).decode("utf-8") if f.get("Imagen_Bytes") else None
        result.append({
            "ID_Foto_Guia": f["ID_Foto_Guia"],
            "Nombre_Foto": f["Nombre_Foto"],
            "Instrucciones": f["Instrucciones"],
            "Segmento": f["Segmento"],
            "imagen_b64": img_b64,
        })
    return result


@app.get("/api/campanas/{id_campana}/mi-entrega")
def get_mi_entrega(
    id_campana: int,
    user=Depends(get_current_user),
):
    """Obtiene o crea el registro de entrega de la tienda del usuario autenticado."""
    tienda = user.get("tienda", "")
    if not tienda:
        raise HTTPException(status_code=400, detail="No tienes una tienda asignada")
    entrega = database.obtener_o_crear_entrega(id_campana, tienda, user["id"])
    if not entrega:
        raise HTTPException(status_code=500, detail="Error al obtener la entrega")
    fotos = database.obtener_fotos_tienda(entrega["ID_Entrega"])
    return {
        **entrega,
        "fotos": {str(k): v for k, v in fotos.items()},
    }


@app.post("/api/campanas/{id_campana}/fotos")
async def subir_foto_tienda(
    id_campana: int,
    id_foto_guia: int,
    foto: UploadFile = File(...),
    user=Depends(get_current_user),
):
    """El gerente sube una foto de su tienda para auditoría."""
    tienda = user.get("tienda", "")
    if not tienda:
        raise HTTPException(status_code=400, detail="No tienes una tienda asignada")

    contenido = await foto.read()
    if len(contenido) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="Imagen demasiado grande (máx 50 MB)")

    # Obtener o crear entrega
    entrega = database.obtener_o_crear_entrega(id_campana, tienda, user["id"])
    if not entrega:
        raise HTTPException(status_code=500, detail="Error al procesar la entrega")
    id_entrega = entrega["ID_Entrega"]

    # Guardar foto en BD con estado inicial
    database.guardar_foto_tienda(id_entrega, id_foto_guia, contenido)

    # Notificar a admins
    database.crear_notificacion_rol(
        "Admin",
        "📸 Nueva Foto de Campaña",
        f"La tienda '{tienda}' ha subido una foto para revisión.",
        "campana",
    )

    # Lanzar auditoría IA en hilo separado
    def thread_auditoria():
        try:
            fotos_guia = database.obtener_fotos_guia(id_campana)
            guia_row = next((g for g in fotos_guia if g["ID_Foto_Guia"] == id_foto_guia), None)
            if guia_row and guia_row.get("Imagen_Bytes"):
                estatus, resultado = gemini_vision.auditar_foto(
                    guia_row["Imagen_Bytes"],
                    contenido,
                    guia_row.get("Instrucciones", ""),
                    guia_row.get("Nombre_Foto", ""),
                )
                database.actualizar_auditoria_foto(id_entrega, id_foto_guia, estatus, resultado)
                # Notificar al gerente
                database.crear_notificacion(
                    user["id"],
                    f"{'✅ Foto Aprobada' if estatus == 'Aprobado' else '⚠️ Foto Requiere Corrección'}",
                    resultado[:200],
                    "campana",
                )
        except Exception as e:
            print("ERROR THREAD AUDITORIA:", e)

    threading.Thread(target=thread_auditoria, daemon=True).start()

    return {"success": True, "message": "Foto subida. Auditoría IA en proceso..."}


@app.post("/api/admin/campanas/crear")
async def crear_campana(
    nombre: str,
    descripcion: Optional[str] = "",
    guia_pdf: UploadFile = File(None),
    user=Depends(require_admin),
):
    """Crea una nueva campaña y desactiva las anteriores (solo admin)."""
    pdf_nombre = None
    pdf_bytes = None
    if guia_pdf:
        pdf_bytes = await guia_pdf.read()
        pdf_nombre = guia_pdf.filename

    id_campana = database.crear_campana(nombre, descripcion, pdf_nombre, pdf_bytes)
    if not id_campana:
        raise HTTPException(status_code=500, detail="Error al crear la campaña")

    # Notificar a gerentes
    database.crear_notificacion_rol(
        "Gerente",
        f"📸 Nueva Campaña Activa: {nombre}",
        descripcion or "Revisa el portal para ver los detalles.",
        "campana",
    )
    return {"id_campana": id_campana, "success": True}


@app.post("/api/admin/campanas/{id_campana}/fotos-guia")
async def agregar_foto_guia(
    id_campana: int,
    nombre_foto: str,
    instrucciones: Optional[str] = "",
    segmento: Optional[str] = "Todos",
    imagen: UploadFile = File(...),
    user=Depends(require_admin),
):
    """Agrega una foto guía a la campaña (solo admin)."""
    contenido = await imagen.read()
    if len(contenido) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail="Imagen demasiado grande")

    id_foto = database.agregar_foto_guia(id_campana, nombre_foto, instrucciones, contenido, segmento)
    if not id_foto:
        raise HTTPException(status_code=500, detail="Error al agregar la foto guía")
    return {"id_foto_guia": id_foto, "success": True}


@app.get("/api/admin/campanas/{id_campana}/entregas")
def get_entregas_campana(id_campana: int, user=Depends(require_admin)):
    """Lista todas las entregas de una campaña (solo admin)."""
    return database.obtener_entregas_campana(id_campana)


@app.get("/api/admin/campanas/{id_campana}/resumen-tiendas")
def get_resumen_tiendas_campana(id_campana: int, user=Depends(require_admin)):
    """Lista TODAS las tiendas con su estatus en la campaña, incluidas las que no entregaron."""
    return database.obtener_resumen_campana_por_tienda(id_campana)


@app.get("/api/admin/campanas/entregas/{id_entrega}/detalle")
def get_detalle_entrega(id_entrega: int, user=Depends(require_admin)):
    """Obtiene el detalle con fotos en base64 de una entrega (solo admin)."""
    detalle = database.obtener_detalle_entrega(id_entrega)
    result = []
    for d in detalle:
        result.append({
            "Nombre_Foto": d["Nombre_Foto"],
            "Instrucciones": d["Instrucciones"],
            "Segmento": d["Segmento"],
            "Estatus_Auditoria": d["Estatus_Auditoria"],
            "Resultado_IA": d["Resultado_IA"],
            "foto_guia_b64": base64.b64encode(d["Foto_Guia"]).decode() if d.get("Foto_Guia") else None,
            "foto_tienda_b64": base64.b64encode(d["Foto_Tienda"]).decode() if d.get("Foto_Tienda") else None,
        })
    return result


@app.put("/api/admin/campanas/entregas/{id_entrega}/visto-bueno")
def dar_visto_bueno(
    id_entrega: int,
    user=Depends(require_admin),
):
    """Otorga el visto bueno zonal a una entrega (solo admin)."""
    row = database.dar_visto_bueno(id_entrega)
    if not row:
        raise HTTPException(status_code=404, detail="Entrega no encontrada")
    # Notificar al gerente
    database.crear_notificacion(
        row["ID_Usuario"],
        "👑 Visto Bueno Otorgado",
        f"Tu entrega de campaña '{row.get('campana_nombre', '')}' ha recibido el visto bueno final.",
        "campana",
    )
    return {"success": True}


@app.post("/api/admin/campanas/depurar-fotos")
def depurar_fotos(user=Depends(require_admin)):
    """Libera espacio eliminando binarios de fotos con más de 3 meses (solo admin)."""
    count = database.depurar_fotos_antiguas()
    return {"success": True, "fotos_depuradas": count}


@app.get("/api/admin/gemini-status")
def gemini_status(user=Depends(require_admin)):
    """Verifica si la API Key de Gemini está configurada."""
    return {"configurada": gemini_vision.verificar_api_key()}


# =========================================
# ENDPOINTS — PRESUPUESTO
# =========================================

class PresupuestoMensualRequest(BaseModel):
    anio: int
    mes: int
    meta: float
    venta_real: Optional[float] = 0.0


@app.get("/api/presupuesto/mensual")
def get_presupuesto_mensual(
    anio: int,
    mes: int,
    user=Depends(get_current_user),
):
    """Retorna el presupuesto mensual de la tienda del usuario autenticado."""
    tienda = user.get("tienda", "")
    if not tienda:
        return {"Meta_Mensual": 0, "Venta_Real": 0}
    return database.obtener_presupuesto_mensual(tienda, anio, mes) or {"Meta_Mensual": 0, "Venta_Real": 0}


@app.put("/api/presupuesto/mensual")
def upsert_presupuesto(
    req: PresupuestoMensualRequest,
    user=Depends(get_current_user),
):
    """Crea o actualiza el presupuesto mensual de la tienda del usuario."""
    tienda = user.get("tienda", "")
    if not tienda:
        raise HTTPException(status_code=400, detail="No tienes una tienda asignada")
    database.upsert_presupuesto_mensual(tienda, req.anio, req.mes, req.meta, req.venta_real)
    return {"success": True}


# =========================================
# PRESUPUESTO OPERATIVO (esquema completo)
# =========================================

class MetasPresupuestoRequest(BaseModel):
    tienda: Optional[str] = None
    anio: int
    mes: int
    meta_venta: float = 0.0
    meta_piezas: int = 0


class VentaDiariaRequest(BaseModel):
    tienda: Optional[str] = None
    fecha: str   # 'YYYY-MM-DD'
    venta_con_iva: float = 0.0
    piezas: int = 0


@app.get("/api/presupuesto/metas")
def get_metas_presupuesto(
    anio: int,
    mes: int,
    tienda: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Retorna las metas de venta y piezas de una tienda para un mes/año."""
    t = tienda if (tienda and user.get("rol") == "Admin") else user.get("tienda", "")
    if not t:
        return {"Meta_Venta": 0.0, "Meta_Piezas": 0}
    return database.obtener_metas_presupuesto(t, anio, mes)


@app.put("/api/presupuesto/metas")
def put_metas_presupuesto(
    req: MetasPresupuestoRequest,
    user=Depends(get_current_user),
):
    """Crea o actualiza las metas mensuales de venta y piezas."""
    t = req.tienda if (req.tienda and user.get("rol") == "Admin") else user.get("tienda", "")
    if not t:
        raise HTTPException(status_code=400, detail="No tienes una tienda asignada")
    ok = database.upsert_metas_presupuesto(t, req.anio, req.mes, req.meta_venta, req.meta_piezas)
    if not ok:
        raise HTTPException(status_code=500, detail="Error al guardar metas")
    return {"success": True}


@app.get("/api/presupuesto/ventas-diarias")
def get_ventas_diarias(
    anio: int,
    mes: int,
    tienda: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Retorna las ventas diarias de una tienda en un mes/año."""
    t = tienda if (tienda and user.get("rol") == "Admin") else user.get("tienda", "")
    if not t:
        return []
    return database.obtener_ventas_diarias(t, mes, anio)


@app.put("/api/presupuesto/venta-diaria")
def put_venta_diaria(
    req: VentaDiariaRequest,
    user=Depends(get_current_user),
):
    """Guarda o actualiza la venta de un día específico."""
    t = req.tienda if (req.tienda and user.get("rol") == "Admin") else user.get("tienda", "")
    if not t:
        raise HTTPException(status_code=400, detail="No tienes una tienda asignada")
    venta_sin_iva = req.venta_con_iva / 1.16
    ok = database.upsert_venta_diaria(t, req.fecha, req.venta_con_iva, venta_sin_iva, req.piezas)
    if not ok:
        raise HTTPException(status_code=500, detail="Error al guardar venta diaria")
    return {"success": True}


@app.get("/api/presupuesto/meses-logrados")
def get_meses_logrados(
    anio: int,
    tienda: Optional[str] = None,
    user=Depends(get_current_user),
):
    """Retorna el resumen anual de meses logrados para una tienda."""
    t = tienda if (tienda and user.get("rol") == "Admin") else user.get("tienda", "")
    if not t:
        return []
    return database.obtener_meses_logrados(t, anio)


@app.get("/api/presupuesto/tiendas")
def get_tiendas_con_zona(user=Depends(get_current_user)):
    """Retorna la lista de tiendas con su zona (solo Admin)."""
    if user.get("rol") != "Admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    return database.obtener_tiendas_con_zona()


# =========================================
# PENDIENTES — RESOLVER
# =========================================

@app.put("/api/pendientes/{id_pendiente}/resolver")
def resolver_pendiente_endpoint(
    id_pendiente: int,
    user=Depends(get_current_user),
):
    """Marca una pregunta pendiente como resuelta (solo Admin)."""
    if user.get("rol") != "Admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    ok = database.resolver_pendiente(id_pendiente)
    if not ok:
        raise HTTPException(status_code=404, detail="Pendiente no encontrado")
    return {"success": True}
