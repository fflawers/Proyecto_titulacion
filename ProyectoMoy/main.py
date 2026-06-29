import flet as ft
import flet_video as fv
import mysql.connector
import requests
import os
import base64
import fitz
import tempfile
import re
import json
import threading
from tkinter import Tk, filedialog
import openpyxl

# =========================================
# CONFIGURACION
# =========================================

BASE_PATH = os.path.dirname(__file__)
ASSETS_PATH = os.path.join(BASE_PATH, "assets")

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '../backend/.env'))

# Valores por defecto
GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
GROQ_MODEL = "llama-3.1-8b-instant"
URL_GROQ = "https://api.groq.com/openai/v1/chat/completions"
GEMINI_API_KEY = ""

DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "los4valtierra",
    "database": "sgh_portal"
}

# Cargar configuracion desde config.json si existe
try:
    config_path = os.path.join(BASE_PATH, "config.json")
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            config_data = json.load(f)
            models = config_data.get("models", [])
            if models:
                primer_modelo = models[0]
                GROQ_API_KEY = primer_modelo.get("apiKey", GROQ_API_KEY)
                GROQ_MODEL = primer_modelo.get("model", GROQ_MODEL)
                print(f"CONFIGURACIÓN CARGADA DESDE config.json: modelo '{GROQ_MODEL}'")
            GEMINI_API_KEY = config_data.get("gemini_api_key", os.getenv("GEMINI_API_KEY", ""))
            print(f"GEMINI API KEY CARGADA DESDE config.json: {'Configurada' if GEMINI_API_KEY else 'No configurada'}")
except Exception as e:
    print("ERROR CARGANDO config.json:", e)

# =========================================
# IMAGENES BASE64
# =========================================

def obtener_64(nombre):

    def buscar_archivo(nombre_buscar):
        target_base = os.path.splitext(nombre_buscar)[0]
        for base in [BASE_PATH, ASSETS_PATH]:
            ruta = os.path.join(base, nombre_buscar)
            if os.path.exists(ruta):
                return ruta
            if os.path.isdir(base):
                for nombre_archivo in os.listdir(base):
                    archivo_base, archivo_ext = os.path.splitext(nombre_archivo)
                    if archivo_base == target_base:
                        return os.path.join(base, nombre_archivo)
                    if archivo_base.startswith(target_base) and archivo_ext.lower() in [".jpeg", ".jpg", ".png"]:
                        return os.path.join(base, nombre_archivo)
        return None

    try:
        ruta = buscar_archivo(nombre)

        if ruta and os.path.exists(ruta):
            ext = os.path.splitext(ruta)[1].lower()
            mime = "image/png" if ext == ".png" else "image/jpeg"
            with open(ruta, "rb") as f:
                contenido = base64.b64encode(f.read()).decode("utf-8")
            return f"data:{mime};base64,{contenido}"

    except Exception as e:
        print("ERROR IMAGEN:", e)

    return None

# =========================================
# MYSQL
# =========================================

def conectar_db():

    try:
        return mysql.connector.connect(**DB_CONFIG)

    except Exception as e:
        print("ERROR MYSQL:", e)
        return None

# =========================================
# CONFIG HELPER & IMAGE OPTIMIZATION / AUDITING
# =========================================

def guardar_config_key(key_name, val):
    try:
        config_path = os.path.join(BASE_PATH, "config.json")
        config_data = {}
        if os.path.exists(config_path):
            with open(config_path, "r", encoding="utf-8") as f:
                try:
                    config_data = json.load(f)
                except Exception:
                    pass
        config_data[key_name] = val
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config_data, f, indent=4)
        return True
    except Exception as e:
        print("ERROR GUARDANDO CONFIG:", e)
        return False

def optimizar_imagen(imagen_bytes):
    try:
        from PIL import Image, ImageEnhance, ImageOps, ImageStat
        import io
        
        img = Image.open(io.BytesIO(imagen_bytes))
        if img.mode != 'RGB':
            img = img.convert('RGB')
            
        max_size = 1024
        if img.width > max_size or img.height > max_size:
            img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
            
        # Aplicar auto-contraste
        img = ImageOps.autocontrast(img, cutoff=2)
        
        # Evaluar brillo para auto-mejorar si está muy oscura
        stat = ImageStat.Stat(img)
        avg_brightness = sum(stat.mean) / 3
        if avg_brightness < 90:
            enhancer = ImageEnhance.Brightness(img)
            img = enhancer.enhance(1.3)
            
        out_io = io.BytesIO()
        img.save(out_io, format='JPEG', quality=85)
        return out_io.getvalue()
    except Exception as e:
        print("ERROR EN OPTIMIZAR_IMAGEN:", e)
        return imagen_bytes

def auditar_foto_con_gemini(guia_bytes, tienda_bytes, instrucciones):
    global GEMINI_API_KEY
    if not GEMINI_API_KEY:
        return "CORREGIR: La API Key de Gemini no está configurada. Vaya al panel de campañas y configúrela."
        
    import base64
    import json
    import requests
    
    try:
        guia_b64 = base64.b64encode(guia_bytes).decode('utf-8')
        tienda_b64 = base64.b64encode(tienda_bytes).decode('utf-8')
        
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={GEMINI_API_KEY}"
        headers = {"Content-Type": "application/json"}
        
        prompt = (
            "Actúa como un auditor visual de campañas de exhibición en Sunglass Hut. "
            "Se te proporcionan dos imágenes:\n"
            "1. La FOTO GUÍA (primera imagen): Es la referencia oficial de cómo debe quedar el montaje.\n"
            "2. La FOTO DE LA TIENDA (segunda imagen): Es el montaje real realizado por la tienda.\n\n"
            f"INSTRUCCIONES DE MONTAJE A VALIDAR:\n{instrucciones}\n\n"
            "Compara la foto de la tienda con la foto guía y con las instrucciones. "
            "Debes identificar si hay elementos faltantes, publicidad errónea, banners mal alineados, "
            "gafas en repisas incorrectas o diferencias significativas.\n"
            "Responde de forma clara, directa y en español.\n"
            "REGLA DE RESPUESTA CRÍTICA:\n"
            "- Si el montaje es correcto y cumple las instrucciones, empieza tu respuesta EXACTAMENTE con 'APROBADO'. Puedes añadir comentarios positivos después.\n"
            "- Si hay errores o diferencias que corregir, empieza tu respuesta EXACTAMENTE con 'CORREGIR' y proporciona una lista numerada con los puntos específicos que se deben solucionar en la tienda."
        )
        
        payload = {
            "contents": [
                {
                    "parts": [
                        {"text": prompt},
                        {
                            "inlineData": {
                                "mimeType": "image/jpeg",
                                "data": guia_b64
                            }
                        },
                        {
                            "inlineData": {
                                "mimeType": "image/jpeg",
                                "data": tienda_b64
                            }
                        }
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.2
            }
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=30)
        if response.status_code == 200:
            res_json = response.json()
            candidates = res_json.get("candidates", [])
            if candidates:
                parts = candidates[0].get("content", {}).get("parts", [])
                if parts:
                    return parts[0].get("text", "Error: No se recibió texto de la IA.")
            return f"Error: Respuesta de Gemini inesperada: {json.dumps(res_json)}"
        else:
            return f"Error en la API de Gemini: Código {response.status_code} - {response.text}"
    except Exception as e:
        return f"Error de conexión con Gemini: {str(e)}"

# =========================================
# NOTIFICACIONES BACKEND
# =========================================

def crear_notificacion(id_usuario, titulo, mensaje, tipo):
    """Inserta una notificación en la base de datos para un usuario específico."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute(
                "INSERT INTO notificaciones (ID_Usuario, Titulo, Mensaje, Tipo) VALUES (%s, %s, %s, %s)",
                (id_usuario, titulo, mensaje, tipo)
            )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR CREANDO NOTIFICACION:", e)
    return False

def crear_notificacion_a_rol(rol, titulo, mensaje, tipo):
    """Crea una notificación para todos los usuarios con un rol determinado."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute("SELECT ID_Usuario FROM usuarios WHERE Rol = %s", (rol,))
            users = cursor.fetchall()
            for u in users:
                cursor.execute(
                    "INSERT INTO notificaciones (ID_Usuario, Titulo, Mensaje, Tipo) VALUES (%s, %s, %s, %s)",
                    (u[0], titulo, mensaje, tipo)
                )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR CREANDO NOTIFICACION ROL:", e)
    return False

def crear_notificacion_a_zona(zona, titulo, mensaje, tipo):
    """Crea una notificación para todos los gerentes de una zona en específico."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute("SELECT ID_Usuario FROM usuarios WHERE Rol = 'Gerente' AND (Zona = %s OR %s = 'Todas')", (zona, zona))
            users = cursor.fetchall()
            for u in users:
                cursor.execute(
                    "INSERT INTO notificaciones (ID_Usuario, Titulo, Mensaje, Tipo) VALUES (%s, %s, %s, %s)",
                    (u[0], titulo, mensaje, tipo)
                )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR CREANDO NOTIFICACION ZONA:", e)
    return False

def cargar_notificaciones(id_usuario):
    """Retorna la lista de notificaciones recientes para un usuario (máximo 15)."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor(dictionary=True)
            cursor.execute(
                "SELECT ID_Notificacion, Titulo, Mensaje, Fecha_Hora, Leida, Tipo FROM notificaciones WHERE ID_Usuario = %s ORDER BY ID_Notificacion DESC LIMIT 15",
                (id_usuario,)
            )
            rows = cursor.fetchall()
            db.close()
            return rows
    except Exception as e:
        print("ERROR CARGANDO NOTIFICACIONES:", e)
    return []

def obtener_cantidad_notificaciones_sin_leer(id_usuario):
    """Retorna el conteo de notificaciones no leídas para un usuario."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute(
                "SELECT COUNT(*) FROM notificaciones WHERE ID_Usuario = %s AND Leida = 0",
                (id_usuario,)
            )
            count = cursor.fetchone()[0]
            db.close()
            return count
    except Exception as e:
        print("ERROR OBTENIENDO SIN LEER:", e)
    return 0

def marcar_notificaciones_leidas(id_usuario):
    """Marca todas las notificaciones de un usuario como leídas."""
    try:
        db = conectar_db()
        if db:
            cursor = db.cursor()
            cursor.execute(
                "UPDATE notificaciones SET Leida = 1 WHERE ID_Usuario = %s",
                (id_usuario,)
            )
            db.commit()
            db.close()
            return True
    except Exception as e:
        print("ERROR MARCANDO LEIDAS:", e)
    return False

# =========================================
# OBTENER PDF DESDE BD
# =========================================

def obtener_pdf_temporal(id_manual):
    """Obtiene un PDF de la BD y lo guarda en un archivo temporal. Retorna (ruta, nombre) o (None, None)."""
    try:
        db = conectar_db()
        if not db:
            return None, None
        cursor = db.cursor(dictionary=True)
        cursor.execute(
            "SELECT Nombre_Archivo, Archivo_PDF FROM manuales WHERE ID_Manual = %s",
            (id_manual,)
        )
        manual = cursor.fetchone()
        db.close()
        if manual:
            ruta = os.path.join(tempfile.gettempdir(), manual["Nombre_Archivo"])
            with open(ruta, "wb") as archivo:
                archivo.write(manual["Archivo_PDF"])
            return ruta, manual["Nombre_Archivo"]
        return None, None
    except Exception as e:
        print("ERROR OBTENER PDF:", e)
        return None, None

def obtener_ruta_escritorio():
    import winreg
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
        path, _ = winreg.QueryValueEx(key, "Desktop")
        return path
    except Exception:
        return os.path.join(os.path.expanduser("~"), "Desktop")

def visualizar_pdf(id_manual, page=None):
    """Abre el PDF en el visor del sistema sin guardar permanentemente."""
    try:
        ruta, nombre = obtener_pdf_temporal(id_manual)
        if ruta:
            os.startfile(ruta)
            if page:
                snack = ft.SnackBar(
                    ft.Text(f"Visualizando: {nombre}", color="#7CFC00", size=16, weight="bold"),
                    bgcolor="#111111", show_close_icon=True, open=True
                )
                page.overlay.append(snack)
                page.update()
        else:
            if page:
                snack = ft.SnackBar(
                    ft.Text("Error: Manual no encontrado.", color="#FF4500", size=16, weight="bold"),
                    bgcolor="#111111", show_close_icon=True, open=True
                )
                page.overlay.append(snack)
                page.update()
    except Exception as e:
        print("ERROR VISUALIZAR:", e)

def descargar_pdf_archivo(id_manual, page=None):
    """Descarga el PDF al escritorio del usuario de forma robusta."""
    try:
        ruta_temp, nombre = obtener_pdf_temporal(id_manual)
        if ruta_temp and nombre:
            escritorio = obtener_ruta_escritorio()
            ruta_destino = os.path.join(escritorio, nombre)
            import shutil
            shutil.copy2(ruta_temp, ruta_destino)
            if page:
                snack = ft.SnackBar(
                    ft.Text(f"PDF guardado en Escritorio: {nombre}", color="#7CFC00", size=16, weight="bold"),
                    bgcolor="#111111", show_close_icon=True, open=True
                )
                page.overlay.append(snack)
                page.update()
        else:
            if page:
                snack = ft.SnackBar(
                    ft.Text("Error: Manual no encontrado.", color="#FF4500", size=16, weight="bold"),
                    bgcolor="#111111", show_close_icon=True, open=True
                )
                page.overlay.append(snack)
                page.update()
    except Exception as e:
        print("ERROR DESCARGA:", e)
        if page:
            snack = ft.SnackBar(
                ft.Text(f"Error al descargar: {e}", color="#FF4500", size=16, weight="bold"),
                bgcolor="#111111", show_close_icon=True, open=True
            )
            page.overlay.append(snack)
            page.update()

# =========================================
# APP
# =========================================

def main(page: ft.Page):

    page.title = "LUXO"

    page.bgcolor = "#000000"

    page.theme_mode = "dark"

    page.window_width = 1100

    page.window_height = 850

    user_info = {
        "id": None,
        "nombre": "",
        "rol": ""
    }
    active_zone_filter = ["Todas"]
    es_admin = lambda: bool(user_info.get("rol") and user_info["rol"].strip().lower() in ("admin", "administrador"))
    selected_lang = ["es"]
        LOCALES = {
        "es": {
            "chat": "Asistente Chat",
            "history": "Mi Historial",
            "checklists": "Checklists 📋",
            "admin_panel": "Panel de Control",
            "logout": "Cerrar Sesión",
            "suggestion_title": "¿Qué te gustaría que tuviera LUXO?",
            "suggestion_hint": "Escribe tu idea aquí...",
            "send": "Enviar",
            "login_title": "SISTEMA LUXO",
            "user_label": "Usuario",
            "pass_label": "Contraseña",
            "login_btn": "INGRESAR",
            "progress": "Progreso",
            "of": "de",
            "completed": "completadas",
            "apertura": "Apertura 🌅",
            "cierre": "Cierre 🌌",
            "venta": "Venta Exitosa 💰",
            "add_task": "Agregar nueva tarea...",
            "no_tasks": "No hay tareas registradas en este checklist.",
            "task_deleted": "Tarea eliminada.",
            "task_added": "Nueva tarea agregada.",
            "refresh": "Recargar bitácoras",
            "edit_options": "Editar Opciones ⚙️",
            "lang_label": "Idioma 🌐",
            "checklist_title": "Bitácoras Operativas Sunglass Hut",
            "checklist_desc": "Completa las actividades obligatorias diarias de tu sucursal. El progreso se reinicia cada día.",
            "manuals_nav": "Manuales 📚",
            "manuals_title": "Manuales y Documentos 📚",
            "manuals_desc": "Consulta, visualiza o descarga los manuales operativos oficiales de Sunglass Hut para tu trabajo diario.",
            "manuals_db_title": "Manuales y Procedimientos Oficiales",
            "no_manuals": "No hay manuales disponibles en el sistema.",
            "version": "Versión",
            "view_pdf": "👁 Visualizar",
            "download_pdf": "⬇ Descargar",
            "pdf_delivered": "Aquí tienes el documento solicitado: {nombre_pdf}.",
            "pdf_not_found": "No encontré un PDF específico relacionado con tu solicitud. Por favor intenta ser más específico con el nombre del manual."
        },
        "en": {
            "chat": "Chat Assistant",
            "history": "My History",
            "checklists": "Checklists 📋",
            "admin_panel": "Admin Panel",
            "logout": "Log Out",
            "suggestion_title": "What would you like LUXO to have?",
            "suggestion_hint": "Write your idea here...",
            "send": "Send",
            "login_title": "LUXO SYSTEM",
            "user_label": "Username",
            "pass_label": "Password",
            "login_btn": "SIGN IN",
            "progress": "Progress",
            "of": "of",
            "completed": "completed",
            "apertura": "Opening 🌅",
            "cierre": "Closing 🌌",
            "venta": "Successful Sale 💰",
            "add_task": "Add new task...",
            "no_tasks": "No tasks registered in this checklist.",
            "task_deleted": "Task deleted.",
            "task_added": "New task added.",
            "refresh": "Refresh checklists",
            "edit_options": "Edit Options ⚙️",
            "lang_label": "Language 🌐",
            "checklist_title": "Sunglass Hut Operating Logs",
            "checklist_desc": "Complete your branch's mandatory daily activities. Progress resets every day.",
            "manuals_nav": "Manuals 📚",
            "manuals_title": "Manuals & Documents 📚",
            "manuals_desc": "Consult, view or download the official Sunglass Hut operational manuals for your daily work.",
            "manuals_db_title": "Official Manuals & Procedures",
            "no_manuals": "No manuals available in the system.",
            "version": "Version",
            "view_pdf": "👁 View PDF",
            "download_pdf": "⬇ Download",
            "pdf_delivered": "Here is the requested document: {nombre_pdf}.",
            "pdf_not_found": "I couldn't find a specific PDF related to your request. Please try to be more specific with the manual name."
        },
        "fr": {
            "chat": "Assistant Chat",
            "history": "Mon Historique",
            "checklists": "Listes de Contrôle 📋",
            "admin_panel": "Panneau de Contrôle",
            "logout": "Se Déconnecter",
            "suggestion_title": "Qu'aimeriez-vous que LUXO ait ?",
            "suggestion_hint": "Écrivez votre idée ici...",
            "send": "Envoyer",
            "login_title": "SYSTÈME LUXO",
            "user_label": "Nom d'utilisateur",
            "pass_label": "Mot de passe",
            "login_btn": "SE CONNECTER",
            "progress": "Progression",
            "of": "sur",
            "completed": "complétées",
            "apertura": "Ouverture 🌅",
            "cierre": "Fermeture 🌌",
            "venta": "Vente Réussie 💰",
            "add_task": "Ajouter une nouvelle tâche...",
            "no_tasks": "Aucune tâche enregistrée dans cette liste.",
            "task_deleted": "Tâche supprimée.",
            "task_added": "Nouvelle tâche ajoutée.",
            "refresh": "Actualiser les listes",
            "edit_options": "Modifier les Options ⚙️",
            "lang_label": "Langue 🌐",
            "checklist_title": "Registres Opérationnels de Sunglass Hut",
            "checklist_desc": "Effectuez les activités quotidiennes obligatoires de votre succursale. La progression est réinitialisée chaque jour.",
            "manuals_nav": "Manuels 📚",
            "manuals_title": "Manuels & Documents 📚",
            "manuals_desc": "Consultez, affichez ou téléchargez les manuels opérationnels officiels de Sunglass Hut pour votre travail quotidien.",
            "manuals_db_title": "Manuels et Procédures Officiels",
            "no_manuals": "Aucun manuel disponible dans le système.",
            "version": "Version",
            "view_pdf": "👁 Visualiser",
            "download_pdf": "⬇ Télécharger",
            "pdf_delivered": "Voici le document demandé : {nombre_pdf}.",
            "pdf_not_found": "Je n'ai pas trouvé de PDF spécifique lié à votre demande. Veuillez essayer d'être plus précis avec le nom du manuel."
        },
        "it": {
            "chat": "Assistente Chat",
            "history": "La Mia Cronologia",
            "checklists": "Liste di Controllo 📋",
            "admin_panel": "Pannello di Controllo",
            "logout": "Disconnettersi",
            "suggestion_title": "Cosa vorresti che avesse LUXO?",
            "suggestion_hint": "Scrivi la tua idea qui...",
            "send": "Invia",
            "login_title": "SISTEMA LUXO",
            "user_label": "Nome utente",
            "pass_label": "Password",
            "login_btn": "ACCEDI",
            "progress": "Progresso",
            "of": "di",
            "completed": "completate",
            "apertura": "Apertura 🌅",
            "cierre": "Chiusura 🌌",
            "venta": "Vendita con Successo 💰",
            "add_task": "Aggiungi nueva attività...",
            "no_tasks": "Nessuna attività registrata in questa lista.",
            "task_deleted": "Attività eliminata.",
            "task_added": "Nuova attività aggiunta.",
            "refresh": "Aggiorna liste",
            "edit_options": "Modifica Opzioni ⚙️",
            "lang_label": "Lingua 🌐",
            "checklist_title": "Registri Operativi Sunglass Hut",
            "checklist_desc": "Completa le attività quotidiane obbligatorie della tua filiale. Il progresso si azzera ogni giorno.",
            "manuals_nav": "Manuali 📚",
            "manuals_title": "Manuali e Documenti 📚",
            "manuals_desc": "Consulta, visualizza o scarica i manuali operativi ufficiali di Sunglass Hut per il tuo lavoro quotidiano.",
            "manuals_db_title": "Manuali e Procedure Ufficiali",
            "no_manuals": "Nessun manuale disponibile nel sistema.",
            "version": "Versione",
            "view_pdf": "👁 Visualizza",
            "download_pdf": "⬇ Scarica",
            "pdf_delivered": "Ecco il documento richiesto: {nombre_pdf}.",
            "pdf_not_found": "Non ho trovato un PDF specifico relativo alla tua richiesta. Per favore, prova ad essere più specifico con il nome del manuale."
        },
        "zh": {
            "chat": "聊天助手",
            "history": "我的历史记录",
            "checklists": "任务清单 📋",
            "admin_panel": "控制面板",
            "logout": "退出登入",
            "suggestion_title": "您希望 LUXO 增加什么功能？",
            "suggestion_hint": "在此写下您的想法...",
            "send": "发送",
            "login_title": "LUXO 系统",
            "user_label": "用户名",
            "pass_label": "密码",
            "login_btn": "登入",
            "progress": "进度",
            "of": "/",
            "completed": "已完成",
            "apertura": "开店准备 🌅",
            "cierre": "打烊关店 🌌",
            "venta": "成功销售 💰",
            "add_task": "添加新任务...",
            "no_tasks": "此清单暂无注册任务。",
            "task_deleted": "任务已删除。",
            "task_added": "新任务已添加。",
            "refresh": "刷新任务栏",
            "edit_options": "编辑选项 ⚙️",
            "lang_label": "语言 🌐",
            "checklist_title": "Sunglass Hut 运营日志",
            "checklist_desc": "完成您分店的每日强制活动。进度每天重置。",
            "manuals_nav": "手册 📚",
            "manuals_title": "手册与文件 📚",
            "manuals_desc": "咨询、查看或下载官方的 Sunglass Hut 运营手册以供日常工作使用。",
            "manuals_db_title": "官方手册与程序",
            "no_manuals": "系统内无可用手册。",
            "version": "版本",
            "view_pdf": "👁 预览",
            "download_pdf": "⬇ 下载",
            "pdf_delivered": "这是您请求的文件: {nombre_pdf}。",
            "pdf_not_found": "我找不到与您的请求相关的特定 PDF。请尝试提供更具体的手册名称。"
        }
    }
    
    def t(key):
        lang = selected_lang[0]
        return LOCALES.get(lang, LOCALES["es"]).get(key, LOCALES["es"].get(key, key))

    dashboard_tab_index = [0]

    # Asegurar que las tablas de checklist existan en la BD al iniciar la app
    try:
        db_init = conectar_db()
        if db_init:
            cursor_init = db_init.cursor()
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS plantillas_checklist (
                    ID_Plantilla INT AUTO_INCREMENT PRIMARY KEY,
                    Categoria INT,
                    Descripcion VARCHAR(255) NOT NULL,
                    Fecha_Creacion DATETIME DEFAULT CURRENT_TIMESTAMP
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS registro_checklist (
                    ID_Registro INT AUTO_INCREMENT PRIMARY KEY,
                    ID_Usuario INT NOT NULL,
                    ID_Plantilla INT NOT NULL,
                    Completado TINYINT(1) DEFAULT 0,
                    Fecha DATE,
                    Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE,
                    FOREIGN KEY (ID_Plantilla) REFERENCES plantillas_checklist(ID_Plantilla) ON DELETE CASCADE,
                    CONSTRAINT unique_user_task_date UNIQUE (ID_Usuario, ID_Plantilla, Fecha)
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS presupuesto_mensual (
                    ID_Presupuesto INT AUTO_INCREMENT PRIMARY KEY,
                    Tienda VARCHAR(100) NOT NULL,
                    Mes INT NOT NULL,
                    Anio INT NOT NULL,
                    Meta_Venta DECIMAL(15, 2) DEFAULT 0.00,
                    Meta_Piezas INT DEFAULT 0,
                    Fecha_Modificacion DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    CONSTRAINT unique_tienda_mes_anio UNIQUE (Tienda, Mes, Anio)
                );
            """)
            cursor_init.execute("""
                CREATE TABLE IF NOT EXISTS presupuesto_diario (
                    ID_Diario INT AUTO_INCREMENT PRIMARY KEY,
                    Tienda VARCHAR(100) NOT NULL,
                    Fecha DATE NOT NULL,
                    Venta_Con_IVA DECIMAL(15, 2) DEFAULT 0.00,
                    Venta_Sin_IVA DECIMAL(15, 2) DEFAULT 0.00,
                    Piezas INT DEFAULT 0,
                    Fecha_Modificacion DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    CONSTRAINT unique_tienda_fecha UNIQUE (Tienda, Fecha)
                );
            """)
            # Asegurar que la columna Abierto existe en la tabla manuales
            try:
                cursor_init.execute("SHOW COLUMNS FROM manuales LIKE 'Abierto'")
                if not cursor_init.fetchone():
                    cursor_init.execute("ALTER TABLE manuales ADD COLUMN Abierto TINYINT(1) DEFAULT 1")
            except Exception as e_col:
                print("ERROR AL AGREGAR COLUMNA Abierto:", e_col)
                
            db_init.commit()
            db_init.close()
    except Exception as e_init:
        print("ERROR INITIALIZING CHECKLIST TABLES:", e_init)

    def mostrar_snack(mensaje, color="#7CFC00"):
        snack = ft.SnackBar(
            ft.Text(mensaje, color=color, size=16, weight="bold"),
            bgcolor="#111111",
            duration=5000,
            show_close_icon=True,
            open=True
        )
        page.overlay.append(snack)
        page.update()

    def speak_text(text):
        if not text:
            return
        def run_speak():
            try:
                import subprocess
                clean_text = text.replace('"', '').replace("'", "").replace("\n", " ").strip()
                cmd = f"Add-Type -AssemblyName System.Speech; (New-Object System.Speech.Synthesis.SpeechSynthesizer).Speak('{clean_text}')"
                subprocess.Popen(["powershell", "-Command", cmd], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            except Exception as e:
                print("ERROR SPEAK:", e)
        import threading
        threading.Thread(target=run_speak, daemon=True).start()

    def clasificar_pregunta_faltante_async(pregunta_texto, id_pend):
        def run_classification():
            try:
                import requests
                headers = {
                    "Authorization": f"Bearer {GROQ_API_KEY}",
                    "Content-Type": "application/json"
                }
                system_msg = {
                    "role": "system",
                    "content": "Clasifica la siguiente pregunta de un usuario de Sunglass Hut en UNA de las siguientes categorías exactas: 'Impresoras', 'Políticas de Venta', 'Sistemas/Terminales', 'Manuales', 'Otros'. Responde ÚNICAMENTE con la palabra de la categoría (una sola palabra, sin comillas ni punto ni explicaciones)."
                }
                user_msg = {
                    "role": "user",
                    "content": pregunta_texto
                }
                payload = {
                    "model": GROQ_MODEL,
                    "messages": [system_msg, user_msg]
                }
                res = requests.post(URL_GROQ, headers=headers, json=payload, timeout=10)
                if res.status_code == 200:
                    data = res.json()
                    if "choices" in data and data["choices"]:
                        categoria = data["choices"][0]["message"]["content"].strip().replace("'", "").replace('"', '').replace(".", "")
                        valid_categories = ['Impresoras', 'Políticas de Venta', 'Sistemas/Terminales', 'Manuales', 'Otros']
                        matched_cat = "Otros"
                        for cat in valid_categories:
                            if cat.lower() in categoria.lower() or categoria.lower() in cat.lower():
                                matched_cat = cat
                                break
                        db_up = conectar_db()
                        if db_up:
                            cursor_up = db_up.cursor()
                            cursor_up.execute(
                                "UPDATE pendientes_actualizacion SET Categoria = %s WHERE ID_Pendiente = %s",
                                (matched_cat, id_pend)
                            )
                            db_up.commit()
                            db_up.close()
                            print(f"Pregunta ID {id_pend} clasificada como: {matched_cat}")
            except Exception as ex:
                print("ERROR EN CLASIFICACION ASYNC:", ex)
        import threading
        threading.Thread(target=run_classification, daemon=True).start()

    def on_broadcast_received(message):
        mostrar_snack(message, color="#7CFC00")
        page.update()
    if hasattr(page, "pubsub") and page.pubsub:
        page.pubsub.subscribe_topic("actualizaciones_luxo", on_broadcast_received)

    def registrar_feedback(id_conv, me_sirvio, comentario, fb_cont):
        if not id_conv:
            return
        try:
            db = conectar_db()
            if db:
                cursor = db.cursor()
                cursor.execute(
                    "UPDATE historial_conversaciones SET Me_Sirvio = %s, Comentario_Feedback = %s WHERE ID_Conversacion = %s",
                    (1 if me_sirvio else 0, comentario if comentario else None, id_conv)
                )
                db.commit()
                db.close()
                fb_cont.content = ft.Text("¡Gracias por calificar la respuesta!", color="#7CFC00", size=11, italic=True)
                page.update()
        except Exception as ex:
            print("ERROR REGISTRAR FEEDBACK:", ex)

    # =========================================
    # DIÁLOGOS DE ARCHIVO (SEGUROS CON TKINTER HILO SEPARADO)
    # =========================================

    def seleccionar_archivo_async(titulo, extensiones, callback):
        def thread_target():
            try:
                root = Tk()
                root.withdraw()
                root.attributes("-topmost", True)
                ruta = filedialog.askopenfilename(
                    title=titulo,
                    filetypes=extensiones
                )
                root.destroy()
                if ruta:
                    callback(ruta)
            except Exception as ex:
                print("ERROR TKINTER DIALOG:", ex)

        threading.Thread(target=thread_target, daemon=True).start()

    def procesar_cargar_pdf(ruta_pdf):
        mostrar_snack("Procesando e insertando PDF...", color="#D8B4FE")
        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return

            cursor = db.cursor()
            with open(ruta_pdf, "rb") as archivo:
                pdf_binario = archivo.read()

            pdf = fitz.open(ruta_pdf)
            texto_extraido = ""
            for pagina in pdf:
                texto_extraido += pagina.get_text()

            nombre_archivo = os.path.basename(ruta_pdf)

            sql = """
            INSERT INTO manuales
            (Titulo, Nombre_Archivo, Archivo_PDF, Contenido_Texto, Categoria, Version)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            valores = (nombre_archivo, nombre_archivo, pdf_binario, texto_extraido, "General", "1.0")
            cursor.execute(sql, valores)
            db.commit()
            db.close()

            crear_notificacion_a_rol("Gerente", "Nuevo Manual Disponible 📚", f"Se ha subido el manual: '{nombre_archivo}'", "manual")

            mostrar_snack(f"Manual '{nombre_archivo}' cargado exitosamente.")
            if hasattr(page, "pubsub") and page.pubsub:
                page.pubsub.send_all_on_topic("actualizaciones_luxo", f"📢 Nuevo manual disponible: '{nombre_archivo}'. LUXO ya se actualizó.")
        except Exception as ex:
            print("ERROR PDF:", ex)
            mostrar_snack(f"Error al cargar PDF: {ex}", color="#FF4500")
    def extraer_texto_excel(ruta_excel):
        """Extrae texto del Excel asociando cada celda con su encabezado de columna.
        Si data_only=True devuelve todo vacío (archivo con fórmulas sin caché),
        reintenta leyendo las fórmulas como texto."""

        def leer_filas(hoja):
            filas = []
            for fila in hoja.iter_rows(values_only=True):
                vals = [str(c).strip() if c is not None else "" for c in fila]
                if any(v for v in vals):
                    filas.append(vals)
            return filas

        texto = ""
        wb = openpyxl.load_workbook(ruta_excel, data_only=True)

        for nombre_hoja in wb.sheetnames:
            hoja = wb[nombre_hoja]
            todas_filas = leer_filas(hoja)

            # Si data_only devolvió todo vacío (fórmulas sin caché), reintenta sin data_only
            if not todas_filas:
                wb2 = openpyxl.load_workbook(ruta_excel, data_only=False)
                hoja2 = wb2[nombre_hoja]
                todas_filas = leer_filas(hoja2)
                wb2.close()
                print(f"[EXCEL] Hoja '{nombre_hoja}': usó modo sin data_only (fórmulas)")
            else:
                # Verificar si más del 80% de las celdas están vacías
                total = sum(len(f) for f in todas_filas)
                vacias = sum(1 for f in todas_filas for v in f if not v)
                if total > 0 and (vacias / total) > 0.8:
                    wb2 = openpyxl.load_workbook(ruta_excel, data_only=False)
                    hoja2 = wb2[nombre_hoja]
                    todas_filas_alt = leer_filas(hoja2)
                    wb2.close()
                    if todas_filas_alt:
                        todas_filas = todas_filas_alt
                        print(f"[EXCEL] Hoja '{nombre_hoja}': reemplazado por modo sin data_only (>80% vacías)")

            texto += f"\n{'='*50}\nHOJA: {nombre_hoja}\n{'='*50}\n"

            if not todas_filas:
                texto += "(Hoja vacía)\n"
                continue

            # Detectar encabezados: primera fila no vacía
            encabezados = []
            for i, val in enumerate(todas_filas[0]):
                encabezados.append(val if val else f"Columna_{i+1}")

            texto += f"COLUMNAS: {' | '.join(encabezados)}\n\n"

            # Cada fila de datos asociada con su encabezado
            for num_fila, fila in enumerate(todas_filas[1:], start=2):
                pares = []
                for i, valor in enumerate(fila):
                    if valor:
                        header = encabezados[i] if i < len(encabezados) else f"Columna_{i+1}"
                        pares.append(f"{header}: {valor}")
                if pares:
                    texto += f"FILA {num_fila}: " + " | ".join(pares) + "\n"

        wb.close()
        return texto

    def procesar_cargar_excel(ruta_excel):
        mostrar_snack("Procesando e insertando Excel...", color="#D8B4FE")
        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return

            cursor = db.cursor()

            with open(ruta_excel, "rb") as archivo:
                excel_binario = archivo.read()

            texto_extraido = extraer_texto_excel(ruta_excel)

            # Debug: imprimir lo extraído en consola
            print("\n===== TEXTO EXTRAÍDO DEL EXCEL =====")
            print(texto_extraido[:20000])
            print("=====================================\n")

            nombre_archivo = os.path.basename(ruta_excel)

            sql = """
            INSERT INTO manuales
            (Titulo, Nombre_Archivo, Archivo_PDF, Contenido_Texto, Categoria, Version)
            VALUES (%s, %s, %s, %s, %s, %s)
            """
            valores = (nombre_archivo, nombre_archivo, excel_binario, texto_extraido, "Excel", "1.0")
            cursor.execute(sql, valores)
            db.commit()
            db.close()

            crear_notificacion_a_rol("Gerente", "Nuevo Manual Excel Cargado 📊", f"Se ha subido el archivo: '{nombre_archivo}'", "manual")

            # Mostrar vista previa de lo que se extrajo
            def cerrar_preview(e):
                page.pop_dialog()

            lineas_preview = texto_extraido.strip().split("\n")[:40]
            preview_str = "\n".join(lineas_preview)
            if len(texto_extraido.strip().split("\n")) > 40:
                preview_str += "\n... (más filas guardadas)"

            dialog_preview = ft.AlertDialog(
                title=ft.Text(f"✅ Excel cargado: {nombre_archivo}", color="#7CFC00", weight="bold"),
                content=ft.Container(
                    content=ft.Column([
                        ft.Text(
                            "Así se leyó tu archivo (verifica que las columnas estén bien):",
                            color="#aaaaaa", size=12
                        ),
                        ft.TextField(
                            value=preview_str,
                            multiline=True,
                            read_only=True,
                            color="white",
                            bgcolor="#0a0a0a",
                            border_color="#444444",
                            min_lines=12,
                            max_lines=16,
                            text_style=ft.TextStyle(font_family="Courier New", size=11)
                        )
                    ], spacing=8),
                    width=650,
                    height=400,
                    padding=10
                ),
                actions=[ft.TextButton("Cerrar", on_click=cerrar_preview)],
                actions_alignment="end",
                bgcolor="#1a1a1a"
            )
            page.show_dialog(dialog_preview)
            if hasattr(page, "pubsub") and page.pubsub:
                page.pubsub.send_all_on_topic("actualizaciones_luxo", f"📢 Nuevo manual disponible: '{nombre_archivo}'. LUXO ya se actualizó.")

        except Exception as ex:
            print("ERROR EXCEL:", ex)
            mostrar_snack(f"Error al cargar Excel: {ex}", color="#FF4500")

    def mostrar_manuales_admin(chat_display=None):
        mostrar_snack("Listando manuales...")

        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return
            cursor = db.cursor(dictionary=True)
            cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version FROM manuales ORDER BY Nombre_Archivo")
            manuales = cursor.fetchall()
            db.close()

            if not manuales:
                mostrar_snack("No hay manuales cargados.")
                return

            if chat_display is not None:
                chat_display.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Text("Manuales cargados:", color="white", weight="bold"),
                            *[
                                ft.Text(f"{m.get('Nombre_Archivo', '')}  |  v{m.get('Version', '')}  |  {m.get('Titulo', '')}", color="white")
                                for m in manuales
                            ]
                        ], spacing=4),
                        bgcolor="#111111",
                        padding=10,
                        border_radius=10
                    )
                )
                page.update()
                return

            items = []
            for m in manuales:
                nombre = m.get("Nombre_Archivo") or ""
                version = m.get("Version") or ""
                titulo = m.get("Titulo") or ""
                items.append(
                    ft.Row([
                        ft.Text(nombre, expand=3, selectable=True),
                        ft.Text(f"v{version}", width=80),
                        ft.Text(titulo, expand=2)
                    ], alignment="center")
                )

            contenido = ft.Column(items, spacing=6)

            def cerrar_dialog(e):
                page.pop_dialog()

            dialog = ft.AlertDialog(
                title=ft.Text("Manuales cargados"),
                content=ft.Container(contenido, width=700, height=320),
                actions=[ft.TextButton("Cerrar", on_click=cerrar_dialog)],
                actions_alignment="end"
            )

            page.show_dialog(dialog)

        except Exception as e:
            print("ERROR LISTAR MANUALES:", e)
            mostrar_snack("Error listando manuales.", color="#FF4500")

    def borrar_manual_admin():
        mostrar_snack("Cargando lista de manuales a eliminar...")
        try:
            db = conectar_db()
            if not db:
                mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                return

            cursor = db.cursor(dictionary=True)
            cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version FROM manuales ORDER BY Nombre_Archivo")
            manuales = cursor.fetchall()
            db.close()

            if not manuales:
                mostrar_snack("No hay manuales para eliminar.")
                return

            def on_confirmar_borrado(id_manual, nombre_archivo, row_control):
                try:
                    db_del = conectar_db()
                    if not db_del:
                        mostrar_snack("Error de conexión", color="#FF4500")
                        return
                    cursor_del = db_del.cursor()

                    # Borrado en cascada seguro para evitar fallos de claves foráneas
                    cursor_del.execute("""
                        DELETE FROM pendientes_actualizacion 
                        WHERE ID_Conversacion IN (
                            SELECT ID_Conversacion FROM historial_conversaciones WHERE ID_Manual = %s
                        )
                    """, (id_manual,))

                    cursor_del.execute("DELETE FROM historial_conversaciones WHERE ID_Manual = %s", (id_manual,))
                    cursor_del.execute("DELETE FROM manuales WHERE ID_Manual = %s", (id_manual,))
                    db_del.commit()
                    db_del.close()

                    mostrar_snack(f"Manual '{nombre_archivo}' eliminado.")
                    lista_manuales_container.controls.remove(row_control)
                    page.update()

                except Exception as ex:
                    print("ERROR BORRADO MANUAL:", ex)
                    mostrar_snack(f"Error al eliminar: {ex}", color="#FF4500")

            lista_manuales_container = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS)

            for m in manuales:
                id_m = m["ID_Manual"]
                nombre = m.get("Nombre_Archivo") or ""
                version = m.get("Version") or ""

                fila = ft.Row(alignment="spaceBetween")

                info_text = ft.Text(
                    f"{nombre} (v{version})", 
                    color="white", 
                    weight="normal",
                    expand=True
                )

                btn_eliminar = ft.IconButton(
                    icon=ft.Icons.DELETE_FOREVER,
                    icon_color="#FF4500",
                    tooltip=f"Eliminar {nombre}",
                    on_click=lambda e, id_man=id_m, nom=nombre, f_ctrl=fila: on_confirmar_borrado(id_man, nom, f_ctrl)
                )

                fila.controls = [info_text, btn_eliminar]
                lista_manuales_container.controls.append(fila)

            def cerrar_dialog(e):
                page.pop_dialog()

            dialog = ft.AlertDialog(
                modal=True,
                title=ft.Text("Eliminar Manuales", color="#D8B4FE", weight="bold"),
                content=ft.Container(
                    content=lista_manuales_container,
                    width=500,
                    height=300,
                    padding=10
                ),
                actions=[
                    ft.TextButton("Cerrar", on_click=cerrar_dialog)
                ],
                actions_alignment="end"
            )

            page.show_dialog(dialog)

        except Exception as e:
            print("ERROR AL ABRIR DIALOGO BORRADO:", e)
            mostrar_snack("Error al cargar menú de eliminación.", color="#FF4500")

    img_avatar = obtener_64("avatar_luxo.png")

    img_fondo = None

    video_path = os.path.join(ASSETS_PATH, "luxo_avatar.mp4")
    if not os.path.exists(video_path):
        video_path = os.path.join(BASE_PATH, "luxo_avatar.mp4")

    # =====================================
    # CERRAR SESION
    # =====================================

    def cerrar_sesion():

        user_info["id"] = None
        user_info["nombre"] = ""
        user_info["rol"] = ""

        page.clean()

        page.add(login_ui)

        page.update()

    # =====================================
    # CHAT
    # =====================================

    def normalizar_texto(texto):
        import unicodedata
        if not texto:
            return ""
        texto = texto.lower()
        texto = ''.join(
            c for c in unicodedata.normalize('NFD', texto)
            if unicodedata.category(c) != 'Mn'
        )
        texto = re.sub(r'[^a-z0-9\s]', '', texto)
        return texto

    def cargar_chat():

        page.clean()

        chat_display = ft.ListView(
            expand=True,
            spacing=10,
            padding=20,
            auto_scroll=True
        )

        # Historial de conversación en memoria para enviar al LLM
        historial_sesion = []

        # Cargar historial de la base de datos de forma silenciosa en memoria para el contexto del LLM (evita problemas de rendimiento en la UI)
        try:
            db = conectar_db()
            if db:
                cursor = db.cursor(dictionary=True)
                cursor.execute("""
                    SELECT Pregunta_Usuario, Respuesta_IA 
                    FROM historial_conversaciones 
                    WHERE ID_Usuario = %s 
                    ORDER BY Fecha_Hora ASC 
                    LIMIT 10
                """, (user_info["id"],))
                historial = cursor.fetchall()
                db.close()
                for row in historial:
                    historial_sesion.append({"role": "user", "content": row["Pregunta_Usuario"]})
                    historial_sesion.append({"role": "assistant", "content": row["Respuesta_IA"]})
        except Exception as e:
            print("ERROR AL CARGAR HISTORIAL EN MEMORIA:", e)

        # Cargar imagen de avatar del usuario si existe
        img_usuario = obtener_64("istockphoto-468228782-612x612")

        # =================================
        # ENVIAR MENSAJE
        # =================================

        def enviar_mensaje(e):
            if not input_msg.value:
                return

            user_text = input_msg.value
            chat_display.controls.append(
                ft.Container(
                    content=ft.Row([
                        ft.Container(
                            content=ft.Image(src=img_usuario, width=35, height=35, fit=ft.controls.box.BoxFit.COVER) if img_usuario else ft.Icon(ft.Icons.PERSON, color="#00FFFF", size=20),
                            width=35,
                            height=35,
                            border_radius=17.5,
                            bgcolor="#333333",
                            clip_behavior=ft.ClipBehavior.HARD_EDGE,
                            alignment=ft.alignment.Alignment(0, 0),
                            border=ft.Border.all(1.5, "#D8B4FE"),
                        ),
                        ft.Text(f"{user_info['nombre']}: {user_text}", color="white", weight="bold", expand=True, selectable=True),
                    ], vertical_alignment="start", spacing=10),
                    bgcolor="#222222",
                    padding=10,
                    border_radius=10
                )
            )
            input_msg.value = ""
            page.update()

            try:
                db = conectar_db()
                if not db:
                    chat_display.controls.append(
                        ft.Text("ERROR: No se pudo conectar a la base de datos de manuales.", color="red")
                    )
                    page.update()
                    return
                cursor = db.cursor(dictionary=True)
                cursor.execute("SELECT * FROM manuales")
                manuales = cursor.fetchall()
                
                # Obtener historial de conversaciones calificadas como útiles (Me_Sirvio = 1) para aprendizaje dinámico
                cursor.execute("""
                    SELECT Pregunta_Usuario, Respuesta_IA 
                    FROM historial_conversaciones 
                    WHERE Me_Sirvio = 1
                """)
                chats_previos = cursor.fetchall()
                db.close()

                # Combinar consulta actual con la última pregunta del usuario (si hay) para dar contexto a la búsqueda de manuales
                contexto_busqueda = user_text
                ultimos_mensajes_usuario = [m["content"] for m in historial_sesion if m["role"] == "user"]
                if ultimos_mensajes_usuario:
                    contexto_busqueda = ultimos_mensajes_usuario[-1] + " " + user_text

                contexto_norm = normalizar_texto(contexto_busqueda)
                palabras_query = re.findall(r"\w+", contexto_norm)
                
                # Stopwords españolas exhaustivas
                stopwords = {
                    "de", "la", "el", "los", "las", "un", "una", "pdf", "manual", "documento", 
                    "archivo", "archivos", "tienes", "telefono", "tienda", "hola", "cual", "es", 
                    "como", "donde", "por", "para", "con", "que", "quiero", "saber", "me", "puedes", 
                    "dar", "darme", "a", "al", "en", "son", "cuales", "esta", "estas", "este", "estos", 
                    "del", "o", "u", "y", "e", "si", "no", "se", "lo", "te", "le", "les", "nos", "mi", 
                    "mis", "tu", "tus", "su", "sus", "ellos", "ellas", "nosotros", "usted", "ustedes", 
                    "mio", "tuyo", "suyo", "aqui", "alli", "alla", "todo", "todos", "toda", "todas", 
                    "uno", "unos", "otro", "otros", "otra", "otras", "hacer", "hace", "hacen", "haciendo", 
                    "ver", "vista", "puede", "pueden", "ser", "esta", "estan", "este", "esto", "del",
                    "sunglass", "hut", "luxottica", "quien", "quienes", "cuando", "como", "cual", "cuales",
                    "que", "porque", "donde", "realizar", "realizo", "realiza", "realizan", "realizado", 
                    "realizando", "paso", "pasos", "guia", "guias", "tutorial", "ayuda", "obtener", 
                    "descargar", "descarga", "bajar", "mostrar", "imprimir", "impresion", "sistema", "sistemas"
                }
                query_palabras = [w for w in palabras_query if w not in stopwords and len(w) >= 2]
                core_keywords = list(query_palabras)
                
                # Diccionario de sinónimos de Sunglass Hut
                SINONIMOS = {
                    "corte": ["caja", "cierre", "arqueo", "corte z"],
                    "cierre": ["corte", "caja", "cerrar", "finalizar"],
                    "cambio": ["devolucion", "garantia", "reemplazo", "ciao", "cancelar", "cancelacion"],
                    "devolucion": ["cambio", "garantia", "ciao", "cancelar", "cancelacion"],
                    "impresora": ["epson", "papel", "ticket", "tickets"],
                    "papel": ["rollo", "rollos", "impresora"],
                    "terminal": ["caja", "ciao", "pinpad", "santander"],
                    "robo": ["3r", "siniestro", "perdida"],
                    "contrasena": ["password", "clave", "lux", "ciao"],
                    "kpi": ["kpis", "indicador", "indicadores", "metas"],
                    "kpis": ["kpi", "indicador", "indicadores", "metas"],
                    "silla": ["ergonomia", "salud", "sillas"],
                    "vacaciones": ["descanso", "dias", "formato"],
                    "cancelacion": ["cancelar", "cancelo", "anular", "anulacion", "eliminar", "devolucion"],
                    "cancelar": ["cancelacion", "cancelo", "anular", "anulacion", "eliminar", "devolucion"],
                    "cancelo": ["cancelar", "cancelacion", "anular", "anulacion", "eliminar", "devolucion"],
                    "ticket": ["tickets", "comprobante", "transaccion", "venta", "cobro"],
                    "tickets": ["ticket", "comprobante", "transaccion", "venta", "cobro"],
                    "venta": ["cobro", "transaccion", "ticket", "vendedor", "comision", "comisiones"],
                    "vendedor": ["comision", "venta", "comisiones", "aclaracion"],
                    "comision": ["comisiones", "vendedor", "venta", "aclaracion"],
                    "comisiones": ["comision", "vendedor", "venta", "aclaracion"],
                    "aclaracion": ["comisiones", "comision", "vendedor", "error"],
                }
                # Expansión de la consulta con sinónimos
                sinonimos_expandidos = []
                for palabra in query_palabras:
                    if palabra in SINONIMOS:
                        for syn in SINONIMOS[palabra]:
                            if syn not in query_palabras and syn not in stopwords:
                                sinonimos_expandidos.append(syn)
                query_palabras.extend(sinonimos_expandidos)
                
                # Normalización inteligente y regla específica para "3R" o "3 R"
                if "3r" in palabras_query or ("3" in palabras_query and "r" in palabras_query):
                    if "3r" not in query_palabras:
                        query_palabras.append("3r")
                    if "robos" not in query_palabras:
                        query_palabras.append("robos")

                descarga_keywords = [
                    "descargar pdf", "descarga el manual", "descargar manual", 
                    "bajar pdf", "bajar el manual", "descargar archivo",
                    "pasa el pdf", "pasar el pdf", "dame el pdf", "quiero el pdf",
                    "mandame el pdf", "enviar pdf", "imprimir pdf", "imprimir manual"
                ]
                lista_keywords = ["listar manuales", "lista de manuales", "mostrar manuales disponibles", "que manuales tienes", "manuales cargados", "cuales son los manuales", "listar los pdf"]
                ask_for_pdf = any(phrase in user_text.lower() for phrase in descarga_keywords) or (
                    ("pdf" in user_text.lower() or "manual" in user_text.lower()) and 
                    any(p in user_text.lower() for p in ["dame", "quiero", "pasa", "descargar", "bajar", "ver", "mostrar", "imprimir", "obtener"])
                )
                ask_for_list = any(phrase in user_text.lower() for phrase in lista_keywords)

                # Función para calcular relevancia de cada manual
                def score_manual(meta):
                    import difflib
                    nombre_archivo_norm = normalizar_texto(meta["nombre"])
                    texto_completo_norm = normalizar_texto(meta["texto"])
                    score = 0
                    
                    if nombre_archivo_norm in contexto_norm:
                        score += 150
                    
                    nombre_palabras = [w for w in re.findall(r"\w+", nombre_archivo_norm) if w not in stopwords]
                    for palabra_archivo in nombre_palabras:
                        for palabra_query in query_palabras:
                            if palabra_query == palabra_archivo:
                                score += 50
                                break
                            ratio = difflib.SequenceMatcher(None, palabra_query, palabra_archivo).ratio()
                            if ratio >= 0.75:
                                score += int(50 * ratio)
                                break
                            elif palabra_query in palabra_archivo or palabra_archivo in palabra_query:
                                score += 20
                                break
                            
                    texto_palabras = texto_completo_norm.split()
                    for palabra in query_palabras:
                        count = texto_palabras.count(palabra)
                        if count > 0:
                            score += 10
                            score += min(count, 10) * 3
                    return score

                manuales_con_score = []
                for manual in manuales:
                    texto_manual = manual.get("Contenido_Texto") or ""
                    nombre_archivo = manual.get("Nombre_Archivo") or ""
                    categoria = manual.get("Categoria") or "General"
                    
                    meta = {
                        "id": manual["ID_Manual"],
                        "nombre": nombre_archivo,
                        "texto": texto_manual,
                        "categoria": categoria,
                        "abierto": manual.get("Abierto") if manual.get("Abierto") is not None else 1
                    }
                    
                    score = score_manual(meta)
                    manuales_con_score.append((score, meta))

                manuales_con_score.sort(key=lambda x: x[0], reverse=True)
                manuales_seleccionados = [m for score, m in manuales_con_score if score >= 15]
                
                modo_sugerencia = False
                sugerencias_nombres = []

                if not manuales_seleccionados:
                    candidatos = [m for score, m in manuales_con_score if score >= 2]
                    if candidatos:
                        modo_sugerencia = True
                        sugerencias_nombres = [m["nombre"] for m in candidatos[:3]]

                # --- BUSCAR CASOS PREVIOS RESUELTOS CON ÉXITO ---
                def score_chat(chat_item):
                    import difflib
                    preg_norm = normalizar_texto(chat_item["Pregunta_Usuario"])
                    score = 0
                    if preg_norm in contexto_norm or contexto_norm in preg_norm:
                        score += 100
                    preg_palabras = [w for w in re.findall(r"\w+", preg_norm) if w not in stopwords]
                    for wp in preg_palabras:
                        for wq in query_palabras:
                            if wq == wp:
                                score += 40
                                break
                            ratio = difflib.SequenceMatcher(None, wq, wp).ratio()
                            if ratio >= 0.8:
                                score += int(40 * ratio)
                                break
                    return score

                chats_con_score = []
                for c_item in chats_previos:
                    sc = score_chat(c_item)
                    chats_con_score.append((sc, c_item))
                
                chats_con_score.sort(key=lambda x: x[0], reverse=True)
                chats_seleccionados = [c_i for sc_val, c_i in chats_con_score if sc_val >= 40]
                chats_seleccionados = chats_seleccionados[:2]
                
                casos_previos_texto = ""
                for idx, c_item in enumerate(chats_seleccionados, start=1):
                    preg_val = c_item.get('Pregunta_Usuario') or ""
                    resp_val = c_item.get('Respuesta_IA') or ""
                    preg_trunc = preg_val[:500]
                    resp_trunc = resp_val[:1500]
                    casos_previos_texto += f"""
CASO PREVIO {idx} (Resuelto con éxito anteriormente):
Pregunta del usuario: {preg_trunc}
Respuesta útil de LUXO: {resp_trunc}
=================================
"""

                manuales_seleccionados = manuales_seleccionados[:2]
                manuales_texto = ""
                id_manual = None
                nombre_pdf = ""
                es_pdf_abierto = True
                
                if manuales_seleccionados:
                    id_manual = manuales_seleccionados[0]["id"]
                    nombre_pdf = manuales_seleccionados[0]["nombre"]
                    es_pdf_abierto = manuales_seleccionados[0].get("abierto", 1) == 1

                # Definir función de división en bloques lógicos/párrafos
                def dividir_texto_en_bloques(texto):
                    if not texto:
                        return []
                    texto = texto.replace("\r\n", "\n")
                    parrafos = texto.split("\n\n")
                    
                    bloques = []
                    header_markers = ["que pasa", "me equivoque", "como ", "un cliente", "cualquier", "en el caso", "que hacer", "me marcaron", "family an", "todo funciona"]
                    
                    for parrafo in parrafos:
                        parrafo_strip = parrafo.strip()
                        if not parrafo_strip:
                            continue
                            
                        lineas = parrafo_strip.split("\n")
                        bloque_actual = []
                        
                        for linea in lineas:
                            linea_strip = linea.strip()
                            if not linea_strip:
                                continue
                                
                            es_inicio = False
                            if (linea_strip.endswith("?") or 
                                linea_strip.endswith("?.") or 
                                linea_strip.endswith("? ") or
                                any(linea_strip.lower().startswith(marker) for marker in ["que hago", "que hacer", "como ", "como hacer", "como realizar"])
                            ):
                                if len(linea_strip) > 15 and not linea_strip.startswith("-") and not linea_strip.startswith("*"):
                                    es_inicio = True
                            else:
                                linea_norm = normalizar_texto(linea_strip)
                                if any(linea_norm.startswith(marker) for marker in header_markers):
                                    if len(linea_strip) > 15 and not linea_strip.startswith("-") and not linea_strip.startswith("*"):
                                        es_inicio = True
                                        
                            if es_inicio and bloque_actual:
                                bloques.append("\n".join(bloque_actual))
                                bloque_actual = []
                            bloque_actual.append(linea)
                            
                        if bloque_actual:
                            bloques.append("\n".join(bloque_actual))
                            
                    bloques_finales = []
                    temp_bloque = ""
                    for b in bloques:
                        b_strip = b.strip()
                        if not b_strip:
                            continue
                        if temp_bloque:
                            temp_bloque += "\n" + b_strip
                        else:
                            temp_bloque = b_strip
                            
                        if len(temp_bloque) >= 120 or temp_bloque.endswith("?"):
                            bloques_finales.append(temp_bloque)
                            temp_bloque = ""
                            
                    if temp_bloque:
                        bloques_finales.append(temp_bloque)
                        
                    return bloques_finales

                bloques_candidatos = []
                for m in manuales_seleccionados:
                    texto = m["texto"]
                    if m["categoria"] == "Excel":
                        lineas = texto.split("\n")
                        lineas_filtradas = []
                        for idx, linea in enumerate(lineas):
                            linea_norm = normalizar_texto(linea)
                            if "columna" in linea_norm or "hoja" in linea_norm or "====" in linea:
                                lineas_filtradas.append(linea)
                                continue
                            if any(q in linea_norm for q in query_palabras):
                                lineas_filtradas.append(linea)
                                if idx + 1 < len(lineas):
                                    linea_sig = lineas[idx + 1]
                                    linea_sig_norm = normalizar_texto(linea_sig)
                                    if not any(marker in linea_sig_norm for marker in ["columna", "hoja"]):
                                        lineas_filtradas.append("   " + linea_sig)
                        texto_excel = "\n".join(lineas_filtradas[:50])
                        bloques_candidatos.append((100, m["nombre"], texto_excel))
                    else:
                        # Dividir manual de texto en bloques y calificarlos independientemente
                        bloques_manual = dividir_texto_en_bloques(texto)
                        for blk in bloques_manual:
                            blk_norm = normalizar_texto(blk)
                            blk_palabras = blk_norm.split()
                            score_b = 0
                            for kw in core_keywords:
                                count = blk_palabras.count(kw)
                                if count > 0:
                                    score_b += 50  # Puntuación base para palabra clave original
                                    score_b += min(count, 5) * 10  # Bono por frecuencia
                                else:
                                    # Solo buscar sinónimos si la palabra clave original no coincide en este bloque
                                    if kw in SINONIMOS:
                                        syn_matches = 0
                                        for syn in SINONIMOS[kw]:
                                            syn_count = blk_palabras.count(syn)
                                            if syn_count > 0:
                                                syn_matches += syn_count
                                        if syn_matches > 0:
                                            score_b += 15  # Puntuación base para sinónimo
                                            score_b += min(syn_matches, 3) * 5  # Bono por frecuencia de sinónimo
                            if score_b >= 20:  # Umbral mínimo absoluto para considerar un bloque
                                bloques_candidatos.append((score_b, m["nombre"], blk))

                # Ordenar todos los bloques y filtrar por umbral relativo (mínimo 60% del puntaje máximo)
                bloques_filtrados = []
                if bloques_candidatos:
                    bloques_candidatos.sort(key=lambda x: x[0], reverse=True)
                    max_score = bloques_candidatos[0][0]
                    # Solo incluimos bloques que tengan al menos el 60% del puntaje máximo
                    for score_val, doc_nombre, blk_texto in bloques_candidatos:
                        if score_val >= (max_score * 0.6):
                            bloques_filtrados.append((score_val, doc_nombre, blk_texto))
                
                bloques_por_doc = {}
                for score_val, doc_nombre, blk_texto in bloques_filtrados[:3]:
                    if doc_nombre not in bloques_por_doc:
                        bloques_por_doc[doc_nombre] = []
                    bloques_por_doc[doc_nombre].append(blk_texto)
                
                for doc_nombre, lista_blks in bloques_por_doc.items():
                    texto_completo_doc = "\n\n[...]\n\n".join(lista_blks)
                    manuales_texto += f"""
DOCUMENTO: {doc_nombre}
CONTENIDO:
{texto_completo_doc}
=================================
"""

                respuesta = ""

                if ask_for_list:
                    if manuales:
                        lista_nombres = ", ".join([m["Nombre_Archivo"] for m in manuales])
                        respuesta = f"Manuales disponibles en el sistema: {lista_nombres}"
                    else:
                        respuesta = "No hay manuales cargados en el sistema."
                    historial_sesion.append({"role": "user", "content": user_text})
                    historial_sesion.append({"role": "assistant", "content": respuesta})
                elif ask_for_pdf:
                    if manuales_con_score and manuales_con_score[0][0] >= 5:
                        mejor = manuales_con_score[0][1]
                        id_manual = mejor["id"]
                        nombre_pdf = mejor["nombre"]
                        respuesta = t("pdf_delivered").format(nombre_pdf=nombre_pdf)
                    else:
                        respuesta = t("pdf_not_found")
                    historial_sesion.append({"role": "user", "content": user_text})
                    historial_sesion.append({"role": "assistant", "content": respuesta})
                else:
                    headers = {
                        "Authorization": f"Bearer {GROQ_API_KEY}",
                        "Content-Type": "application/json"
                    }

                    lang_names = {
                        "es": "Spanish (Español)",
                        "en": "English (Inglés)",
                        "fr": "French (Français)",
                        "it": "Italian (Italiano)",
                        "zh": "Chinese Mandarin (中文/普通话)"
                    }
                    target_lang_name = lang_names.get(selected_lang[0], "Spanish (Español)")

                    if modo_sugerencia:
                        mensaje_sistema = {
                            "role": "system",
                            "content": f"""Eres LUXO, asistente operativo de Sunglass Hut.
                            
El usuario realizó una consulta, pero no logramos identificar con certeza un manual relacionado en nuestro sistema.
Sin embargo, se encontraron estos documentos candidatos:
{", ".join(sugerencias_nombres)}

Por favor, responde de manera muy natural y amable. Pregúntale si su duda se refiere a alguno de estos temas. Usa la frase "¿Quizás quisiste decir...?" y presenta las opciones de forma clara (ej. viñetas con los nombres de los manuales sugeridos) para que el usuario pueda elegir o reformular su pregunta.
No inventes información sobre el contenido de los manuales.

INSTRUCCIÓN DE IDIOMA IMPORTANTE:
Debes responder al usuario strictly in the language: {target_lang_name}. If the provided manuals or user query are in another language (like Spanish), please translate them on the fly and reply in {target_lang_name} naturally and fluently.
"""
                        }
                    else:
                        mensaje_sistema = {
                            "role": "system",
                            "content": f"""Eres LUXO, asistente operativo inteligente de Sunglass Hut.

INSTRUCCIÓN CRÍTICA DE SEGURIDAD (CERO ALUCINACIONES Y CERO CONSEJOS EXTERNOS):
1. Debes responder basándote ÚNICAMENTE y de forma estricta en la información provista en la sección "DOCUMENTOS / MANUALES". Está terminantemente PROHIBIDO usar tu conocimiento general o inventar procedimientos.
2. No agregues pasos, consejos, recomendaciones, oficinas a visitar, personas a contactar (como jefe zonal o área de control) ni soluciones alternativas que no estén escritas textualmente en el manual provisto. Si el manual no lo dice, NO debes sugerirlo bajo ninguna circunstancia.
3. Si la respuesta no está detallada textualmente en los manuales proporcionados, debes responder únicamente: "Por el momento no cuento con esta información."
4. Las conversaciones generales, saludos y despedidas (como "hola", "gracias", "¿cómo estás?", "¿qué puedes hacer?") puedes responderlos de forma natural, amable y profesional sin necesidad de consultar documentos.
5. Al responder sobre advertencias o prohibiciones operativas (por ejemplo, no realizar devoluciones o cambios bajo ciertas condiciones), cita o parafrasea de forma muy cercana las advertencias literales que están escritas en el manual (ejemplo: "ES IMPORTANTE QUE NO REALICES LA DEVOLUCIÓN" o "PUEDES PEDIRLE A TU GERENTE QUE REALICE UNA ACLARACIÓN DE COMISIONES").

DOCUMENTOS / MANUALES PROPORCIONADOS:
{manuales_texto if manuales_texto.strip() else "(No hay documentos relacionados para esta consulta)"}

CASOS PREVIOS RESUELTOS CON ÉXITO (EXPERIENCIA DE LUXO):
{casos_previos_texto if casos_previos_texto.strip() else "(No hay casos previos similares registrados)"}

INSTRUCCIÓN DE IDIOMA IMPORTANTE:
Debes responder al usuario strictly in the language: {target_lang_name}. If the provided manuals or user query are in another language (like Spanish), please translate them on the fly and reply in {target_lang_name} naturally and fluently.
"""
                        }

                    historial_sesion.append({"role": "user", "content": user_text})
                    mensajes_api = [mensaje_sistema]
                    mensajes_api.extend(historial_sesion[-10:])

                    payload = {
                        "model": GROQ_MODEL,
                        "messages": mensajes_api
                    }

                    res = requests.post(URL_GROQ, headers=headers, json=payload)
                    if res.status_code == 200:
                        try:
                            data = res.json()
                            if "choices" in data and data["choices"]:
                                respuesta = data["choices"][0]["message"]["content"]
                            else:
                                respuesta = "Ocurrió un error consultando la IA."
                        except Exception as e:
                            print("AI PARSE ERROR:", e)
                            respuesta = "Ocurrió un error consultando la IA."
                    else:
                        print("AI CONNECTION ERROR:", res.status_code, res.text)
                        respuesta = f"Error de conexión con la IA ({res.status_code})."

                    historial_sesion.append({"role": "assistant", "content": respuesta})

                # --- REGISTRAR EN BASE DE DATOS ANTES DE RENDERIZAR PARA OBTENER EL ID ---
                id_conversacion = None
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        sql_historial = """
                        INSERT INTO historial_conversaciones
                        (
                            ID_Usuario,
                            ID_Manual,
                            Pregunta_Usuario,
                            Respuesta_IA,
                            Fecha_Hora,
                            Fue_Respondida_Con_Manual
                        )
                        VALUES (%s, %s, %s, %s, NOW(), %s)
                        """
                        cursor.execute(
                            sql_historial,
                            (
                                user_info["id"],
                                id_manual,
                                user_text,
                                respuesta,
                                1 if id_manual else 0,
                            )
                        )
                        db.commit()
                        id_conversacion = cursor.lastrowid

                        if "Por el momento no cuento con esta información" in respuesta:
                            sql_pendiente = """
                            INSERT INTO pendientes_actualizacion
                            (
                                ID_Conversacion,
                                Pregunta_Faltante
                            )
                            VALUES (%s, %s)
                            """
                            cursor.execute(sql_pendiente, (id_conversacion, user_text))
                            db.commit()
                            id_pendiente = cursor.lastrowid
                            clasificar_pregunta_faltante_async(user_text, id_pendiente)
                        db.close()
                except Exception as ex:
                    print("ERROR AL REGISTRAR CONVERSACIÓN:", ex)

                # --- CREAR CONTENEDOR DE FEEDBACK SI SE OBTUVO EL ID DE CONVERSACIÓN ---
                feedback_container = ft.Container(alignment=ft.alignment.Alignment(-1, 0))
                if id_conversacion:
                    def on_thumbs_up(ev, conv_id=id_conversacion, fb_cont=feedback_container):
                        comment_input = ft.TextField(
                            label="Cuéntanos qué te sirvió o agrega un comentario (opcional)",
                            multiline=True,
                            min_lines=2,
                            max_lines=4,
                            border_color="#9D50BB",
                            color="white"
                        )
                        
                        def submit_comment(e):
                            comentario = comment_input.value.strip()
                            registrar_feedback(conv_id, True, comentario, fb_cont)
                            page.pop_dialog()
                            page.update()
                            
                        def cancel_comment(e):
                            registrar_feedback(conv_id, True, "", fb_cont)
                            page.pop_dialog()
                            page.update()

                        dialog_comment = ft.AlertDialog(
                            title=ft.Text("Calificar respuesta como Útil", color="#7CFC00", weight="bold", size=16),
                            content=ft.Container(
                                content=comment_input,
                                width=450,
                                height=120
                            ),
                            actions=[
                                ft.TextButton("Calificar Directo", on_click=cancel_comment),
                                ft.ElevatedButton("Enviar y Calificar", on_click=submit_comment, bgcolor="#7CFC00", color="white")
                            ],
                            actions_alignment="end",
                            bgcolor="#111111"
                        )
                        page.show_dialog(dialog_comment)
                        
                    def on_thumbs_down(ev, conv_id=id_conversacion, fb_cont=feedback_container):
                        comment_input = ft.TextField(
                            label="Cuéntanos por qué no te sirvió (opcional)",
                            multiline=True,
                            min_lines=2,
                            max_lines=4,
                            border_color="#9D50BB",
                            color="white"
                        )
                        
                        def submit_comment(e):
                            comentario = comment_input.value.strip()
                            registrar_feedback(conv_id, False, comentario, fb_cont)
                            page.pop_dialog()
                            page.update()
                            
                        def cancel_comment(e):
                            registrar_feedback(conv_id, False, "", fb_cont)
                            page.pop_dialog()
                            page.update()

                        def crear_ticket_click(e):
                            comentario = comment_input.value.strip()
                            if not comentario:
                                comentario = "El bot no respondió correctamente a la pregunta: " + user_text
                            registrar_feedback(conv_id, False, comentario, fb_cont)
                            try:
                                db_t = conectar_db()
                                if db_t:
                                    cursor_t = db_t.cursor()
                                    cursor_t.execute("""
                                        INSERT INTO tickets_soporte (ID_Usuario, Detalle_Problema)
                                        VALUES (%s, %s)
                                    """, (user_info["id"], f"Pregunta: {user_text}\nRespuesta Luxo: {respuesta}\nComentario: {comentario}"))
                                    db_t.commit()
                                    db_t.close()
                                    mostrar_snack("¡Ticket de Soporte creado con éxito!", color="#7CFC00")
                            except Exception as ex:
                                print("ERROR AL CREAR TICKET:", ex)
                                mostrar_snack("Error al registrar el ticket.", color="red")
                            page.pop_dialog()
                            page.update()

                        dialog_comment = ft.AlertDialog(
                            title=ft.Text("Calificar respuesta como No Útil", color="#FF4500", weight="bold", size=16),
                            content=ft.Container(
                                content=comment_input,
                                width=450,
                                height=120
                            ),
                            actions=[
                                ft.TextButton("Calificar Directo", on_click=cancel_comment),
                                ft.ElevatedButton("Crear Ticket 🎫", on_click=crear_ticket_click, bgcolor="#9D50BB", color="white"),
                                ft.ElevatedButton("Enviar y Calificar", on_click=submit_comment, bgcolor="#FF4500", color="white")
                            ],
                            actions_alignment="end",
                            bgcolor="#111111"
                        )
                        page.show_dialog(dialog_comment)

                    feedback_buttons = ft.Row([
                        ft.Text("¿Te sirvió la respuesta?", color="#aaaaaa", size=11),
                        ft.IconButton(
                            icon=ft.Icons.THUMB_UP_OUTLINED,
                            icon_size=15,
                            icon_color="#7CFC00",
                            tooltip="Sí, fue útil",
                            on_click=on_thumbs_up
                        ),
                        ft.IconButton(
                            icon=ft.Icons.THUMB_DOWN_OUTLINED,
                            icon_size=15,
                            icon_color="#FF4500",
                            tooltip="No fue útil",
                            on_click=on_thumbs_down
                        ),
                        ft.IconButton(
                            icon=ft.Icons.VOLUME_UP_ROUNDED,
                            icon_size=15,
                            icon_color="#00FFFF",
                            tooltip="Escuchar respuesta",
                            on_click=lambda e: speak_text(respuesta)
                        )
                    ], spacing=5, alignment="start", vertical_alignment="center")
                    
                    feedback_container.content = feedback_buttons

                # Renderizar mensaje de Luxo
                # --- BUSCAR ACTIVOS VISUALES INTERACTIVOS (Assets) ---
                keyword_assets = {
                    "impresora": "ayuda_impresora.png",
                    "epson": "ayuda_impresora.png",
                    "papel": "ayuda_impresora.png",
                    "terminal": "guia_terminal.png",
                    "caja": "guia_terminal.png",
                    "devolucion": "guia_devolucion.png",
                    "devoluciones": "guia_devolucion.png",
                    "politica de devolucion": "guia_devolucion.png",
                    "politica de devoluciones": "guia_devolucion.png",
                }
                
                matched_asset = None
                clean_resp_lower = respuesta.lower()
                # No mostrar imagenes de guia si la respuesta indica no disponibilidad de la informacion o si se adjunta un PDF
                if "no cuento con" not in clean_resp_lower and id_manual is None:
                    for kw, asset_filename in keyword_assets.items():
                        if kw in clean_resp_lower:
                            asset_path = os.path.join("assets", asset_filename)
                            if os.path.exists(asset_path):
                                matched_asset = asset_path
                                break

                luxo_column_controls = [
                    ft.Row([
                        ft.Container(
                            content=fv.Video(
                                playlist=[fv.VideoMedia(video_path)],
                                playlist_mode=fv.PlaylistMode.LOOP,
                                autoplay=True,
                                muted=True,
                                controls=None,
                                width=35,
                                height=35,
                            ) if os.path.exists(video_path) else (
                                ft.Image(src=img_avatar, width=35, height=35) if img_avatar else ft.Text("L", color="white", weight="bold")
                            ),
                            width=35,
                            height=35,
                            border_radius=17.5,
                            clip_behavior=ft.ClipBehavior.HARD_EDGE,
                            border=ft.Border.all(1.5, "#00FFFF"),
                        ),
                        ft.Text(f"LUXO: {respuesta}", color="white", weight="bold", expand=True, selectable=True),
                    ], vertical_alignment="start", spacing=10),
                    ft.Row([
                        ft.Container(width=45),
                        feedback_container
                    ])
                ]

                if matched_asset:
                    luxo_column_controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Text("💡 Guía Visual Relacionada:", color="#00FFFF", size=12, weight="bold"),
                                ft.Image(src=matched_asset, width=400, height=220, border_radius=10)
                            ], spacing=5),
                            bgcolor="#1e1e1e",
                            padding=10,
                            border_radius=8,
                            border=ft.Border.all(1, "#333333"),
                            margin=ft.Margin(left=45, top=5, right=0, bottom=0)
                        )
                    )

                chat_display.controls.append(
                    ft.Container(
                        content=ft.Column(luxo_column_controls, spacing=5),
                        bgcolor="#111111",
                        padding=10,
                        border_radius=10
                    )
                )

                if id_manual and es_pdf_abierto:
                    chat_display.controls.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.Icons.PICTURE_AS_PDF, color="#D8B4FE"),
                                ft.Text(f"{nombre_pdf}", color="white", weight="bold", expand=True),
                                ft.ElevatedButton(
                                    t("view_pdf"),
                                    on_click=lambda e, idm=id_manual: visualizar_pdf(idm, page),
                                    bgcolor="#6E48AA",
                                    color="white"
                                ),
                                ft.ElevatedButton(
                                    t("download_pdf"),
                                    on_click=lambda e, idm=id_manual: descargar_pdf_archivo(idm, page),
                                    bgcolor="#444444",
                                    color="white"
                                ),
                            ], alignment="center", spacing=10),
                            bgcolor="#1a1a2e",
                            padding=10,
                            border_radius=10
                        )
                    )

            except Exception as e:
                chat_display.controls.append(
                    ft.Text(f"ERROR: {e}", color="red")
                )

            page.update()

        # =================================
        # INPUT
        # =================================

        input_msg = ft.TextField(
            hint_text="Escribe tu consulta...",
            expand=True,
            on_submit=enviar_mensaje,
            border_color="#9D50BB",
            color="white",
            bgcolor="#111111"
        )

        # =================================
        # VISTAS DEL PANEL DINÁMICO
        # =================================

        def mostrar_instrucciones_dictado(e):
            def cerrar_dictado_dialog(ev):
                page.pop_dialog()
            dialog = ft.AlertDialog(
                title=ft.Text("🎙️ Dictado por Voz (Speech-to-Text)", color="#D8B4FE", weight="bold"),
                content=ft.Column([
                    ft.Text("Para escribir usando tu voz, puedes usar las funciones de dictado nativas de tu dispositivo:", color="white", size=14),
                    ft.Divider(height=10, color="#333333"),
                    ft.Row([
                        ft.Icon(ft.Icons.PHONE_ANDROID, color="#7CFC00"),
                        ft.Text("En Celular (Android / iPhone):", weight="bold", color="white")
                    ]),
                    ft.Text("1. Toca la barra de texto para escribir.\n2. Abre el teclado virtual y presiona el icono de micrófono que viene incorporado en tu teclado (al lado de la barra espaciadora o en la barra de sugerencias).\n3. Empieza a hablar.", color="#aaaaaa"),
                    ft.Divider(height=10, color="transparent"),
                    ft.Row([
                        ft.Icon(ft.Icons.COMPUTER, color="#00FFFF"),
                        ft.Text("En Computadora (Windows):", weight="bold", color="white")
                    ]),
                    ft.Text("1. Haz clic en la barra de texto para escribir.\n2. Presiona la combinación de teclas Windows + H.\n3. Asegúrate de tener un micrófono activo en tu configuración de sonido de Windows.", color="#aaaaaa"),
                ], tight=True, spacing=10),
                actions=[
                    ft.TextButton("Entendido", on_click=cerrar_dictado_dialog)
                ]
            )
            page.show_dialog(dialog)

        def build_chat_view():
            return ft.Column([
                ft.Row([
                    ft.Text("Luxo AI Assistant", size=24, color="#D8B4FE", weight="bold"),
                    ft.Container(expand=True),
                    ft.Icon(ft.Icons.MIC_ROUNDED, color="#aaaaaa", size=16),
                    ft.Text("Toca el icono de micrófono 🎙️ para ver cómo dictar con voz", color="#aaaaaa", size=11, italic=True),
                ], vertical_alignment="center"),
                ft.Container(
                    content=ft.SelectionArea(content=chat_display),
                    expand=True,
                    bgcolor="#000000",
                    border_radius=20,
                    padding=10,
                    border=ft.Border.all(2, "#D8B4FE"),
                    shadow=[
                        ft.BoxShadow(
                            color="#D8B4FE",
                            blur_radius=15,
                            spread_radius=1,
                        )
                    ]
                ),
                ft.Row([
                    input_msg,
                    ft.IconButton(
                        icon=ft.Icons.MIC_ROUNDED,
                        icon_color="#D8B4FE",
                        tooltip="Instrucciones de Dictado",
                        on_click=mostrar_instrucciones_dictado
                    ),
                    ft.ElevatedButton(
                        "ENVIAR",
                        on_click=enviar_mensaje,
                        bgcolor="#6E48AA",
                        color="white"
                    )
                ], spacing=10)
            ], expand=True)

        def build_historial_view():
            historial_list = ft.Column(spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cargar_lista_historial():
                historial_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT Pregunta_Usuario, Respuesta_IA, Fecha_Hora 
                            FROM historial_conversaciones 
                            WHERE ID_Usuario = %s 
                            ORDER BY Fecha_Hora DESC 
                            LIMIT 30
                        """, (user_info["id"],))
                        historial = cursor.fetchall()
                        db.close()
                        
                        if not historial:
                            historial_list.controls.append(
                                ft.Container(
                                    content=ft.Text("No tienes consultas anteriores registradas.", color="#aaaaaa", size=14),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    expand=True
                                )
                            )
                        else:
                            for row in historial:
                                fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                                historial_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"💬 Pregunta: {row['Pregunta_Usuario']}", color="white", weight="bold"),
                                            ft.Text(f"🤖 Respuesta: {row['Respuesta_IA']}", color="#D8B4FE"),
                                        ], spacing=4),
                                        bgcolor="#222222",
                                        padding=15,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR HISTORIAL VIEW:", ex)
                    historial_list.controls.append(ft.Text("Error al cargar el historial.", color="red"))
                page.update()
                
            cargar_lista_historial()
            
            def confirmar_borrado_historial(e):
                def on_confirmar(ev):
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor()
                            cursor.execute("""
                                DELETE FROM pendientes_actualizacion 
                                WHERE ID_Conversacion IN (
                                    SELECT ID_Conversacion FROM historial_conversaciones WHERE ID_Usuario = %s
                                )
                            """, (user_info["id"],))
                            cursor.execute("DELETE FROM historial_conversaciones WHERE ID_Usuario = %s", (user_info["id"],))
                            db.commit()
                            db.close()
                            mostrar_snack("Historial borrado correctamente.")
                            cargar_lista_historial()
                            chat_display.controls.clear()
                            historial_sesion.clear()
                            page.pop_dialog()
                            page.update()
                    except Exception as ex:
                        print("ERROR BORRAR HISTORIAL:", ex)
                        mostrar_snack("Error al borrar el historial.", color="red")
                
                def on_cancelar(ev):
                    page.pop_dialog()
                    
                dialog_confirm = ft.AlertDialog(
                    title=ft.Text("Confirmar Borrado", color="#FF4500", weight="bold"),
                    content=ft.Text("¿Seguro que deseas borrar todo tu historial de conversaciones? Esta acción no se puede deshacer."),
                    actions=[
                        ft.TextButton("Cancelar", on_click=on_cancelar),
                        ft.ElevatedButton("Borrar Todo", on_click=on_confirmar, bgcolor="#FF4500", color="white")
                    ],
                    actions_alignment="end",
                    bgcolor="#111111"
                )
                page.show_dialog(dialog_confirm)

            btn_clear = ft.ElevatedButton(
                "Borrar Historial",
                icon=ft.Icons.DELETE_SWEEP,
                on_click=confirmar_borrado_historial,
                bgcolor="#FF4500",
                color="white"
            )

            return ft.SelectionArea(
                content=ft.Column([
                    ft.Row([
                        ft.Text("Historial de Consultas", size=24, color="#D8B4FE", weight="bold"),
                        btn_clear
                    ], alignment="spaceBetween", vertical_alignment="center"),
                    ft.Divider(height=20, color="#333333"),
                    historial_list
                ], expand=True)
            )

        def build_stats_tab():
            total_consultas = 0
            utiles = 0
            no_utiles = 0
            total_tickets = 0
            tickets_resueltos = 0
            negatives_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            positives_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            logins_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            categories_cnt = []
            
            # Variables de cumplimiento por zona
            checklist_pct = 0.0
            total_stores_zone = 0
            
            # Métricas de campaña activa por zona
            total_campaign_stores = 0
            aprobadas_ia_cnt = 0
            visto_bueno_cnt = 0
            rechazadas_ia_cnt = 0
            pendientes_cnt = 0
            sin_entrega_cnt = 0
            
            try:
                db = conectar_db()
                if db:
                    cursor = db.cursor(dictionary=True)
                    cursor.execute("SELECT COUNT(*) as cnt FROM historial_conversaciones")
                    total_consultas = cursor.fetchone()["cnt"]
                    
                    cursor.execute("SELECT COUNT(*) as cnt FROM historial_conversaciones WHERE Me_Sirvio = 1")
                    utiles = cursor.fetchone()["cnt"]
                    
                    cursor.execute("SELECT COUNT(*) as cnt FROM historial_conversaciones WHERE Me_Sirvio = 0")
                    no_utiles = cursor.fetchone()["cnt"]

                    # Conteo de tickets
                    cursor.execute("SELECT COUNT(*) as cnt FROM tickets_soporte")
                    total_tickets = cursor.fetchone()["cnt"]
                    
                    cursor.execute("SELECT COUNT(*) as cnt FROM tickets_soporte WHERE Estatus = 'Resuelto'")
                    tickets_resueltos = cursor.fetchone()["cnt"]

                    # Consultas agrupadas por categoría
                    cursor.execute("""
                        SELECT COALESCE(Categoria, 'Otros') as cat, COUNT(*) as cnt 
                        FROM pendientes_actualizacion 
                        GROUP BY Categoria
                    """)
                    categories_cnt = cursor.fetchall()
                    
                    # --- CÁLCULO DE CUMPLIMIENTO POR ZONA ---
                    zona_act = active_zone_filter[0]
                    
                    # 1. Porcentaje de cumplimiento de bitácoras (Checklists) diarias
                    cursor.execute("SELECT COUNT(*) as cnt FROM plantillas_checklist")
                    total_plantillas = cursor.fetchone()["cnt"] or 1
                    
                    # Buscar gerentes de la zona
                    if zona_act != "Todas":
                        cursor.execute("SELECT ID_Usuario, Tienda FROM usuarios WHERE Rol = 'Gerente' AND Tienda IS NOT NULL AND Tienda != '' AND Zona = %s", (zona_act,))
                    else:
                        cursor.execute("SELECT ID_Usuario, Tienda FROM usuarios WHERE Rol = 'Gerente' AND Tienda IS NOT NULL AND Tienda != ''")
                    gerentes_zona = cursor.fetchall()
                    total_stores_zone = len(gerentes_zona)
                    
                    completed_sum = 0
                    if total_stores_zone > 0:
                        for g_z in gerentes_zona:
                            cursor.execute("""
                                SELECT COUNT(DISTINCT ID_Plantilla) as cnt 
                                FROM registro_checklist 
                                WHERE ID_Usuario = %s AND Fecha = CURDATE() AND Completado = 1
                            """, (g_z["ID_Usuario"],))
                            completed_sum += cursor.fetchone()["cnt"]
                        checklist_pct = (completed_sum / (total_stores_zone * total_plantillas)) * 100
                    else:
                        checklist_pct = 0.0

                    # 2. Métricas de campañas del mes
                    cursor.execute("SELECT ID_Campana FROM campanas WHERE Estatus = 'Activa'")
                    active_camp_row = cursor.fetchone()
                    if active_camp_row:
                        id_active_camp = active_camp_row["ID_Campana"]
                        total_campaign_stores = total_stores_zone
                        
                        if total_campaign_stores > 0:
                            placeholders = ",".join(["%s"] * total_campaign_stores)
                            ids_gerentes = [g_z["ID_Usuario"] for g_z in gerentes_zona]
                            
                            # Aprobado IA
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Aprobado_IA' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            aprobadas_ia_cnt = cursor.fetchone()["cnt"]
                            
                            # Visto Bueno
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Visto_Bueno' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            visto_bueno_cnt = cursor.fetchone()["cnt"]
                            
                            # Rechazado IA
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Rechazado_IA' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            rechazadas_ia_cnt = cursor.fetchone()["cnt"]
                            
                            # Pendiente
                            cursor.execute(f"""
                                SELECT COUNT(*) as cnt 
                                FROM campana_entregas_tienda 
                                WHERE ID_Campana = %s AND Estatus = 'Pendiente' AND ID_Usuario IN ({placeholders})
                            """, [id_active_camp] + ids_gerentes)
                            pendientes_cnt = cursor.fetchone()["cnt"]
                            
                            sin_entrega_cnt = total_campaign_stores - (aprobadas_ia_cnt + visto_bueno_cnt + rechazadas_ia_cnt + pendientes_cnt)
                    
                    # Cargar negativos
                    cursor.execute("""
                        SELECT h.Fecha_Hora, u.Nombre_Completo, h.Pregunta_Usuario, h.Respuesta_IA, h.Comentario_Feedback
                        FROM historial_conversaciones h
                        JOIN usuarios u ON h.ID_Usuario = u.ID_Usuario
                        WHERE h.Me_Sirvio = 0
                        ORDER BY h.Fecha_Hora DESC
                    """)
                    negatives = cursor.fetchall()
                    
                    # Cargar positivos
                    cursor.execute("""
                        SELECT h.Fecha_Hora, u.Nombre_Completo, h.Pregunta_Usuario, h.Respuesta_IA, h.Comentario_Feedback
                        FROM historial_conversaciones h
                        JOIN usuarios u ON h.ID_Usuario = u.ID_Usuario
                        WHERE h.Me_Sirvio = 1
                        ORDER BY h.Fecha_Hora DESC
                    """)
                    positives = cursor.fetchall()

                    # Cargar inicios de sesión
                    cursor.execute("""
                        SELECT s.Fecha_Login, u.Nombre_Completo, s.Direccion_IP, s.Ubicacion_Ciudad, s.Ubicacion_Pais
                        FROM sesiones s
                        JOIN usuarios u ON s.ID_Usuario = u.ID_Usuario
                        ORDER BY s.Fecha_Login DESC
                        LIMIT 15
                    """)
                    logins = cursor.fetchall()
                    
                    db.close()
                    
                    if not negatives:
                        negatives_list.controls.append(ft.Text("No hay respuestas calificadas negativamente.", color="#aaaaaa", size=14))
                    else:
                        for row in negatives:
                            fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                            comentario = row["Comentario_Feedback"] or "(Sin comentario)"
                            negatives_list.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"👤 Usuario: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                        ], alignment="spaceBetween"),
                                        ft.Text(f"💬 Pregunta: {row['Pregunta_Usuario']}", color="white", weight="bold"),
                                        ft.Text(f"🤖 Respuesta de Luxo: {row['Respuesta_IA']}", color="#D8B4FE"),
                                        ft.Container(
                                            content=ft.Text(f"Razón: {comentario}", color="#FF4500", size=12, italic=True),
                                            bgcolor="#3d1f1f",
                                            padding=8,
                                            border_radius=5
                                        )
                                    ], spacing=4),
                                    bgcolor="#222222",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
                            
                    if not positives:
                        positives_list.controls.append(ft.Text("No hay respuestas calificadas positivamente.", color="#aaaaaa", size=14))
                    else:
                        for row in positives:
                            fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                            comentario = row["Comentario_Feedback"] or "(Sin comentario)"
                            positives_list.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"👤 Usuario: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                        ], alignment="spaceBetween"),
                                        ft.Text(f"💬 Pregunta: {row['Pregunta_Usuario']}", color="white", weight="bold"),
                                        ft.Text(f"🤖 Respuesta de Luxo: {row['Respuesta_IA']}", color="#D8B4FE"),
                                        ft.Container(
                                            content=ft.Text(f"Comentario: {comentario}", color="#7CFC00", size=12, italic=True),
                                            bgcolor="#1b3d1f",
                                            padding=8,
                                            border_radius=5
                                        )
                                    ], spacing=4),
                                    bgcolor="#222222",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )

                    if not logins:
                        logins_list.controls.append(ft.Text("No hay registros de inicio de sesión.", color="#aaaaaa", size=14))
                    else:
                        for row in logins:
                            fecha = row["Fecha_Login"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Login"] else ""
                            ip = row["Direccion_IP"] or "Desconocida"
                            ciudad = row["Ubicacion_Ciudad"] or "Desconocida"
                            pais = row["Ubicacion_Pais"] or "Desconocido"
                            logins_list.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                            ft.Text(f"👤 {row['Nombre_Completo']}", color="white", weight="bold"),
                                        ], alignment="spaceBetween"),
                                        ft.Row([
                                            ft.Text(f"🌐 IP: {ip}", color="#00FFFF", size=12),
                                            ft.Text(f"📍 {ciudad}, {pais}", color="#D8B4FE", size=12),
                                        ], alignment="spaceBetween"),
                                    ], spacing=4),
                                    bgcolor="#222222",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
            except Exception as ex:
                print("ERROR STATS TAB:", ex)
                
            eficacia = 0.0
            if (utiles + no_utiles) > 0:
                eficacia = (utiles / (utiles + no_utiles)) * 100

            # 1. KPI Total Consultas
            kpi_card_1 = ft.Container(
                content=ft.Column([
                    ft.Text("Total Consultas", color="#aaaaaa", size=11, weight="bold"),
                    ft.Text(str(total_consultas), color="#D8B4FE", size=26, weight="bold"),
                    ft.Icon(ft.Icons.CHAT_ROUNDED, color="#D8B4FE", size=22)
                ], spacing=8, alignment="center", horizontal_alignment="center"),
                bgcolor="#111111",
                border=ft.Border.all(1.5, "#D8B4FE"),
                border_radius=10,
                padding=10,
                width=150,
                height=140
            )

            # 2. KPI Tickets de Soporte (Barra de progreso horizontal)
            tickets_pct = tickets_resueltos / total_tickets if total_tickets > 0 else 0.0
            tickets_bar = ft.Column([
                ft.Text(f"Resueltos: {tickets_resueltos} de {total_tickets}", color="#00FFFF", size=11, weight="bold"),
                ft.ProgressBar(value=tickets_pct, color="#00FFFF", bgcolor="#222222", width=120),
                ft.Text(f"Tasa de Resolución: {int(tickets_pct * 100)}%", color="#aaaaaa", size=10)
            ], horizontal_alignment="center", spacing=5, alignment="center")
            
            kpi_card_2 = ft.Container(
                content=ft.Column([
                    ft.Text("Soporte Técnico 🎫", color="#aaaaaa", size=11, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    tickets_bar
                ], spacing=5, alignment="center", horizontal_alignment="center"),
                bgcolor="#111111",
                border=ft.Border.all(1.5, "#00FFFF"),
                border_radius=10,
                padding=10,
                width=170,
                height=140
            )

            # 3. KPI Eficacia de IA (Anillo de progreso circular)
            eficacia_ring = ft.Container(
                content=ft.Stack([
                    ft.ProgressRing(
                        value=eficacia / 100 if (utiles + no_utiles) > 0 else 0.0,
                        stroke_width=8,
                        color="#7CFC00" if eficacia >= 75 else ("#00FFFF" if eficacia >= 50 else "#FF4500"),
                        bgcolor="#222222",
                        width=85,
                        height=85
                    ),
                    ft.Container(
                        content=ft.Text(f"{eficacia:.0f}%", color="white", size=15, weight="bold"),
                        alignment=ft.alignment.Alignment(0, 0)
                    )
                ], width=85, height=85),
                alignment=ft.alignment.Alignment(0, 0)
            )
            
            kpi_card_3 = ft.Container(
                content=ft.Column([
                    ft.Text("Eficacia de la IA", color="#aaaaaa", size=11, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    eficacia_ring
                ], spacing=2, alignment="center", horizontal_alignment="center"),
                bgcolor="#111111",
                border=ft.Border.all(1.5, "#7CFC00"),
                border_radius=10,
                padding=10,
                width=170,
                height=140
            )

            # 4. KPI Categorías Faltantes (Gráfico de barras personalizado)
            bar_controls = []
            max_val = max([c["cnt"] for c in categories_cnt]) if categories_cnt else 1
            cat_colors = {
                "Impresoras": "#7CFC00",
                "Políticas de Venta": "#00FFFF",
                "Sistemas/Terminales": "#FF4500",
                "Manuales": "#D8B4FE",
                "Otros": "#888888"
            }
            
            for c in categories_cnt:
                cat_name = c["cat"]
                count = c["cnt"]
                color = cat_colors.get(cat_name, "#888888")
                
                # Proportional height (max 65px, min 10px)
                height = (count / max_val) * 65 if max_val > 0 else 10
                height = max(height, 15)
                
                bar_controls.append(
                    ft.Column([
                        ft.Text(str(count), color="white", size=9, weight="bold"),
                        ft.Container(
                            width=22,
                            height=height,
                            bgcolor=color,
                            border_radius=ft.BorderRadius(top_left=3, top_right=3, bottom_left=0, bottom_right=0),
                            shadow=ft.BoxShadow(color=color, blur_radius=5, spread_radius=0.1)
                        ),
                        ft.Text(cat_name[:4] + ".." if len(cat_name) > 4 else cat_name, color="#aaaaaa", size=8)
                    ], horizontal_alignment="center", spacing=2, alignment="end")
                )
            
            bar_chart_row = ft.Row(
                bar_controls,
                spacing=8,
                alignment="center",
                vertical_alignment="end"
            ) if bar_controls else ft.Text("Sin preguntas faltantes", color="#aaaaaa", size=11, italic=True)
            
            kpi_card_4 = ft.Container(
                content=ft.Column([
                    ft.Text("Categorías IA Faltantes", color="#aaaaaa", size=11, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    bar_chart_row
                ], spacing=2, alignment="center", horizontal_alignment="center"),
                bgcolor="#111111",
                border=ft.Border.all(1.5, "#A100F2"),
                border_radius=10,
                padding=10,
                width=240,
                height=140
            )

            kpi_row = ft.Row([
                kpi_card_1,
                kpi_card_2,
                kpi_card_3,
                kpi_card_4
            ], spacing=15, alignment="center")

            lists_row = ft.Row([
                ft.Column([
                    ft.Text("Auditoría de Inicios de Sesión (🌐)", size=15, color="#00FFFF", weight="bold"),
                    ft.Divider(height=5, color="transparent"),
                    logins_list
                ], expand=True, spacing=5),
                ft.VerticalDivider(width=20, color="#333333"),
                ft.Column([
                    ft.Text("Respuestas Insatisfactorias (👎)", size=15, color="#FF4500", weight="bold"),
                    ft.Divider(height=5, color="transparent"),
                    negatives_list
                ], expand=True, spacing=5),
                ft.VerticalDivider(width=20, color="#333333"),
                ft.Column([
                    ft.Text("Respuestas Satisfactorias (👍)", size=15, color="#7CFC00", weight="bold"),
                    ft.Divider(height=5, color="transparent"),
                    positives_list
                ], expand=True, spacing=5)
            ], expand=True, spacing=10)

            # --- WIDGETS DE CUMPLIMIENTO ZONAL ---
            checklist_ring = ft.Container(
                content=ft.Stack([
                    ft.ProgressRing(
                        value=checklist_pct / 100.0,
                        stroke_width=8,
                        color="#7CFC00" if checklist_pct >= 75 else ("#00FFFF" if checklist_pct >= 50 else "#FF4500"),
                        bgcolor="#222222",
                        width=85,
                        height=85
                    ),
                    ft.Container(
                        content=ft.Text(f"{checklist_pct:.0f}%", color="white", size=15, weight="bold"),
                        alignment=ft.alignment.Alignment(0, 0)
                    )
                ], width=85, height=85),
                alignment=ft.alignment.Alignment(0, 0)
            )

            compliance_card = ft.Container(
                content=ft.Column([
                    ft.Text("Bitácoras Diarias", color="#aaaaaa", size=11, weight="bold"),
                    ft.Text(f"{zona_act}", color="#00FFFF", size=12, weight="bold"),
                    ft.Divider(height=2, color="transparent"),
                    checklist_ring,
                    ft.Text(f"Tiendas: {total_stores_zone}", size=11, color="#aaaaaa")
                ], spacing=3, alignment="center", horizontal_alignment="center"),
                bgcolor="#111111",
                border=ft.Border.all(1.5, "#00FFFF"),
                border_radius=10,
                padding=10,
                width=170,
                height=180
            )

            def build_horizontal_bar(label, count, total, color):
                pct = (count / total * 100) if total > 0 else 0.0
                bar_width = (count / total * 200) if total > 0 else 0
                return ft.Row([
                    ft.Text(label, size=11, color="white", width=95),
                    ft.Stack([
                        ft.Container(width=200, height=12, bgcolor="#222222", border_radius=3),
                        ft.Container(width=max(bar_width, 4) if count > 0 else 0, height=12, bgcolor=color, border_radius=3, shadow=ft.BoxShadow(color=color, blur_radius=3, spread_radius=0.1))
                    ]),
                    ft.Text(f"{count} ({pct:.0f}%)", size=11, color="#aaaaaa", width=65)
                ], spacing=5, alignment="start", vertical_alignment="center")

            campaign_chart_card = ft.Container(
                content=ft.Column([
                    ft.Text("Estado de Campaña Mensual", color="#aaaaaa", size=11, weight="bold"),
                    ft.Text(f"Zona: {zona_act}", color="#D8B4FE", size=12, weight="bold"),
                    ft.Divider(height=1, color="#333333"),
                    build_horizontal_bar("Visto Bueno 👑", visto_bueno_cnt, total_campaign_stores, "#00FF7F"),
                    build_horizontal_bar("Aprobado IA 🤖", aprobadas_ia_cnt, total_campaign_stores, "#7CFC00"),
                    build_horizontal_bar("Rechazado IA ⚠️", rechazadas_ia_cnt, total_campaign_stores, "#FF4500"),
                    build_horizontal_bar("Pendiente ⏳", pendientes_cnt, total_campaign_stores, "#FFD700"),
                    build_horizontal_bar("Sin Entrega ❌", sin_entrega_cnt, total_campaign_stores, "#888888")
                ], spacing=6, alignment="center", horizontal_alignment="start"),
                bgcolor="#111111",
                border=ft.Border.all(1.5, "#D8B4FE"),
                border_radius=10,
                padding=12,
                width=400,
                height=180
            )

            compliance_row = ft.Row([
                compliance_card,
                campaign_chart_card
            ], spacing=15, alignment="center")

            return ft.Column([
                ft.Row([
                    ft.Text("Métricas de Control Geográfico", size=18, color="white", weight="bold"),
                    ft.Container(
                        content=ft.Text(f"Filtro: {zona_act}", size=12, color="black", weight="bold"),
                        bgcolor="#00FFFF",
                        padding=ft.Padding(left=8, right=8, top=3, bottom=3),
                        border_radius=5
                    )
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Divider(height=5, color="transparent"),
                compliance_row,
                ft.Divider(height=15, color="#333333"),
                ft.Text("Indicadores de Uso y Calidad (Global)", size=18, color="white", weight="bold"),
                ft.Divider(height=5, color="transparent"),
                kpi_row,
                ft.Divider(height=20, color="#333333"),
                lists_row
            ], expand=True, spacing=10)

        def build_missing_questions_tab():
            questions_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cargar_preguntas():
                questions_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT ID_Pendiente, Pregunta_Faltante, Fecha_Registro, Categoria 
                            FROM pendientes_actualizacion 
                            WHERE Estatus = 'Pendiente'
                            ORDER BY Fecha_Registro DESC
                        """)
                        preguntas = cursor.fetchall()
                        db.close()
                        
                        if not preguntas:
                            questions_list.controls.append(
                                ft.Container(
                                    content=ft.Text("No hay preguntas pendientes de actualizar.", color="#7CFC00", size=14),
                                    alignment=ft.alignment.Alignment(0, 0),
                                    expand=True
                                )
                            )
                        else:
                            for row in preguntas:
                                id_p = row["ID_Pendiente"]
                                pregunta = row["Pregunta_Faltante"]
                                fecha = row["Fecha_Registro"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Registro"] else ""
                                cat_text = row["Categoria"] or "Clasificando..."
                                
                                # Definir color del tag según la categoría
                                cat_colors = {
                                    "Impresoras": "#7CFC00",
                                    "Políticas de Venta": "#00FFFF",
                                    "Sistemas/Terminales": "#FF4500",
                                    "Manuales": "#D8B4FE",
                                    "Otros": "#aaaaaa"
                                }
                                cat_color = cat_colors.get(cat_text, "#aaaaaa")
                                
                                category_badge = ft.Container(
                                    content=ft.Text(cat_text, color="black", size=10, weight="bold"),
                                    bgcolor=cat_color,
                                    padding=ft.Padding(left=8, right=8, top=3, bottom=3),
                                    border_radius=5
                                )
                                
                                def resolver_click(e, id_pend=id_p):
                                    try:
                                        db_res = conectar_db()
                                        if db_res:
                                            cursor_res = db_res.cursor()
                                            cursor_res.execute(
                                                "UPDATE pendientes_actualizacion SET Estatus = 'Resuelto' WHERE ID_Pendiente = %s",
                                                (id_pend,)
                                            )
                                            db_res.commit()
                                            db_res.close()
                                            mostrar_snack("Pregunta marcada como resuelta.")
                                            cargar_preguntas()
                                            page.update()
                                    except Exception as ex:
                                        print("ERROR MARCAR RESUELTO:", ex)
                                
                                questions_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Column([
                                                ft.Row([
                                                    ft.Text(pregunta, color="white", weight="bold", size=14),
                                                    category_badge
                                                ], spacing=10, vertical_alignment="center"),
                                                ft.Text(f"Registrada el: {fecha}", color="#aaaaaa", size=11)
                                            ], spacing=3, expand=True),
                                            ft.ElevatedButton(
                                                "Marcar Resuelta",
                                                icon=ft.Icons.CHECK,
                                                bgcolor="#6E48AA",
                                                color="white",
                                                on_click=resolver_click
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#222222",
                                        padding=12,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR MISSING QUESTIONS:", ex)
                    questions_list.controls.append(ft.Text("Error al cargar las preguntas pendientes.", color="red"))
                page.update()
                    
            cargar_preguntas()
            
            return ft.Column([
                ft.Text("Preguntas que la IA no pudo responder (Falta Información)", size=18, color="white", weight="bold"),
                ft.Text("Usa esta lista para identificar qué manuales o temas faltan en el sistema y cárgalos.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                questions_list
            ], expand=True)

        def build_manuals_tab():
            manuals_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cargar_manuales():
                manuals_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version, Abierto FROM manuales ORDER BY Nombre_Archivo")
                        manuales = cursor.fetchall()
                        db.close()
                        
                        manuals_list.controls.append(ft.Text("Manuales Base de Datos (PDF/Excel)", size=14, color="#00FFFF", weight="bold"))
                        if not manuales:
                            manuals_list.controls.append(ft.Text("No hay manuales cargados en la base de datos.", color="#aaaaaa", size=12))
                        else:
                            for m in manuales:
                                id_m = m["ID_Manual"]
                                nombre = m.get("Nombre_Archivo") or ""
                                version = m.get("Version") or ""
                                es_abierto = m.get("Abierto") if m.get("Abierto") is not None else 1
                                
                                def borrar_click(e, id_man=id_m, nom=nombre):
                                    def on_confirmar(ev):
                                        try:
                                            db_del = conectar_db()
                                            if db_del:
                                                cursor_del = db_del.cursor()
                                                cursor_del.execute("""
                                                    DELETE FROM pendientes_actualizacion 
                                                    WHERE ID_Conversacion IN (
                                                        SELECT ID_Conversacion FROM historial_conversaciones WHERE ID_Manual = %s
                                                    )
                                                """, (id_man,))
                                                cursor_del.execute("DELETE FROM historial_conversaciones WHERE ID_Manual = %s", (id_man,))
                                                cursor_del.execute("DELETE FROM manuales WHERE ID_Manual = %s", (id_man,))
                                                db_del.commit()
                                                db_del.close()
                                                mostrar_snack(f"Manual '{nom}' eliminado.")
                                                cargar_manuales()
                                                page.pop_dialog()
                                                page.update()
                                        except Exception as ex:
                                            print("ERROR BORRAR MANUAL:", ex)
                                            mostrar_snack("Error al borrar manual.", color="red")
                                            
                                    def on_cancelar(ev):
                                        page.pop_dialog()
                                        
                                    dialog_confirm = ft.AlertDialog(
                                        title=ft.Text("Confirmar Borrado", color="#FF4500", weight="bold"),
                                        content=ft.Text(f"¿Seguro que deseas borrar el archivo \"{nom}\"?"),
                                        actions=[
                                            ft.TextButton("Cancelar", on_click=on_cancelar),
                                            ft.ElevatedButton("Sí, Borrar", on_click=on_confirmar, bgcolor="#FF4500", color="white")
                                        ],
                                        actions_alignment="end",
                                        bgcolor="#111111"
                                    )
                                    page.show_dialog(dialog_confirm)

                                def toggle_abierto_click(e, id_man=id_m, act_abierto=es_abierto):
                                    nuevo_estado = 0 if act_abierto == 1 else 1
                                    try:
                                        db_toggle = conectar_db()
                                        if db_toggle:
                                            cursor_toggle = db_toggle.cursor()
                                            cursor_toggle.execute(
                                                "UPDATE manuales SET Abierto = %s WHERE ID_Manual = %s",
                                                (nuevo_estado, id_man)
                                            )
                                            db_toggle.commit()
                                            db_toggle.close()
                                            mostrar_snack("Estado del manual actualizado.")
                                            cargar_manuales()
                                            page.update()
                                    except Exception as ex:
                                        print("ERROR TOGGLE ABIERTO:", ex)

                                icon_toggle = ft.Icons.LOCK_OPEN_ROUNDED if es_abierto == 1 else ft.Icons.LOCK_ROUNDED
                                color_toggle = "#7CFC00" if es_abierto == 1 else "#FF4500"
                                tooltip_toggle = "Manual Abierto (Click para bloquear)" if es_abierto == 1 else "Manual Cerrado (Click para desbloquear)"

                                btn_lock = ft.IconButton(
                                    icon=icon_toggle,
                                    icon_color=color_toggle,
                                    tooltip=tooltip_toggle,
                                    on_click=lambda e, id_man=id_m, act=es_abierto: toggle_abierto_click(e, id_man, act)
                                )
 
                                manuals_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.INSERT_DRIVE_FILE, color="#00FFFF"),
                                            ft.Column([
                                                ft.Text(nombre, color="white", weight="bold", size=14),
                                                ft.Text(f"Versión: {version}", color="#aaaaaa", size=11)
                                            ], spacing=3, expand=True),
                                            btn_lock,
                                            ft.IconButton(
                                                icon=ft.Icons.DELETE_FOREVER,
                                                icon_color="#FF4500",
                                                tooltip="Eliminar manual",
                                                on_click=borrar_click
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#222222",
                                        padding=10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR MANUALS LIST:", ex)
                    manuals_list.controls.append(ft.Text("Error al cargar la lista de manuales.", color="red"))
                
                # --- Recursos Multimedia Locales (assets/) ---
                manuals_list.controls.append(ft.Divider(height=15, color="#333333"))
                manuals_list.controls.append(ft.Text("Recursos Multimedia Locales (assets/)", size=14, color="#A100F2", weight="bold"))
                try:
                    os.makedirs("assets", exist_ok=True)
                    archivos = [f for f in os.listdir("assets") if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.mp4', '.mov', '.avi'))]
                    if not archivos:
                        manuals_list.controls.append(ft.Text("No hay archivos multimedia cargados en assets/.", color="#aaaaaa", size=12))
                    else:
                        for filename in archivos:
                            def borrar_media_click(e, fn=filename):
                                def on_confirmar_media(ev):
                                    try:
                                        ruta_media = os.path.join("assets", fn)
                                        if os.path.exists(ruta_media):
                                            os.remove(ruta_media)
                                        mostrar_snack(f"Archivo multimedia '{fn}' eliminado.")
                                        cargar_manuales()
                                        page.pop_dialog()
                                        page.update()
                                    except Exception as ex:
                                        print("ERROR BORRAR MEDIA:", ex)
                                        mostrar_snack("Error al borrar archivo multimedia.", color="red")
                                        
                                def on_cancelar_media(ev):
                                    page.pop_dialog()
                                    
                                dialog_confirm_media = ft.AlertDialog(
                                    title=ft.Text("Confirmar Borrado de Multimedia", color="#FF4500", weight="bold"),
                                    content=ft.Text(f"¿Seguro que deseas borrar el archivo multimedia \"{fn}\"?"),
                                    actions=[
                                        ft.TextButton("Cancelar", on_click=on_cancelar_media),
                                        ft.ElevatedButton("Sí, Borrar", on_click=on_confirmar_media, bgcolor="#FF4500", color="white")
                                    ],
                                    actions_alignment="end",
                                    bgcolor="#111111"
                                )
                                page.show_dialog(dialog_confirm_media)

                            is_video = filename.lower().endswith(('.mp4', '.mov', '.avi'))
                            icon_media = ft.Icons.PLAY_CIRCLE_FILL if is_video else ft.Icons.IMAGE
                            color_media = "#A100F2"
                            
                            manuals_list.controls.append(
                                ft.Container(
                                    content=ft.Row([
                                        ft.Icon(icon_media, color=color_media),
                                        ft.Column([
                                            ft.Text(filename, color="white", weight="bold", size=14),
                                            ft.Text("Ubicación: local assets/", color="#aaaaaa", size=11)
                                        ], spacing=3, expand=True),
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE_FOREVER,
                                            icon_color="#FF4500",
                                            tooltip="Eliminar multimedia",
                                            on_click=borrar_media_click
                                        )
                                    ], alignment="spaceBetween", vertical_alignment="center"),
                                    bgcolor="#222222",
                                    padding=10,
                                    border_radius=8,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
                except Exception as ex:
                    print("ERROR LISTA MULTIMEDIA:", ex)
                    manuals_list.controls.append(ft.Text("Error al cargar la lista de archivos multimedia.", color="red"))
                
                page.update()
                    
            cargar_manuales()
            
            def on_pdf_cargado(ruta):
                procesar_cargar_pdf(ruta)
                def reload_after_delay():
                    import time
                    time.sleep(1.5)
                    cargar_manuales()
                    page.update()
                threading.Thread(target=reload_after_delay, daemon=True).start()
                
            def on_excel_cargado(ruta):
                procesar_cargar_excel(ruta)
                def reload_after_delay():
                    import time
                    time.sleep(1.5)
                    cargar_manuales()
                    page.update()
                threading.Thread(target=reload_after_delay, daemon=True).start()

            def on_multimedia_cargado(ruta):
                try:
                    os.makedirs("assets", exist_ok=True)
                    nombre_archivo = os.path.basename(ruta)
                    destino = os.path.join("assets", nombre_archivo)
                    shutil.copy(ruta, destino)
                    mostrar_snack(f"Archivo multimedia '{nombre_archivo}' cargado con éxito en assets/.")
                    
                    # Recargar después de cargar
                    def reload_after_delay():
                        import time
                        time.sleep(1.5)
                        cargar_manuales()
                        page.update()
                    threading.Thread(target=reload_after_delay, daemon=True).start()
                except Exception as ex:
                    print("ERROR CARGANDO MULTIMEDIA:", ex)
                    mostrar_snack("Error al guardar archivo multimedia.", color="red")

            btn_pdf = ft.ElevatedButton(
                "Cargar PDF",
                icon=ft.Icons.PICTURE_AS_PDF,
                bgcolor="#6E48AA",
                color="white",
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar PDF para cargar",
                    [("PDF files", "*.pdf")],
                    on_pdf_cargado
                )
            )
            
            btn_excel = ft.ElevatedButton(
                "Cargar Excel",
                icon=ft.Icons.TABLE_CHART,
                bgcolor="#1f6f43",
                color="white",
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar Excel para cargar",
                    [("Excel files", "*.xlsx *.xls"), ("Todos los archivos", "*.*")],
                    on_excel_cargado
                )
            )

            btn_media = ft.ElevatedButton(
                "Cargar Multimedia",
                icon=ft.Icons.PERM_MEDIA,
                bgcolor="#A100F2",
                color="white",
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar Imagen, GIF o Video",
                    [
                        ("Archivos Multimedia", "*.png *.jpg *.jpeg *.gif *.mp4 *.mov *.avi"),
                        ("Imágenes (*.png, *.jpg, *.gif)", "*.png *.jpg *.jpeg *.gif"),
                        ("Videos (*.mp4, *.mov)", "*.mp4 *.mov *.avi"),
                        ("Todos los archivos", "*.*")
                    ],
                    on_multimedia_cargado
                )
            )

            return ft.Column([
                ft.Row([
                    ft.Text("Manuales y Documentos de Sunglass Hut", size=18, color="white", weight="bold"),
                    ft.Row([btn_pdf, btn_excel, btn_media], spacing=10)
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Divider(height=10, color="transparent"),
                manuals_list
            ], expand=True)

        def build_suggestions_tab():
            suggestions_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cargar_sugerencias():
                suggestions_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            CREATE TABLE IF NOT EXISTS sugerencias_luxo (
                                ID_Sugerencia INT AUTO_INCREMENT PRIMARY KEY,
                                ID_Usuario INT NOT NULL,
                                Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                                Sugerencia TEXT NOT NULL,
                                FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE
                            )
                        """)
                        db.commit()
                        
                        cursor.execute("""
                            SELECT s.Fecha_Hora, u.Nombre_Completo, s.Sugerencia 
                            FROM sugerencias_luxo s
                            JOIN usuarios u ON s.ID_Usuario = u.ID_Usuario
                            ORDER BY s.Fecha_Hora DESC
                        """)
                        sugerencias = cursor.fetchall()
                        db.close()
                        
                        if not sugerencias:
                            suggestions_list.controls.append(
                                ft.Text("No hay sugerencias registradas de los usuarios.", color="#aaaaaa", size=14)
                            )
                        else:
                            for row in sugerencias:
                                fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                                suggestions_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                                ft.Text(f"👤 Usuario: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                            ], alignment="spaceBetween"),
                                            ft.Text(row["Sugerencia"], color="white", size=14),
                                        ], spacing=5),
                                        bgcolor="#222222",
                                        padding=15,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR AL CARGAR SUGERENCIAS:", ex)
                    suggestions_list.controls.append(ft.Text("Error al cargar las sugerencias de los usuarios.", color="red"))
                page.update()
                
            cargar_sugerencias()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Sugerencias de los Usuarios", size=18, color="white", weight="bold"),
                    ft.IconButton(icon=ft.Icons.REFRESH, tooltip="Recargar sugerencias", on_click=lambda e: cargar_sugerencias())
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Lista de comentarios y propuestas enviados por los usuarios a través del recuadro de la barra lateral.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                suggestions_list
            ], expand=True)

        def build_support_tickets_tab():
            tickets_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cargar_tickets():
                tickets_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("""
                            SELECT t.ID_Ticket, t.Fecha_Hora, u.Nombre_Completo, t.Detalle_Problema, t.Respuesta_Soporte, t.Estatus
                            FROM tickets_soporte t
                            JOIN usuarios u ON t.ID_Usuario = u.ID_Usuario
                            ORDER BY t.Fecha_Hora DESC
                        """)
                        tickets = cursor.fetchall()
                        db.close()
                        
                        if not tickets:
                            tickets_list.controls.append(ft.Text("No hay tickets de soporte registrados.", color="#aaaaaa", size=14))
                        else:
                            for row in tickets:
                                id_t = row["ID_Ticket"]
                                fecha = row["Fecha_Hora"].strftime("%d/%m/%Y %H:%M") if row["Fecha_Hora"] else ""
                                estatus = row["Estatus"]
                                respuesta_t = row["Respuesta_Soporte"] or "(Sin respuesta aún)"
                                
                                is_abierto = estatus == "Abierto"
                                status_color = "#FF4500" if is_abierto else "#7CFC00"
                                
                                # Campo de respuesta de soporte para el admin
                                resp_input = ft.TextField(
                                    label="Escribe la solución...",
                                    value=row["Respuesta_Soporte"] or "",
                                    multiline=True,
                                    min_lines=1,
                                    max_lines=3,
                                    border_color="#9D50BB",
                                    color="white",
                                    text_size=12,
                                    expand=True
                                )
                                
                                def resolver_ticket_click(e, ticket_id=id_t, r_input=resp_input):
                                    solucion = r_input.value.strip()
                                    if not solucion:
                                        mostrar_snack("Por favor escribe una solución antes de resolver.", color="red")
                                        return
                                    try:
                                        db_res = conectar_db()
                                        if db_res:
                                            cursor_res = db_res.cursor()
                                            cursor_res.execute("""
                                                UPDATE tickets_soporte 
                                                SET Estatus = 'Resuelto', Respuesta_Soporte = %s 
                                                WHERE ID_Ticket = %s
                                            """, (solucion, ticket_id))
                                            db_res.commit()
                                            db_res.close()
                                            mostrar_snack("Ticket resuelto con éxito.", color="#7CFC00")
                                            cargar_tickets()
                                            page.update()
                                    except Exception as ex:
                                        print("ERROR RESOLVER TICKET:", ex)
                                
                                tickets_list.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Text(f"📅 {fecha}", color="#aaaaaa", size=11),
                                                ft.Text(f"👤 Reporta: {row['Nombre_Completo']}", color="#aaaaaa", size=11),
                                                ft.Container(
                                                    content=ft.Text(estatus.upper(), color="black", size=9, weight="bold"),
                                                    bgcolor=status_color,
                                                    padding=ft.Padding(left=6, right=6, top=2, bottom=2),
                                                    border_radius=3
                                                )
                                            ], alignment="spaceBetween"),
                                            ft.Text(row["Detalle_Problema"], color="white", size=13),
                                            ft.Divider(height=10, color="#444444"),
                                            ft.Row([
                                                ft.Text("Solución de Soporte:", color="#aaaaaa", size=12, weight="bold"),
                                            ]),
                                            ft.Row([
                                                resp_input,
                                                ft.ElevatedButton(
                                                    "Resolver",
                                                    icon=ft.Icons.CHECK_CIRCLE,
                                                    bgcolor="#7CFC00" if is_abierto else "#444444",
                                                    color="black" if is_abierto else "white",
                                                    on_click=resolver_ticket_click,
                                                    disabled=not is_abierto
                                                )
                                            ], spacing=10) if is_abierto else (
                                                ft.Text(respuesta_t, color="#7CFC00", size=13, italic=True)
                                            )
                                        ], spacing=5),
                                        bgcolor="#222222",
                                        padding=15,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR EN CARGAR TICKETS:", ex)
                    tickets_list.controls.append(ft.Text("Error al cargar los tickets de soporte.", color="red"))
                page.update()
                
            cargar_tickets()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Bandeja de Tickets de Soporte Técnico", size=18, color="white", weight="bold"),
                    ft.IconButton(icon=ft.Icons.REFRESH, tooltip="Recargar tickets", on_click=lambda e: cargar_tickets())
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Visualiza y responde a las fallas o inconsistencias operativas reportadas por los asociados.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                tickets_list
            ], expand=True)

        def build_checklists_view():
            apertura_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            cierre_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            venta_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            progress_apertura = ft.ProgressBar(value=0.0, color="#7CFC00", bgcolor="#222222")
            progress_cierre = ft.ProgressBar(value=0.0, color="#00FFFF", bgcolor="#222222")
            progress_venta = ft.ProgressBar(value=0.0, color="#A100F2", bgcolor="#222222")
            
            text_apertura = ft.Text(f"{t('progress')}: 0%", color="#7CFC00", size=13, weight="bold")
            text_cierre = ft.Text(f"{t('progress')}: 0%", color="#00FFFF", size=13, weight="bold")
            text_venta = ft.Text(f"{t('progress')}: 0%", color="#A100F2", size=13, weight="bold")
            
            def calcular_progreso(categoria, col, p_bar, p_text):
                try:
                    total = 0
                    completados = 0
                    for container in col.controls:
                        if isinstance(container, ft.Container) and container.content:
                            content = container.content
                            chk = None
                            if isinstance(content, ft.Row) and content.controls:
                                # Modo Admin: Row([Checkbox, IconButton])
                                chk = content.controls[0]
                            elif isinstance(content, ft.Checkbox):
                                # Modo Asociado: Checkbox directo
                                chk = content
                            
                            if isinstance(chk, ft.Checkbox):
                                total += 1
                                if chk.value:
                                    completados += 1
                                    
                    val = 0.0
                    if total > 0:
                        val = completados / total
                    p_bar.value = val
                    p_text.value = f"{t('progress')}: {int(val * 100)}% ({completados} {t('of')} {total} {t('completed')})"
                except Exception as ex:
                    print("ERROR CALCULAR PROGRESO CHECKLIST:", ex)
                page.update()

            def toggle_tarea(id_plantilla, completado_val, categoria, col, p_bar, p_text):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        if completado_val:
                            cursor.execute("""
                                INSERT INTO registro_checklist (ID_Usuario, ID_Plantilla, Completado, Fecha, Fecha_Hora)
                                VALUES (%s, %s, 1, CURDATE(), NOW())
                                ON DUPLICATE KEY UPDATE Completado = 1, Fecha_Hora = NOW()
                            """, (user_info["id"], id_plantilla))
                        else:
                            cursor.execute("""
                                DELETE FROM registro_checklist 
                                WHERE ID_Usuario = %s AND ID_Plantilla = %s AND Fecha = CURDATE()
                            """, (user_info["id"], id_plantilla))
                        db.commit()
                        db.close()
                except Exception as ex:
                    print("ERROR TOGGLE TAREA CHECKLIST:", ex)
                calcular_progreso(categoria, col, p_bar, p_text)

            def cargar_checklist_por_categoria(categoria, col, p_bar, p_text):
                col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Plantilla, Descripcion FROM plantillas_checklist WHERE Categoria = %s ORDER BY ID_Plantilla ASC", (categoria,))
                        tareas = cursor.fetchall()
                        
                        cursor.execute("""
                            SELECT ID_Plantilla FROM registro_checklist 
                            WHERE ID_Usuario = %s AND Fecha = CURDATE() AND Completado = 1
                        """, (user_info["id"],))
                        completadas_hoy = {row["ID_Plantilla"] for row in cursor.fetchall()}
                        db.close()
                        
                        if not tareas:
                            col.controls.append(ft.Text(t("no_tasks"), color="#aaaaaa", italic=True))
                        else:
                            for t_item in tareas:
                                id_pl = t_item["ID_Plantilla"]
                                desc = t_item["Descripcion"]
                                esta_completada = id_pl in completadas_hoy
                                
                                checkbox = ft.Checkbox(
                                    label=desc,
                                    value=esta_completada,
                                    label_style=ft.TextStyle(color="white", size=13),
                                    fill_color="#7CFC00" if categoria == 1 else ("#00FFFF" if categoria == 2 else "#A100F2"),
                                    expand=True if es_admin() else False
                                )
                                
                                checkbox.on_change = lambda e, i_p=id_pl, chk=checkbox: toggle_tarea(
                                    i_p, 
                                    chk.value, 
                                    categoria, 
                                    col, 
                                    p_bar, 
                                    p_text
                                )
                                
                                if es_admin():
                                    def make_delete_click(i_p=id_pl):
                                        def delete_item(e):
                                            try:
                                                db_del = conectar_db()
                                                if db_del:
                                                    cursor_del = db_del.cursor()
                                                    cursor_del.execute("DELETE FROM registro_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                    cursor_del.execute("DELETE FROM plantillas_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                    db_del.commit()
                                                    db_del.close()
                                                    mostrar_snack(t("task_deleted"))
                                                    # Reload all checklists to keep UI in sync
                                                    cargar_checklist_por_categoria(1, apertura_list, progress_apertura, text_apertura)
                                                    cargar_checklist_por_categoria(2, cierre_list, progress_cierre, text_cierre)
                                                    cargar_checklist_por_categoria(3, venta_list, progress_venta, text_venta)
                                            except Exception as ex:
                                                print("ERROR ELIMINAR TAREA:", ex)
                                                mostrar_snack("Error", color="red")
                                        return delete_item

                                    container_content = ft.Row([
                                        checkbox,
                                        ft.IconButton(
                                            icon=ft.Icons.DELETE_OUTLINE_ROUNDED,
                                            icon_color="#FF4500",
                                            tooltip=t("lang_label") if False else "Delete task", # using English tooltip directly for simplicity
                                            on_click=make_delete_click()
                                        )
                                    ], alignment="spaceBetween", expand=True)
                                else:
                                    container_content = checkbox

                                col.controls.append(
                                    ft.Container(
                                        content=container_content,
                                        bgcolor="#111111",
                                        padding=10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#222222")
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGAR CHECKLIST POR CATEGORIA:", ex)
                    col.controls.append(ft.Text("Error", color="red"))
                calcular_progreso(categoria, col, p_bar, p_text)

            def build_admin_inline_form(categoria, col, p_bar, p_text):
                input_new_task = ft.TextField(
                    label=t("add_task"),
                    expand=True,
                    border_color="#7CFC00" if categoria == 1 else ("#00FFFF" if categoria == 2 else "#A100F2"),
                    label_style=ft.TextStyle(color="#aaaaaa", size=11),
                    text_style=ft.TextStyle(color="white", size=13),
                    height=45
                )
                
                def agregar_inline_click(e):
                    desc_val = input_new_task.value.strip()
                    if not desc_val:
                        return
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor()
                            cursor.execute("INSERT INTO plantillas_checklist (Categoria, Descripcion) VALUES (%s, %s)", (categoria, desc_val))
                            db.commit()
                            db.close()
                            input_new_task.value = ""
                            mostrar_snack(t("task_added"))
                            # Reload this checklist category
                            cargar_checklist_por_categoria(categoria, col, p_bar, p_text)
                    except Exception as ex:
                        print("ERROR AGREGAR TAREA INLINE:", ex)
                        mostrar_snack("Error", color="red")
                
                btn_add = ft.IconButton(
                    icon=ft.Icons.ADD_CIRCLE_OUTLINE_ROUNDED,
                    icon_color="#7CFC00" if categoria == 1 else ("#00FFFF" if categoria == 2 else "#A100F2"),
                    tooltip=t("add_task"),
                    on_click=agregar_inline_click
                )
                
                return ft.Container(
                    content=ft.Row([
                        input_new_task,
                        btn_add
                    ], spacing=10),
                    margin=ft.Margin(left=0, top=0, right=0, bottom=10)
                )

            cargar_checklist_por_categoria(1, apertura_list, progress_apertura, text_apertura)
            cargar_checklist_por_categoria(2, cierre_list, progress_cierre, text_cierre)
            cargar_checklist_por_categoria(3, venta_list, progress_venta, text_venta)

            tabs_checklist = ft.Tabs(
                selected_index=0,
                animation_duration=300,
                length=3,
                expand=True,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.TabBar(
                            tabs=[
                                ft.Tab(label=t("apertura"), icon=ft.Icons.LIGHT_MODE),
                                ft.Tab(label=t("cierre"), icon=ft.Icons.NIGHTLIGHT_ROUNDED),
                                ft.Tab(label=t("venta"), icon=ft.Icons.MONETIZATION_ON_ROUNDED)
                            ]
                        ),
                        ft.TabBarView(
                            expand=True,
                            controls=[
                                ft.Column([
                                    ft.Divider(height=10, color="transparent"),
                                    text_apertura,
                                    progress_apertura,
                                    ft.Divider(height=10, color="transparent"),
                                    build_admin_inline_form(1, apertura_list, progress_apertura, text_apertura) if es_admin() else ft.Container(),
                                    apertura_list
                                ], expand=True),
                                ft.Column([
                                    ft.Divider(height=10, color="transparent"),
                                    text_cierre,
                                    progress_cierre,
                                    ft.Divider(height=10, color="transparent"),
                                    build_admin_inline_form(2, cierre_list, progress_cierre, text_cierre) if es_admin() else ft.Container(),
                                    cierre_list
                                ], expand=True),
                                ft.Column([
                                    ft.Divider(height=10, color="transparent"),
                                    text_venta,
                                    progress_venta,
                                    ft.Divider(height=10, color="transparent"),
                                    build_admin_inline_form(3, venta_list, progress_venta, text_venta) if es_admin() else ft.Container(),
                                    venta_list
                                ], expand=True)
                            ]
                        )
                    ]
                )
            )

            # Botones del encabezado de checklist
            header_buttons = []
            if es_admin():
                def ir_a_editar_checklists(e):
                    dashboard_tab_index[0] = 5  # Selecciona la pestaña 6: Editar Checklists
                    cambiar_vista("dashboard")
                header_buttons.append(
                    ft.ElevatedButton(
                        t("edit_options"),
                        icon=ft.Icons.EDIT_ROUNDED,
                        bgcolor="#9D50BB",
                        color="white",
                        on_click=ir_a_editar_checklists
                    )
                )
            
            header_buttons.append(
                ft.IconButton(
                    icon=ft.Icons.REFRESH_ROUNDED,
                    tooltip=t("refresh"),
                    on_click=lambda e: [
                        cargar_checklist_por_categoria(1, apertura_list, progress_apertura, text_apertura),
                        cargar_checklist_por_categoria(2, cierre_list, progress_cierre, text_cierre),
                        cargar_checklist_por_categoria(3, venta_list, progress_venta, text_venta)
                    ]
                )
            )

            return ft.Column([
                ft.Row([
                    ft.Text(t("checklist_title"), size=24, color="#D8B4FE", weight="bold"),
                    ft.Row(header_buttons, spacing=10)
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text(t("checklist_desc"), color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                tabs_checklist
            ], expand=True)

        def build_admin_checklist_tab():
            tasks_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            dropdown_cat = ft.Dropdown(
                label="Seleccionar Tipo de Checklist",
                value="1",
                options=[
                    ft.dropdown.Option("1", "Apertura 🌅"),
                    ft.dropdown.Option("2", "Cierre 🌌"),
                    ft.dropdown.Option("3", "Venta Exitosa 💰")
                ],
                width=300,
                border_color="#9D50BB"
            )
            
            input_desc = ft.TextField(
                label="Nueva instrucción de tarea...",
                expand=True,
                border_color="#9D50BB"
            )

            def cargar_tareas_admin(e=None):
                tasks_list.controls.clear()
                cat_val = int(dropdown_cat.value)
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Plantilla, Descripcion FROM plantillas_checklist WHERE Categoria = %s ORDER BY ID_Plantilla DESC", (cat_val,))
                        tareas = cursor.fetchall()
                        db.close()
                        
                        if not tareas:
                            tasks_list.controls.append(ft.Text("No hay tareas registradas en este checklist.", color="#aaaaaa", italic=True))
                        else:
                            for t in tareas:
                                id_pl = t["ID_Plantilla"]
                                desc = t["Descripcion"]
                                
                                def make_eliminar_click(i_p=id_pl, d_t=desc):
                                    def eliminar_tarea_click(ev):
                                        try:
                                            db_del = conectar_db()
                                            if db_del:
                                                cursor_del = db_del.cursor()
                                                cursor_del.execute("DELETE FROM registro_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                cursor_del.execute("DELETE FROM plantillas_checklist WHERE ID_Plantilla = %s", (i_p,))
                                                db_del.commit()
                                                db_del.close()
                                                mostrar_snack(f"Tarea eliminada con éxito.")
                                                cargar_tareas_admin()
                                        except Exception as ex:
                                            print("ERROR ELIMINAR TAREA ADMIN:", ex)
                                            mostrar_snack("Error al eliminar la tarea.", color="red")
                                    return eliminar_tarea_click
                                
                                tasks_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.CHECKLIST_ROUNDED, color="#00FFFF"),
                                            ft.Text(desc, color="white", size=13, expand=True),
                                            ft.IconButton(
                                                icon=ft.Icons.DELETE_FOREVER,
                                                icon_color="#FF4500",
                                                tooltip="Eliminar tarea",
                                                on_click=make_eliminar_click()
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#111111",
                                        padding=10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#222222")
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGAR TAREAS ADMIN:", ex)
                page.update()

            def agregar_tarea_click(e):
                desc_val = input_desc.value.strip()
                if not desc_val:
                    mostrar_snack("Por favor escribe la descripción de la tarea.", color="red")
                    return
                cat_val = int(dropdown_cat.value)
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("INSERT INTO plantillas_checklist (Categoria, Descripcion) VALUES (%s, %s)", (cat_val, desc_val))
                        db.commit()
                        db.close()
                        input_desc.value = ""
                        mostrar_snack("Nueva tarea agregada al checklist.")
                        cargar_tareas_admin()
                except Exception as ex:
                    print("ERROR AGREGAR TAREA ADMIN:", ex)
                    mostrar_snack("Error al guardar la nueva tarea.", color="red")
                page.update()

            dropdown_cat.on_change = cargar_tareas_admin
            cargar_tareas_admin()

            btn_agregar = ft.ElevatedButton(
                "Agregar Tarea",
                icon=ft.Icons.ADD,
                bgcolor="#6E48AA",
                color="white",
                on_click=agregar_tarea_click
            )

            return ft.Column([
                ft.Row([
                    ft.Text("Administración de Checklists Operativos", size=18, color="white", weight="bold"),
                    dropdown_cat
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text("Agrega o elimina tareas específicas del checklist seleccionado. Los cambios se verán reflejados al instante en el portal de los asociados.", color="#aaaaaa", size=13),
                ft.Divider(height=10, color="transparent"),
                ft.Row([
                    input_desc,
                    btn_agregar
                ], spacing=10),
                ft.Divider(height=15, color="#333333"),
                tasks_list
            ], expand=True)

        def build_dashboard_view():
            return ft.Column([
                ft.Text("Panel de Control Operativo", size=24, color="#D8B4FE", weight="bold"),
                ft.Divider(height=20, color="#333333"),
                ft.Tabs(
                    selected_index=dashboard_tab_index[0],
                    on_change=lambda e: dashboard_tab_index.__setitem__(0, e.control.selected_index),
                    animation_duration=300,
                    length=7,
                    expand=True,
                    content=ft.Column(
                        expand=True,
                        controls=[
                            ft.TabBar(
                                tabs=[
                                    ft.Tab(label="Estadísticas", icon=ft.Icons.BAR_CHART),
                                    ft.Tab(label="Preguntas sin Contestar", icon=ft.Icons.QUESTION_MARK_ROUNDED),
                                    ft.Tab(label="Gestión de Manuales", icon=ft.Icons.FOLDER_OPEN_ROUNDED),
                                    ft.Tab(label="Sugerencias", icon=ft.Icons.LIGHTBULB_ROUNDED),
                                    ft.Tab(label="Soporte 🎫", icon=ft.Icons.CONFIRMATION_NUMBER_ROUNDED),
                                    ft.Tab(label="Editar Checklists 📋", icon=ft.Icons.CHECKLIST_ROUNDED),
                                    ft.Tab(label="Tareas Consolidadas 📊", icon=ft.Icons.ASSIGNMENT)
                                ]
                            ),
                            ft.TabBarView(
                                expand=True,
                                controls=[
                                    build_stats_tab(),
                                    build_missing_questions_tab(),
                                    build_manuals_tab(),
                                    build_suggestions_tab(),
                                    build_support_tickets_tab(),
                                    build_admin_checklist_tab(),
                                    build_tareas_admin_tab()
                                ]
                            )
                        ]
                    )
                )
            ], expand=True)

        # --- SISTEMA DE TAREAS OPERATIVAS ---
        
        def verificar_y_cerrar_tareas_vencidas():
            try:
                db = conectar_db()
                if db:
                    cursor = db.cursor()
                    cursor.execute("""
                        UPDATE tareas 
                        SET Estatus = 'Cerrada' 
                        WHERE Estatus = 'Activa' 
                        AND Fecha_Limite IS NOT NULL 
                        AND Fecha_Limite < NOW()
                    """)
                    db.commit()
                    db.close()
            except Exception as e:
                print("Error actualizando tareas vencidas:", e)

        def descargar_consolidado_async(task_id, task_title, nombre_original):
            def thread_target():
                try:
                    db = conectar_db()
                    if not db:
                        mostrar_snack("Error: No se pudo conectar a la base de datos.", color="#FF4500")
                        return
                    cursor = db.cursor(dictionary=True)
                    cursor.execute("SELECT Plantilla_Bytes, Nombre_Plantilla FROM tareas WHERE ID_Tarea = %s", (task_id,))
                    task = cursor.fetchone()
                    if not task or not task["Plantilla_Bytes"]:
                        mostrar_snack("Error: No se encontró la plantilla de esta tarea.", color="#FF4500")
                        db.close()
                        return
                        
                    query = """
                        SELECT r.Tienda, u.Nombre_Completo as Gerente, r.Fecha_Envio, r.Respuestas_JSON 
                        FROM respuestas_tarea r
                        JOIN usuarios u ON r.ID_Usuario = u.ID_Usuario
                        WHERE r.ID_Tarea = %s
                    """
                    params = [task_id]
                    zona_act = active_zone_filter[0]
                    if zona_act != "Todas":
                        query += " AND u.Zona = %s"
                        params.append(zona_act)
                    cursor.execute(query, tuple(params))
                    respuestas = cursor.fetchall()
                    db.close()
                    
                    import io
                    import openpyxl
                    
                    template_bytes = task["Plantilla_Bytes"]
                    f_in = io.BytesIO(template_bytes)
                    wb = openpyxl.load_workbook(f_in)
                    ws = wb.active
                    
                    header_row_idx = None
                    headers = {}
                    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if any(cell is not None for cell in row):
                            header_row_idx = r_idx
                            for c_idx, cell in enumerate(row, start=1):
                                if cell is not None:
                                    headers[str(cell).strip()] = c_idx
                            break
                            
                    if not header_row_idx:
                        mostrar_snack("Error: Plantilla Excel inválida.", color="#FF4500")
                        return
                        
                    next_row = ws.max_row + 1
                    if next_row <= header_row_idx:
                        next_row = header_row_idx + 1
                        
                    for r_item in respuestas:
                        tienda = r_item["Tienda"]
                        gerente = r_item["Gerente"]
                        fecha = r_item["Fecha_Envio"].strftime("%Y-%m-%d %H:%M:%S") if r_item["Fecha_Envio"] else ""
                        vals = json.loads(r_item["Respuestas_JSON"])
                        
                        for h_name, col_idx in headers.items():
                            h_lower = h_name.lower().strip()
                            if any(kw == h_lower for kw in ["tienda", "sucursal"]):
                                ws.cell(row=next_row, column=col_idx, value=tienda)
                            elif any(kw == h_lower for kw in ["gerente", "usuario"]):
                                ws.cell(row=next_row, column=col_idx, value=gerente)
                            elif any(kw in h_lower for kw in ["fecha", "fecha de envio", "fecha_envio"]):
                                ws.cell(row=next_row, column=col_idx, value=fecha)
                            else:
                                val = vals.get(h_name, "")
                                try:
                                    if "." in str(val):
                                        ws.cell(row=next_row, column=col_idx, value=float(val))
                                    else:
                                        ws.cell(row=next_row, column=col_idx, value=int(val))
                                except ValueError:
                                    ws.cell(row=next_row, column=col_idx, value=val)
                        next_row += 1
                        
                    out_buf = io.BytesIO()
                    wb.save(out_buf)
                    consolidated_bytes = out_buf.getvalue()
                    
                    root = Tk()
                    root.withdraw()
                    root.attributes("-topmost", True)
                    
                    default_name = f"Consolidado_{task_title.replace(' ', '_')}.xlsx"
                    ruta_guardar = filedialog.asksaveasfilename(
                        title="Guardar Consolidado Excel",
                        initialfile=default_name,
                        defaultextension=".xlsx",
                        filetypes=[("Archivos de Excel", "*.xlsx")]
                    )
                    root.destroy()
                    
                    if ruta_guardar:
                        with open(ruta_guardar, "wb") as f_out:
                            f_out.write(consolidated_bytes)
                        mostrar_snack(f"Consolidado guardado en: {os.path.basename(ruta_guardar)}", color="#00FF7F")
                except Exception as ex:
                    print("Error al exportar consolidado:", ex)
                    mostrar_snack("Error al exportar reporte consolidado.", color="#FF4500")

            threading.Thread(target=thread_target, daemon=True).start()

        def build_tareas_admin_tab():
            txt_titulo = ft.TextField(label="Título de la Tarea", width=400, border_color="#D8B4FE")
            txt_desc = ft.TextField(label="Descripción / Instrucciones", multiline=True, min_lines=2, max_lines=4, border_color="#D8B4FE")
            txt_fecha_limite = ft.TextField(
                label="Fecha/Hora Límite",
                hint_text="Usa los botones de calendario y reloj",
                width=280,
                border_color="#D8B4FE",
                read_only=True
            )
            
            selected_date = [None]
            selected_time = [None]
            
            def update_limite_text():
                d_str = selected_date[0].strftime("%Y-%m-%d") if selected_date[0] else ""
                t_str = selected_time[0].strftime("%H:%M") if selected_time[0] else ""
                if d_str and t_str:
                    txt_fecha_limite.value = f"{d_str} {t_str}"
                elif d_str:
                    txt_fecha_limite.value = f"{d_str} 00:00"
                else:
                    txt_fecha_limite.value = ""
                page.update()
                
            def date_picked(e):
                if date_picker.value:
                    selected_date[0] = date_picker.value
                    update_limite_text()
                    
            def time_picked(e):
                if time_picker.value:
                    selected_time[0] = time_picker.value
                    update_limite_text()
                    
            date_picker = ft.DatePicker(
                on_change=date_picked,
                help_text="Seleccionar fecha límite"
            )
            time_picker = ft.TimePicker(
                on_change=time_picked,
                help_text="Seleccionar hora límite"
            )
            
            # Evitar duplicados en overlays de la página
            found_dp = False
            found_tp = False
            for ctrl in page.overlay:
                if isinstance(ctrl, ft.DatePicker):
                    date_picker = ctrl
                    date_picker.on_change = date_picked
                    found_dp = True
                elif isinstance(ctrl, ft.TimePicker):
                    time_picker = ctrl
                    time_picker.on_change = time_picked
                    found_tp = True
                    
            if not found_dp:
                page.overlay.append(date_picker)
            if not found_tp:
                page.overlay.append(time_picker)
                
            def show_date_picker(e):
                date_picker.open = True
                page.update()
                
            def show_time_picker(e):
                time_picker.open = True
                page.update()
                
            btn_fecha = ft.IconButton(
                icon=ft.Icons.CALENDAR_MONTH_ROUNDED,
                icon_color="#00FFFF",
                tooltip="Elegir Fecha",
                on_click=show_date_picker
            )
            btn_hora = ft.IconButton(
                icon=ft.Icons.ACCESS_TIME_ROUNDED,
                icon_color="#00FFFF",
                tooltip="Elegir Hora",
                on_click=show_time_picker
            )
            
            fecha_limite_row = ft.Row([
                txt_fecha_limite,
                btn_fecha,
                btn_hora
            ], spacing=5, vertical_alignment="center")
            
            selected_file_path = [None]
            selected_file_bytes = [None]
            selected_columns = [[]]
            
            lbl_plantilla = ft.Text("Ningún archivo Excel seleccionado.", color="#aaaaaa", italic=True)
            
            def on_file_selected(ruta):
                try:
                    with open(ruta, "rb") as f:
                        file_bytes = f.read()
                    import io
                    import openpyxl
                    f_in = io.BytesIO(file_bytes)
                    wb = openpyxl.load_workbook(f_in, data_only=True)
                    ws = wb.active
                    
                    header_row = None
                    for row in ws.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            header_row = [str(cell).strip() for cell in row if cell is not None]
                            break
                            
                    if not header_row:
                        mostrar_snack("Error: El Excel no contiene cabeceras.", color="#FF4500")
                        return
                        
                    omit_keywords = {"tienda", "sucursal", "gerente", "usuario", "fecha", "fecha_envio", "fecha de envio", "hora"}
                    cols = []
                    for col in header_row:
                        col_lower = col.lower().strip()
                        if not any(kw in col_lower for kw in omit_keywords):
                            cols.append(col)
                            
                    selected_file_path[0] = ruta
                    selected_file_bytes[0] = file_bytes
                    selected_columns[0] = cols
                    
                    lbl_plantilla.value = f"Plantilla: {os.path.basename(ruta)}\nColumnas: {', '.join(cols)}"
                    lbl_plantilla.color = "#00FFFF"
                    page.update()
                except Exception as ex:
                    print("Error leyendo excel:", ex)
                    mostrar_snack("Error al leer la plantilla Excel.", color="#FF4500")

            btn_seleccionar = ft.ElevatedButton(
                "Subir Plantilla Excel (.xlsx)",
                icon=ft.Icons.UPLOAD_FILE_ROUNDED,
                color="white",
                bgcolor="#A100F2",
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar Plantilla Excel",
                    [("Archivos de Excel", "*.xlsx")],
                    on_file_selected
                )
            )
            
            def crear_tarea_click(e):
                if not txt_titulo.value.strip():
                    mostrar_snack("El título es obligatorio.", color="#FF4500")
                    return
                if not selected_file_bytes[0]:
                    mostrar_snack("Debes subir una plantilla Excel (.xlsx).", color="#FF4500")
                    return
                    
                limite_val = None
                if txt_fecha_limite.value.strip():
                    try:
                        from datetime import datetime
                        limite_val = datetime.strptime(txt_fecha_limite.value.strip(), "%Y-%m-%d %H:%M")
                    except ValueError:
                        mostrar_snack("Formato de fecha inválido. Usa YYYY-MM-DD HH:MM", color="#FF4500")
                        return
                
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("""
                            INSERT INTO tareas (Titulo, Descripcion, Fecha_Limite, Estatus, Nombre_Plantilla, Plantilla_Bytes, Columnas_JSON, Creado_Por)
                            VALUES (%s, %s, %s, 'Activa', %s, %s, %s, %s)
                        """, (
                            txt_titulo.value.strip(),
                            txt_desc.value.strip(),
                            limite_val,
                            os.path.basename(selected_file_path[0]),
                            selected_file_bytes[0],
                            json.dumps(selected_columns[0]),
                            user_info.get("id")
                        ))
                        db.commit()
                        db.close()
                        
                        # Notificar a las tiendas de la zona activa
                        zona_notif = active_zone_filter[0] if 'active_zone_filter' in locals() or 'active_zone_filter' in globals() else "Todas"
                        crear_notificacion_a_zona(zona_notif, "Nueva Tarea Asignada 📋", f"Nueva tarea: '{txt_titulo.value.strip()}'", "tarea")
                        
                        mostrar_snack("Tarea creada con éxito.", color="#00FF7F")
                        txt_titulo.value = ""
                        txt_desc.value = ""
                        txt_fecha_limite.value = ""
                        selected_file_bytes[0] = None
                        selected_file_path[0] = None
                        selected_columns[0] = []
                        lbl_plantilla.value = "Ningún archivo Excel seleccionado."
                        lbl_plantilla.color = "#aaaaaa"
                        
                        cargar_tareas_admin()
                        page.update()
                except Exception as ex:
                    print("Error al insertar tarea:", ex)
                    mostrar_snack("Error al guardar la tarea en base de datos.", color="#FF4500")

            btn_crear = ft.ElevatedButton(
                "Crear Tarea Operativa",
                icon=ft.Icons.ADD_TASK_ROUNDED,
                color="white",
                bgcolor="#00FFFF",
                on_click=crear_tarea_click
            )
            
            tareas_column = ft.Column(spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cambiar_estatus_tarea(t_id, nuevo_estatus):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("UPDATE tareas SET Estatus = %s WHERE ID_Tarea = %s", (nuevo_estatus, t_id))
                        db.commit()
                        db.close()
                        mostrar_snack(f"Tarea cambiada a: {nuevo_estatus}", color="#00FF7F")
                        cargar_tareas_admin()
                except Exception as ex:
                    print("Error al cambiar estatus:", ex)
                    mostrar_snack("Error al cambiar el estatus de la tarea.", color="#FF4500")
            
            def eliminar_tarea(t_id):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("DELETE FROM tareas WHERE ID_Tarea = %s", (t_id,))
                        db.commit()
                        db.close()
                        mostrar_snack("Tarea eliminada con éxito.", color="#FF4500")
                        cargar_tareas_admin()
                except Exception as ex:
                    print("Error al eliminar tarea:", ex)
                    mostrar_snack("Error al eliminar la tarea.", color="#FF4500")

            def cargar_tareas_admin():
                verificar_y_cerrar_tareas_vencidas()
                tareas_column.controls.clear()
                try:
                    db = conectar_db()
                    if not db:
                        tareas_column.controls.append(ft.Text("Error de conexión a la BD.", color="red"))
                        page.update()
                        return
                    cursor = db.cursor(dictionary=True)
                    cursor.execute("SELECT ID_Tarea, Titulo, Descripcion, Fecha_Limite, Estatus, Nombre_Plantilla FROM tareas ORDER BY ID_Tarea DESC")
                    tareas_res = cursor.fetchall()
                    
                    zona_act = active_zone_filter[0]
                    if zona_act != "Todas":
                        cursor.execute("SELECT ID_Usuario, Nombre_Completo, Tienda FROM usuarios WHERE Rol = 'Gerente' AND Tienda IS NOT NULL AND Tienda != '' AND Zona = %s", (zona_act,))
                    else:
                        cursor.execute("SELECT ID_Usuario, Nombre_Completo, Tienda FROM usuarios WHERE Rol = 'Gerente' AND Tienda IS NOT NULL AND Tienda != ''")
                    gerentes = cursor.fetchall()
                    
                    for task in tareas_res:
                        t_id = task["ID_Tarea"]
                        titulo = task["Titulo"]
                        desc = task["Descripcion"]
                        fecha_lim = task["Fecha_Limite"]
                        estatus = task["Estatus"]
                        plantilla = task["Nombre_Plantilla"]
                        
                        cursor.execute("SELECT ID_Usuario, Tienda, Fecha_Envio FROM respuestas_tarea WHERE ID_Tarea = %s", (t_id,))
                        res_list = cursor.fetchall()
                        res_dict = {r_item["Tienda"]: r_item["Fecha_Envio"] for r_item in res_list}
                        
                        checklist_controls = []
                        checklist_controls.append(ft.Text("Cumplimiento de Tiendas:", size=12, color="#D8B4FE", weight="bold"))
                        
                        if not gerentes:
                            checklist_controls.append(ft.Text("No hay gerentes con tienda asignada en el sistema.", color="#aaaaaa", italic=True, size=11))
                        else:
                            for g in gerentes:
                                tienda_name = g["Tienda"]
                                gerente_name = g["Nombre_Completo"]
                                
                                if tienda_name in res_dict:
                                    fecha_envio_str = res_dict[tienda_name].strftime("%Y-%m-%d %H:%M") if res_dict[tienda_name] else ""
                                    checklist_controls.append(
                                        ft.Row([
                                            ft.Icon(ft.Icons.CHECK_CIRCLE_ROUNDED, color="#00FF7F", size=16),
                                            ft.Text(f"{tienda_name} - {gerente_name} (Entregado: {fecha_envio_str})", color="white", size=11)
                                        ], spacing=5)
                                    )
                                else:
                                    checklist_controls.append(
                                        ft.Row([
                                            ft.Icon(ft.Icons.CANCEL_ROUNDED, color="#FF4500", size=16),
                                            ft.Text(f"{tienda_name} - {gerente_name} (Pendiente)", color="#aaaaaa", size=11)
                                        ], spacing=5)
                                    )
                        
                        sw_active = ft.Switch(
                            label=f"Estatus: {estatus}", 
                            value=(estatus == "Activa"),
                            active_color="#00FFFF",
                            on_change=lambda e, task_id=t_id: cambiar_estatus_tarea(task_id, "Activa" if e.control.value else "Cerrada")
                        )
                        
                        limite_str = fecha_lim.strftime("%Y-%m-%d %H:%M") if fecha_lim else "Sin límite"
                        
                        tareas_column.controls.append(
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([
                                        ft.Text(titulo, size=16, color="white", weight="bold", expand=True),
                                        ft.IconButton(ft.Icons.DELETE_ROUNDED, icon_color="#FF4500", icon_size=18, on_click=lambda e, task_id=t_id: eliminar_tarea(task_id)),
                                    ], alignment="spaceBetween"),
                                    ft.Text(desc, size=13, color="#cccccc"),
                                    ft.Row([
                                        ft.Text(f"📅 Límite: {limite_str}", size=11, color="#aaaaaa"),
                                        ft.Text(f"📎 Plantilla: {plantilla}", size=11, color="#00FFFF"),
                                    ], spacing=15),
                                    ft.Divider(height=1, color="#333333"),
                                    ft.Column(checklist_controls, spacing=4),
                                    ft.Row([
                                        sw_active,
                                        ft.ElevatedButton(
                                            "Descargar Consolidado",
                                            icon=ft.Icons.DOWNLOAD_ROUNDED,
                                            bgcolor="#00FF7F",
                                            color="black",
                                            on_click=lambda e, task_id=t_id, t_title=titulo, p_orig=plantilla: descargar_consolidado_async(task_id, t_title, p_orig)
                                        )
                                    ], alignment="spaceBetween", vertical_alignment="center")
                                ], spacing=10),
                                padding=15,
                                bgcolor="#111111",
                                border_radius=10,
                                border=ft.Border.all(1, "#333333")
                            )
                        )
                    db.close()
                    page.update()
                except Exception as ex:
                    print("Error al cargar tareas:", ex)
                    tareas_column.controls.append(ft.Text("Error al cargar la lista de tareas.", color="red"))
                    page.update()

            cargar_tareas_admin()
            
            return ft.Row([
                ft.Container(
                    content=ft.Column([
                        ft.Text("Crear Nueva Tarea", size=18, color="#D8B4FE", weight="bold"),
                        ft.Divider(height=5, color="#D8B4FE"),
                        txt_titulo,
                        txt_desc,
                        fecha_limite_row,
                        ft.Row([btn_seleccionar, lbl_plantilla], spacing=10, vertical_alignment="center"),
                        ft.Divider(height=10, color="transparent"),
                        btn_crear
                    ], spacing=12, scroll=ft.ScrollMode.ALWAYS),
                    width=450,
                    padding=15,
                    bgcolor="#0d0d0d",
                    border_radius=10,
                    border=ft.Border.all(1, "#222222")
                ),
                ft.Container(
                    content=ft.Column([
                        ft.Text("Monitoreo y Consolidación", size=18, color="#D8B4FE", weight="bold"),
                        ft.Divider(height=5, color="#D8B4FE"),
                        tareas_column
                    ], spacing=12),
                    expand=True,
                    padding=15,
                    bgcolor="#0d0d0d",
                    border_radius=10,
                    border=ft.Border.all(1, "#222222")
                )
            ], spacing=20, expand=True)

        def build_tareas_gerente_view():
            tienda_user = user_info.get("tienda") or ""
            
            tareas_pendientes = ft.Column(spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True)
            tareas_completadas = ft.Column(spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            form_container = ft.Container(visible=False, width=400)
            
            def ver_formulario(task):
                form_container.content = None
                form_container.visible = True
                
                t_id = task["ID_Tarea"]
                titulo = task["Titulo"]
                cols = json.loads(task["Columnas_JSON"])
                
                inputs = {}
                input_controls = []
                
                input_controls.append(ft.Text(f"Reportar: {titulo}", size=18, color="#00FFFF", weight="bold"))
                input_controls.append(ft.Text(task["Descripcion"], size=13, color="#cccccc"))
                input_controls.append(ft.Divider(height=10, color="#333333"))
                
                for c_name in cols:
                    c_lower = c_name.lower().strip()
                    is_number = any(kw in c_lower for kw in ["monto", "venta", "pieza", "cantidad", "total", "precio"])
                    
                    inp = ft.TextField(
                        label=c_name,
                        border_color="#00FFFF",
                        keyboard_type=ft.KeyboardType.NUMBER if is_number else ft.KeyboardType.TEXT,
                        width=350
                    )
                    inputs[c_name] = inp
                    input_controls.append(inp)
                
                def enviar_reporte_click(e):
                    answers = {}
                    for c_name, inp in inputs.items():
                        val = inp.value.strip()
                        if not val:
                            mostrar_snack(f"El campo '{c_name}' es obligatorio.", color="#FF4500")
                            return
                        answers[c_name] = val
                        
                    try:
                        db = conectar_db()
                        if db:
                            cursor = db.cursor()
                            cursor.execute("""
                                INSERT INTO respuestas_tarea (ID_Tarea, ID_Usuario, Tienda, Respuestas_JSON)
                                VALUES (%s, %s, %s, %s)
                            """, (t_id, user_info["id"], tienda_user, json.dumps(answers)))
                            db.commit()
                            db.close()
                            
                            mostrar_snack("Reporte enviado con éxito.", color="#00FF7F")
                            form_container.visible = False
                            cargar_tareas_gerente()
                    except Exception as ex:
                        print("Error al guardar respuesta:", ex)
                        mostrar_snack("Error al guardar el reporte.", color="#FF4500")
                
                btn_enviar = ft.ElevatedButton(
                    "Enviar Reporte",
                    icon=ft.Icons.SEND_ROUNDED,
                    color="white",
                    bgcolor="#00FFFF",
                    on_click=enviar_reporte_click
                )
                btn_cancelar = ft.TextButton(
                    "Cancelar",
                    color="#FF4500",
                    on_click=lambda e: form_container.set_attr("visible", False) or page.update()
                )
                
                input_controls.append(ft.Row([btn_enviar, btn_cancelar], spacing=10))
                
                form_container.content = ft.Container(
                    content=ft.Column(input_controls, spacing=12),
                    padding=20,
                    bgcolor="#0d0d0d",
                    border_radius=10,
                    border=ft.Border.all(1, "#00FFFF")
                )
                page.update()

            def cargar_tareas_gerente():
                verificar_y_cerrar_tareas_vencidas()
                tareas_pendientes.controls.clear()
                tareas_completadas.controls.clear()
                
                if not tienda_user:
                    tareas_pendientes.controls.append(ft.Text("Advertencia: No tienes una tienda asignada en tu perfil. Contacta al Administrador.", color="#FF4500"))
                    page.update()
                    return
                    
                try:
                    db = conectar_db()
                    if not db:
                        tareas_pendientes.controls.append(ft.Text("Error de conexión a la BD.", color="red"))
                        page.update()
                        return
                    cursor = db.cursor(dictionary=True)
                    
                    cursor.execute("SELECT ID_Tarea, Fecha_Envio, Respuestas_JSON FROM respuestas_tarea WHERE Tienda = %s", (tienda_user,))
                    respuestas_enviadas = cursor.fetchall()
                    res_dict = {r["ID_Tarea"]: r for r in respuestas_enviadas}
                    
                    cursor.execute("SELECT ID_Tarea, Titulo, Descripcion, Fecha_Limite, Estatus, Columnas_JSON FROM tareas ORDER BY ID_Tarea DESC")
                    tareas_res = cursor.fetchall()
                    db.close()
                    
                    for task in tareas_res:
                        t_id = task["ID_Tarea"]
                        titulo = task["Titulo"]
                        desc = task["Descripcion"]
                        fecha_lim = task["Fecha_Limite"]
                        estatus = task["Estatus"]
                        
                        limite_str = fecha_lim.strftime("%Y-%m-%d %H:%M") if fecha_lim else "Sin límite"
                        
                        if t_id in res_dict:
                            ans_data = json.loads(res_dict[t_id]["Respuestas_JSON"])
                            fecha_envio_str = res_dict[t_id]["Fecha_Envio"].strftime("%Y-%m-%d %H:%M") if res_dict[t_id]["Fecha_Envio"] else ""
                            
                            resumen_column = ft.Column(spacing=4)
                            resumen_column.controls.append(ft.Text("Datos reportados:", size=12, color="#aaaaaa", weight="bold"))
                            for k, v in ans_data.items():
                                resumen_column.controls.append(ft.Text(f"  • {k}: {v}", size=12, color="white"))
                                
                            tareas_completadas.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(titulo, size=15, color="white", weight="bold"),
                                            ft.Row([
                                                ft.Icon(ft.Icons.CHECK_CIRCLE_ROUNDED, color="#00FF7F", size=16),
                                                ft.Text("Completada", color="#00FF7F", size=12)
                                            ], spacing=5)
                                        ], alignment="spaceBetween"),
                                        ft.Text(desc, size=13, color="#bbbbbb"),
                                        ft.Text(f"Enviado el: {fecha_envio_str}", size=11, color="#aaaaaa"),
                                        ft.Divider(height=1, color="#333333"),
                                        resumen_column
                                    ], spacing=8),
                                    padding=15,
                                    bgcolor="#111111",
                                    border_radius=10,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
                        elif estatus == "Activa":
                            tareas_pendientes.controls.append(
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(titulo, size=15, color="white", weight="bold", expand=True),
                                            ft.ElevatedButton(
                                                "Llenar Reporte",
                                                icon=ft.Icons.EDIT_ROUNDED,
                                                bgcolor="#00FFFF",
                                                color="black",
                                                on_click=lambda e, t=task: ver_formulario(t)
                                            )
                                        ], alignment="spaceBetween"),
                                        ft.Text(desc, size=13, color="#bbbbbb"),
                                        ft.Text(f"📅 Límite de entrega: {limite_str}", size=11, color="#FF8C00")
                                    ], spacing=8),
                                    padding=15,
                                    bgcolor="#111111",
                                    border_radius=10,
                                    border=ft.Border.all(1, "#333333")
                                )
                            )
                            
                    if not tareas_pendientes.controls:
                        tareas_pendientes.controls.append(ft.Text("No tienes tareas pendientes por reportar. ¡Buen trabajo! 🎉", color="#00FF7F", size=13))
                    if not tareas_completadas.controls:
                        tareas_completadas.controls.append(ft.Text("No has completado ninguna tarea todavía.", color="#aaaaaa", italic=True, size=12))
                        
                    page.update()
                except Exception as ex:
                    print("Error al cargar tareas gerente:", ex)
                    tareas_pendientes.controls.append(ft.Text("Error al cargar tareas.", color="red"))
                    page.update()
            
            cargar_tareas_gerente()
            
            return ft.Row([
                ft.Container(
                    content=ft.Column([
                        ft.Text(f"Tareas Operativas — Tienda: {tienda_user}", size=18, color="white", weight="bold"),
                        ft.Divider(height=10, color="#333333"),
                        ft.Tabs(
                            length=2,
                            expand=True,
                            content=ft.Column(
                                expand=True,
                                controls=[
                                    ft.TabBar(
                                        tabs=[
                                            ft.Tab(label="Pendientes 📋", icon=ft.Icons.PENDING_ACTIONS_ROUNDED),
                                            ft.Tab(label="Completadas ✅", icon=ft.Icons.ASSIGNMENT_TURNED_IN_ROUNDED)
                                        ]
                                    ),
                                    ft.TabBarView(
                                        expand=True,
                                        controls=[
                                            tareas_pendientes,
                                            tareas_completadas
                                        ]
                                    )
                                ]
                            )
                        )
                    ], spacing=10),
                    expand=True,
                    padding=15,
                    bgcolor="#080808",
                    border_radius=10,
                    border=ft.Border.all(1, "#222222")
                ),
                form_container
            ], spacing=20, expand=True)

        def build_tareas_view():
            if es_admin():
                return build_tareas_admin_tab()
            else:
                return build_tareas_gerente_view()

        def build_campanas_view():
            if es_admin():
                return build_campanas_admin_view()
            else:
                return build_campanas_gerente_view()

        def build_campanas_admin_view():
            # Estado de fotos guia en creacion de campaña
            # Formato: {"nombre": "...", "instrucciones": "...", "foto_bytes": b"...", "segmento": "Todos", "img_preview": ft.Image}
            guias_creacion = []
            
            # Contenedor para lista de guias en creacion
            guias_col = ft.Column(spacing=10)
            
            nombre_campana = ft.TextField(label="Nombre de la Campaña", border_color="#D8B4FE")
            desc_campana = ft.TextField(label="Instrucciones / Descripción de la Campaña", border_color="#D8B4FE", multiline=True, min_lines=2)
            
            # PDF de la guia
            pdf_guia_bytes = [None]
            pdf_guia_nombre = [None]
            text_pdf_info = ft.Text("No se ha cargado PDF de guía de instalación", color="#aaaaaa", italic=True)
            
            def on_pdf_guia_cargado(path):
                try:
                    import os
                    with open(path, "rb") as f:
                        pdf_guia_bytes[0] = f.read()
                    pdf_guia_nombre[0] = os.path.basename(path)
                    text_pdf_info.value = f"PDF Cargado: {pdf_guia_nombre[0]}"
                    text_pdf_info.color = "#00FF7F"
                    mostrar_snack(f"Guía PDF '{pdf_guia_nombre[0]}' cargada correctamente.", color="#7CFC00")
                    page.update()
                except Exception as ex:
                    print("ERROR CARGANDO PDF GUIA:", ex)
                    mostrar_snack("Error al cargar el archivo PDF.", color="red")

            btn_cargar_pdf_guia = ft.ElevatedButton(
                "Cargar Guía PDF (Opcional)",
                icon=ft.Icons.PICTURE_AS_PDF,
                bgcolor="#9D50BB",
                color="white",
                on_click=lambda e: seleccionar_archivo_async(
                    "Seleccionar PDF de la Guía de Instalación",
                    [("PDF files", "*.pdf")],
                    on_pdf_guia_cargado
                )
            )

            def refrescar_guias_creacion():
                guias_col.controls.clear()
                for i, g in enumerate(guias_creacion):
                    def make_on_click(idx):
                        return lambda e: seleccionar_archivo_async(
                            f"Seleccionar Foto Guía {idx+1}",
                            [("Imágenes", "*.png *.jpg *.jpeg")],
                            lambda path: on_guia_file_selected(idx, path)
                        )
                    
                    def make_on_delete(idx):
                        return lambda e: eliminar_guia_creacion(idx)
                        
                    img_preview = g.get("img_preview")
                    if not img_preview:
                        if g.get("foto_bytes"):
                            import base64
                            img_b64 = base64.b64encode(g["foto_bytes"]).decode("utf-8")
                            img_preview = ft.Image(src=f"data:image/jpeg;base64,{img_b64}", width=120, height=120, fit=ft.ImageFit.CONTAIN)
                            g["img_preview"] = img_preview
                        else:
                            img_preview = ft.Icon(ft.Icons.IMAGE, size=40, color="#555555")
                            
                    dd_guia = ft.Dropdown(
                        label="Formato / Segmento de Tienda",
                        value=g["segmento"],
                        options=[
                            ft.dropdown.Option("Todos", "Todos"),
                            ft.dropdown.Option("Formato 6.000/2.0", "Formato 6.000/2.0"),
                            ft.dropdown.Option("Formato Inline 4.0", "Formato Inline 4.0"),
                            ft.dropdown.Option("Formato Inline Skin", "Formato Inline Skin"),
                            ft.dropdown.Option("Formato Inline Boxes", "Formato Inline Boxes"),
                            ft.dropdown.Option("Formato Open Airs (Kioskos)", "Formato Open Airs (Kioskos)"),
                            ft.dropdown.Option("Formato Inline Skin Kiosko", "Formato Inline Skin Kiosko")
                        ],
                        border_color="#333333",
                        width=350
                    )
                    dd_guia.on_change = lambda e, idx=i: actualizar_guia_campo(idx, "segmento", e.control.value)

                    guias_col.controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Text(f"Foto Guía #{i+1}", weight="bold", color="#00FFFF"),
                                    ft.IconButton(ft.Icons.DELETE, icon_color="#FF4500", on_click=make_on_delete(i))
                                ], alignment="spaceBetween"),
                                ft.Row([
                                    ft.Column([
                                        ft.TextField(
                                            label="Nombre de la Foto (ej. Muro Oakley)",
                                            value=g["nombre"],
                                            border_color="#333333",
                                            on_change=lambda e, idx=i: actualizar_guia_campo(idx, "nombre", e.control.value),
                                            width=350
                                        ),
                                        dd_guia,
                                        ft.TextField(
                                            label="Instrucciones para la IA (ej. Logo centrado, sin espacios vacíos)",
                                            value=g["instrucciones"],
                                            border_color="#333333",
                                            multiline=True,
                                            min_lines=2,
                                            on_change=lambda e, idx=i: actualizar_guia_campo(idx, "instrucciones", e.control.value),
                                            width=350
                                        ),
                                    ], spacing=5, expand=True),
                                    ft.Column([
                                        img_preview,
                                        ft.ElevatedButton(
                                            "Subir Guía",
                                            icon=ft.Icons.UPLOAD,
                                            bgcolor="#D8B4FE",
                                            color="black",
                                            on_click=make_on_click(i)
                                        )
                                    ], horizontal_alignment="center", spacing=5)
                                ], spacing=15)
                            ]),
                            bgcolor="#222222",
                            padding=12,
                            border_radius=8,
                            border=ft.Border.all(1, "#333333")
                        )
                    )
                
                # Botón "+" al final para añadir más guías cómodamente
                guias_col.controls.append(
                    ft.Container(
                        content=ft.Row([
                            ft.IconButton(
                                icon=ft.Icons.ADD_CIRCLE_ROUNDED,
                                icon_color="#00FFFF",
                                icon_size=36,
                                tooltip="Añadir otra Foto Guía",
                                on_click=agregar_guia_creacion
                            ),
                            ft.Text("Añadir otra Foto Guía (+)", color="#00FFFF", weight="bold", size=14)
                        ], alignment="center"),
                        margin=ft.Margin(left=0, top=10, right=0, bottom=10)
                    )
                )
                page.update()

            def actualizar_guia_campo(idx, campo, valor):
                if idx < len(guias_creacion):
                    guias_creacion[idx][campo] = valor

            def on_guia_file_selected(idx, path):
                try:
                    with open(path, "rb") as f:
                        raw_bytes = f.read()
                    enhanced_bytes = optimizar_imagen(raw_bytes)
                    if idx < len(guias_creacion):
                        guias_creacion[idx]["foto_bytes"] = enhanced_bytes
                        guias_creacion[idx]["img_preview"] = None
                        refrescar_guias_creacion()
                        mostrar_snack(f"Foto {idx+1} cargada y optimizada.", color="#7CFC00")
                except Exception as ex:
                    print("ERROR CARGANDO GUIA:", ex)
                    mostrar_snack("Error al cargar la foto.", color="red")

            def eliminar_guia_creacion(idx):
                if idx < len(guias_creacion):
                    guias_creacion.pop(idx)
                    refrescar_guias_creacion()

            def agregar_guia_creacion(e):
                guias_creacion.append({
                    "nombre": "",
                    "instrucciones": "",
                    "segmento": "Todos",
                    "foto_bytes": None,
                    "img_preview": None
                })
                refrescar_guias_creacion()

            def guardar_campana_click(e):
                nom = nombre_campana.value.strip()
                desc = desc_campana.value.strip()
                if not nom:
                    mostrar_snack("Por favor ingrese un nombre de campaña.", color="red")
                    return
                if not guias_creacion:
                    mostrar_snack("Debe añadir al menos una foto guía.", color="red")
                    return
                # Verificar que todas tengan foto y nombre
                for i, g in enumerate(guias_creacion):
                    if not g["nombre"].strip():
                        mostrar_snack(f"La foto guía #{i+1} no tiene nombre.", color="red")
                        return
                    if not g["foto_bytes"]:
                        mostrar_snack(f"La foto guía #{i+1} no tiene imagen.", color="red")
                        return
                
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        # Desactivar otras campañas
                        cursor.execute("UPDATE campanas SET Estatus = 'Inactiva' WHERE Estatus = 'Activa'")
                        # Insertar nueva campaña
                        cursor.execute(
                            "INSERT INTO campanas (Nombre, Descripcion, Estatus, Guia_PDF_Bytes, Guia_PDF_Nombre) VALUES (%s, %s, 'Activa', %s, %s)",
                            (nom, desc, pdf_guia_bytes[0], pdf_guia_nombre[0])
                        )
                        id_campana = cursor.lastrowid
                        
                        # Insertar fotos guia
                        for g in guias_creacion:
                            cursor.execute(
                                "INSERT INTO campana_fotos_guia (ID_Campana, Nombre_Foto, Instrucciones, Imagen_Bytes, Segmento) VALUES (%s, %s, %s, %s, %s)",
                                (id_campana, g["nombre"], g["instrucciones"], g["foto_bytes"], g["segmento"])
                            )
                        db.commit()
                        db.close()
                        
                        # Notificar a todas las tiendas
                        crear_notificacion_a_rol("Gerente", "Nueva Campaña Mensual 📸", f"Se ha activado la campaña: '{nom}'", "campana")
                        
                        nombre_campana.value = ""
                        desc_campana.value = ""
                        pdf_guia_bytes[0] = None
                        pdf_guia_nombre[0] = None
                        text_pdf_info.value = "No se ha cargado PDF de guía de instalación"
                        text_pdf_info.color = "#aaaaaa"
                        guias_creacion.clear()
                        refrescar_guias_creacion()
                        mostrar_snack("¡Campaña guardada y activada con éxito!", color="#7CFC00")
                        # Recargar panel de entregas
                        cargar_entregas_admin()
                except Exception as ex:
                    print("ERROR GUARDANDO CAMPANA:", ex)
                    mostrar_snack("Error de base de datos al guardar campaña.", color="red")

            # --- PANEL DE ENTREGAS ---
            entregas_col = ft.Column(spacing=10)
            detalle_entrega_col = ft.Column(spacing=15)
            
            def cargar_entregas_admin():
                entregas_col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Buscar campaña activa
                        cursor.execute("SELECT ID_Campana, Nombre FROM campanas WHERE Estatus = 'Activa'")
                        campana = cursor.fetchone()
                        if not campana:
                            entregas_col.controls.append(ft.Text("No hay ninguna campaña activa actualmente.", color="#aaaaaa", italic=True))
                            db.close()
                            page.update()
                            return
                        
                        id_campana = campana["ID_Campana"]
                        entregas_col.controls.append(ft.Text(f"Campaña Activa: {campana['Nombre']}", size=14, color="#D8B4FE", weight="bold"))
                        
                        # Obtener todas las entregas de esta campaña
                        query = """
                            SELECT e.ID_Entrega, e.Tienda, e.Fecha_Envio, e.Estatus, u.Nombre_Completo, u.Segmento as Segmento_Tienda
                            FROM campana_entregas_tienda e
                            JOIN usuarios u ON e.ID_Usuario = u.ID_Usuario
                            WHERE e.ID_Campana = %s
                        """
                        params = [id_campana]
                        zona_act = active_zone_filter[0]
                        if zona_act != "Todas":
                            query += " AND u.Zona = %s"
                            params.append(zona_act)
                        query += " ORDER BY e.Fecha_Envio DESC"
                        
                        cursor.execute(query, tuple(params))
                        entregas = cursor.fetchall()
                        db.close()
                        
                        if not entregas:
                            entregas_col.controls.append(ft.Text("Ninguna tienda ha enviado fotos todavía.", color="#aaaaaa", italic=True))
                        else:
                            for ent in entregas:
                                est_color = "#FF4500" if ent["Estatus"] == "Rechazado_IA" else ("#00FF7F" if ent["Estatus"] == "Visto_Bueno" else "#FFD700")
                                status_badge = ft.Container(
                                    content=ft.Text(ent["Estatus"].upper().replace("_", " "), size=10, weight="bold", color="black"),
                                    bgcolor=est_color,
                                    padding=ft.Padding(left=10, right=10, top=5, bottom=5),
                                    border_radius=4
                                )
                                
                                def make_view_details(id_ent, tienda_name):
                                    return lambda e: ver_detalle_entrega_admin(id_ent, tienda_name)
                                    
                                format_text = ent['Segmento_Tienda'] or "Sin Segmento"
                                entregas_col.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Column([
                                                ft.Text(f"Tienda: {ent['Tienda']} ({format_text})", weight="bold", color="white"),
                                                ft.Text(f"Enviado por: {ent['Nombre_Completo']} - {ent['Fecha_Envio']}", size=12, color="#aaaaaa")
                                            ], spacing=2, expand=True),
                                            status_badge,
                                            ft.IconButton(ft.Icons.CHEVRON_RIGHT, icon_color="#00FFFF", on_click=make_view_details(ent["ID_Entrega"], ent["Tienda"]))
                                        ], alignment="spaceBetween"),
                                        bgcolor="#222222",
                                        padding=12,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGANDO ENTREGAS ADMIN:", ex)
                    entregas_col.controls.append(ft.Text("Error al cargar las entregas.", color="red"))
                page.update()

            def ver_detalle_entrega_admin(id_entrega, tienda_name):
                detalle_entrega_col.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Obtener fotos entregadas por la tienda y su respectiva foto guia
                        cursor.execute("""
                            SELECT f.ID_Foto_Tienda, f.Estatus_Auditoria, f.Resultado_IA, f.Imagen_Bytes as Foto_Tienda,
                                   g.Nombre_Foto, g.Instrucciones, g.Imagen_Bytes as Foto_Guia, g.Segmento as Segmento_Foto
                            FROM campana_fotos_tienda f
                            JOIN campana_fotos_guia g ON f.ID_Foto_Guia = g.ID_Foto_Guia
                            WHERE f.ID_Entrega = %s
                        """, (id_entrega,))
                        fotos = cursor.fetchall()
                        
                        # Obtener estatus de la entrega
                        cursor.execute("SELECT Estatus FROM campana_entregas_tienda WHERE ID_Entrega = %s", (id_entrega,))
                        entrega_row = cursor.fetchone()
                        db.close()
                        
                        detalle_entrega_col.controls.append(
                            ft.Row([
                                ft.IconButton(ft.Icons.ARROW_BACK, icon_color="#00FFFF", on_click=lambda e: volver_a_lista_entregas()),
                                ft.Text(f"Detalle de Entrega - {tienda_name}", size=16, color="#00FFFF", weight="bold")
                            ], spacing=10)
                        )
                        
                        if not fotos:
                            detalle_entrega_col.controls.append(ft.Text("No hay fotos en esta entrega.", color="#aaaaaa", italic=True))
                        else:
                            for f in fotos:
                                # Imagen de guia y de tienda en base64
                                import base64
                                img_guia_b64 = base64.b64encode(f["Foto_Guia"]).decode("utf-8")
                                img_tienda_b64 = base64.b64encode(f["Foto_Tienda"]).decode("utf-8")
                                
                                card_border_color = "#00FF7F" if f["Estatus_Auditoria"] == "Aprobado" else ("#FF4500" if f["Estatus_Auditoria"] == "Corregir" else "#333333")
                                
                                detail_card = ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(f"Sección: {f['Nombre_Foto']}", size=14, color="#D8B4FE", weight="bold"),
                                            ft.Container(
                                                content=ft.Text(f"Segmento: {f['Segmento_Foto']}", size=9, color="black", weight="bold"),
                                                bgcolor="#00FFFF",
                                                padding=3,
                                                border_radius=3
                                            )
                                        ], alignment="spaceBetween"),
                                        ft.Text(f"Instrucciones: {f['Instrucciones']}", size=12, color="#aaaaaa"),
                                        ft.Row([
                                            ft.Column([
                                                ft.Text("FOTO GUÍA", size=10, color="#aaaaaa", weight="bold"),
                                                ft.Image(src=f"data:image/jpeg;base64,{img_guia_b64}", width=200, height=150, fit=ft.ImageFit.CONTAIN)
                                            ], horizontal_alignment="center"),
                                            ft.Column([
                                                ft.Text("FOTO TIENDA", size=10, color="#aaaaaa", weight="bold"),
                                                ft.Image(src=f"data:image/jpeg;base64,{img_tienda_b64}", width=200, height=150, fit=ft.ImageFit.CONTAIN)
                                            ], horizontal_alignment="center")
                                        ], spacing=20, alignment="center"),
                                        ft.Divider(height=10, color="#333333"),
                                        ft.Text(f"Estatus IA: {f['Estatus_Auditoria'].upper()}", color="#00FF7F" if f['Estatus_Auditoria'] == 'Aprobado' else "#FF4500", weight="bold", size=12),
                                        ft.Text(f"Análisis de IA:\n{f['Resultado_IA'] or 'Sin revisión.'}", size=12, color="white")
                                    ], spacing=10),
                                    bgcolor="#222222",
                                    padding=15,
                                    border_radius=8,
                                    border=ft.Border.all(1.5, card_border_color)
                                )
                                detalle_entrega_col.controls.append(detail_card)
                                
                            # Botón de visto bueno
                            if entrega_row and entrega_row["Estatus"] != "Visto_Bueno":
                                def on_visto_bueno_click(e, ent_id=id_entrega, t_name=tienda_name):
                                    dar_visto_bueno_entrega(ent_id, t_name)
                                    
                                detalle_entrega_col.controls.append(
                                    ft.Row([
                                        ft.ElevatedButton(
                                            "Dar Visto Bueno Zonal 👑",
                                            icon=ft.Icons.CHECK_CIRCLE,
                                            bgcolor="#00FF7F",
                                            color="black",
                                            on_click=on_visto_bueno_click
                                        )
                                    ], alignment="center")
                                )
                            else:
                                detalle_entrega_col.controls.append(
                                    ft.Row([
                                        ft.Container(
                                            content=ft.Row([
                                                ft.Icon(ft.Icons.CHECK_CIRCLE, color="#00FF7F"),
                                                ft.Text("Esta entrega tiene el Visto Bueno del Jefe Zonal", color="#00FF7F", weight="bold")
                                            ], spacing=5),
                                            padding=10,
                                            bgcolor="#112211",
                                            border_radius=8,
                                            border=ft.Border.all(1, "#00FF7F")
                                        )
                                    ], alignment="center")
                                )
                        
                        entregas_tabs.selected_index = 1 # Ir a la pestaña de entregas
                        entregas_col.visible = False
                        detalle_entrega_col.visible = True
                except Exception as ex:
                    print("ERROR MOSTRANDO DETALLE ENTREGA ADMIN:", ex)
                    detalle_entrega_col.controls.append(ft.Text("Error al cargar detalles de la entrega.", color="red"))
                page.update()

            def volver_a_lista_entregas():
                detalle_entrega_col.visible = False
                entregas_col.visible = True
                cargar_entregas_admin()

            def dar_visto_bueno_entrega(id_entrega, tienda_name):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        cursor.execute("UPDATE campana_entregas_tienda SET Estatus = 'Visto_Bueno' WHERE ID_Entrega = %s", (id_entrega,))
                        db.commit()
                        
                        # Obtener el ID_Usuario de la entrega y el Nombre de la campaña para enviar la notificación
                        cursor.execute("""
                            SELECT e.ID_Usuario, c.Nombre 
                            FROM campana_entregas_tienda e
                            JOIN campanas c ON e.ID_Campana = c.ID_Campana
                            WHERE e.ID_Entrega = %s
                        """, (id_entrega,))
                        row_ent = cursor.fetchone()
                        if row_ent:
                            id_gerente = row_ent[0]
                            camp_name = row_ent[1]
                            crear_notificacion(id_gerente, "Visto Bueno Otorgado 👑", f"Tu entrega de campaña '{camp_name}' ha recibido el visto bueno final.", "campana")
                            
                        db.close()
                        mostrar_snack(f"Visto Bueno otorgado para {tienda_name}.", color="#7CFC00")
                        volver_a_lista_entregas()
                except Exception as ex:
                    print("ERROR DANDO VISTO BUENO:", ex)
                    mostrar_snack("Error al guardar estatus.", color="red")

            # --- CONFIGURACIÓN GEMINI KEY ---
            api_key_input = ft.TextField(
                label="Gemini API Key",
                value=GEMINI_API_KEY,
                password=True,
                can_reveal_password=True,
                border_color="#D8B4FE",
                width=450
            )
            
            def guardar_gemini_key_click(e):
                global GEMINI_API_KEY
                k = api_key_input.value.strip()
                if not k:
                    mostrar_snack("Por favor ingrese una clave válida.", color="red")
                    return
                if guardar_config_key("gemini_api_key", k):
                    GEMINI_API_KEY = k
                    mostrar_snack("API Key de Gemini guardada correctamente.", color="#7CFC00")
                else:
                    mostrar_snack("Error al guardar la clave en config.json.", color="red")

            def depurar_fotos_viejas_click(e):
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        # Count how many would be affected
                        cursor.execute("""
                            SELECT COUNT(*) 
                            FROM campana_fotos_tienda ft
                            JOIN campana_entregas_tienda et ON ft.ID_Entrega = et.ID_Entrega
                            WHERE et.Fecha_Envio < DATE_SUB(NOW(), INTERVAL 3 MONTH)
                              AND ft.Imagen_Bytes IS NOT NULL
                        """)
                        filas_a_depurar = cursor.fetchone()[0]
                        
                        if filas_a_depurar == 0:
                            mostrar_snack("No hay imágenes de más de 3 meses para depurar.", color="#00FFFF")
                            db.close()
                            return
                            
                        # Perform the update
                        cursor.execute("""
                            UPDATE campana_fotos_tienda ft
                            JOIN campana_entregas_tienda et ON ft.ID_Entrega = et.ID_Entrega
                            SET ft.Imagen_Bytes = NULL
                            WHERE et.Fecha_Envio < DATE_SUB(NOW(), INTERVAL 3 MONTH)
                        """)
                        filas_depuradas = cursor.rowcount
                        
                        # Optimize table
                        cursor.execute("OPTIMIZE TABLE campana_fotos_tienda")
                        cursor.fetchall() # Consume results of OPTIMIZE TABLE
                        
                        db.commit()
                        db.close()
                        mostrar_snack(f"Mantenimiento exitoso: Se eliminaron {filas_depuradas} fotos antiguas. Base de datos optimizada.", color="#7CFC00")
                except Exception as ex:
                    print("ERROR DEPURANDO ALMACENAMIENTO:", ex)
                    mostrar_snack("Error al ejecutar la depuración de base de datos.", color="red")

            config_key_view = ft.Column([
                ft.Text("Configuración de IA de Visión (Gemini)", size=16, color="#00FFFF", weight="bold"),
                ft.Text("La API Key se guarda localmente en el archivo config.json para autorizar las solicitudes a Gemini 1.5 Flash.", color="#aaaaaa", size=13),
                ft.Row([
                    api_key_input,
                    ft.ElevatedButton(
                        "Guardar Clave 💾",
                        bgcolor="#9D50BB",
                        color="white",
                        on_click=guardar_gemini_key_click
                    )
                ], spacing=10),
                ft.Divider(height=15, color="#333333"),
                ft.Container(
                    content=ft.Column([
                        ft.Text("Mantenimiento y Almacenamiento 🧹", size=16, color="#00FFFF", weight="bold"),
                        ft.Text("Depura el almacenamiento de base de datos liberando espacio ocupado por imágenes binarias de campañas con más de 3 meses de antigüedad. Se conserva la metadata y las auditorías de IA para el historial.", color="#aaaaaa", size=13),
                        ft.ElevatedButton(
                            "Liberar Almacenamiento (Fotos > 3 Meses) 🧹",
                            icon=ft.Icons.CLEANING_SERVICES_ROUNDED,
                            bgcolor="#FF4500",
                            color="white",
                            on_click=depurar_fotos_viejas_click
                        )
                    ], spacing=10),
                    padding=15,
                    bgcolor="#111111",
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                )
            ], spacing=10)

            # Evitar error de content/tabs en Tabs constructor usando TabBar y TabBarView
            entregas_tabs = ft.Tabs(
                selected_index=0,
                length=3,
                expand=True,
                content=ft.Column(
                    expand=True,
                    controls=[
                        ft.TabBar(
                            tabs=[
                                ft.Tab(label="Crear Campaña 📸", icon=ft.Icons.ADD_A_PHOTO),
                                ft.Tab(label="Revisar Entregas 📋", icon=ft.Icons.CHECKLIST),
                                ft.Tab(label="Configuración IA ⚙", icon=ft.Icons.SETTINGS)
                            ]
                        ),
                        ft.TabBarView(
                            expand=True,
                            controls=[
                                # Tab 1 content: Crear Campaña
                                ft.Column([
                                    nombre_campana,
                                    desc_campana,
                                    ft.Row([
                                        btn_cargar_pdf_guia,
                                        text_pdf_info
                                    ], spacing=15, vertical_alignment="center"),
                                    ft.Row([
                                        ft.Text("Secciones / Fotos requeridas de la Campaña", size=14, color="#D8B4FE", weight="bold"),
                                        ft.ElevatedButton(
                                            "Añadir Foto Guía",
                                            icon=ft.Icons.ADD,
                                            bgcolor="#00FFFF",
                                            color="black",
                                            on_click=agregar_guia_creacion
                                        )
                                    ], alignment="spaceBetween", vertical_alignment="center"),
                                    guias_col,
                                    ft.Divider(height=15, color="#333333"),
                                    ft.Row([
                                        ft.ElevatedButton(
                                            "Activar y Guardar Campaña 💾",
                                            icon=ft.Icons.SAVE,
                                            bgcolor="#00FF7F",
                                            color="black",
                                            height=45,
                                            on_click=guardar_campana_click
                                        )
                                    ], alignment="center")
                                ], spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True),
                                
                                # Tab 2 content: Revisar Entregas
                                ft.Column([
                                    entregas_col,
                                    detalle_entrega_col
                                ], spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True),
                                
                                # Tab 3 content: Configuración IA
                                ft.Column([
                                    config_key_view
                                ], spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True)
                            ]
                        )
                    ]
                )
            )
            
            # Cargar guías iniciales y entregas
            agregar_guia_creacion(None)
            cargar_entregas_admin()
            detalle_entrega_col.visible = False
            
            return ft.Column([
                ft.Row([
                    ft.Text("Fotos de Campaña — Administrador", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Define las fotos guía del mes para las exhibiciones de Sunglass Hut y audita las entregas de las tiendas.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                entregas_tabs
            ], expand=True)

        def build_campanas_gerente_view():
            gerente_campana_col = ft.Column(spacing=15, scroll=ft.ScrollMode.ALWAYS, expand=True)
            u_id = user_info.get("id")
            
            # Cargar el segmento y la zona del usuario desde la BD si no están en user_info
            if "segmento" not in user_info or "zona" not in user_info:
                try:
                    db_u = conectar_db()
                    if db_u:
                        cur_u = db_u.cursor(dictionary=True)
                        cur_u.execute("SELECT Segmento, Zona FROM usuarios WHERE ID_Usuario = %s", (u_id,))
                        user_row = cur_u.fetchone()
                        db_u.close()
                        if user_row:
                            user_info["segmento"] = user_row["Segmento"] if user_row["Segmento"] else "Todos"
                            user_info["zona"] = user_row["Zona"] if user_row["Zona"] else "Zona Centro"
                        else:
                            user_info["segmento"] = "Todos"
                            user_info["zona"] = "Zona Centro"
                except Exception as ex_u:
                    print("ERROR CARGANDO SEGMENTO/ZONA USUARIO:", ex_u)
                    user_info["segmento"] = "Todos"
                    user_info["zona"] = "Zona Centro"

            segmento_actual = user_info.get("segmento") or "Todos"
            zona_actual = user_info.get("zona") or "Zona Centro"

            def cambiar_segmento_gerente(e):
                nuevo_seg = e.control.value
                user_info["segmento"] = nuevo_seg
                try:
                    db_seg = conectar_db()
                    if db_seg:
                        cursor_seg = db_seg.cursor()
                        cursor_seg.execute("UPDATE usuarios SET Segmento = %s WHERE ID_Usuario = %s", (nuevo_seg, u_id))
                        db_seg.commit()
                        db_seg.close()
                        mostrar_snack(f"Segmento de tienda actualizado a: {nuevo_seg}", color="#7CFC00")
                except Exception as ex:
                    print("ERROR ACTUALIZANDO SEGMENTO GERENTE:", ex)
                cargar_campana_gerente()

            def cambiar_zona_gerente(e):
                nueva_zona = e.control.value
                user_info["zona"] = nueva_zona
                try:
                    db_z = conectar_db()
                    if db_z:
                        cursor_z = db_z.cursor()
                        cursor_z.execute("UPDATE usuarios SET Zona = %s WHERE ID_Usuario = %s", (nueva_zona, u_id))
                        db_z.commit()
                        db_z.close()
                        mostrar_snack(f"Zona de tienda actualizada a: {nueva_zona}", color="#7CFC00")
                except Exception as ex:
                    print("ERROR ACTUALIZANDO ZONA GERENTE:", ex)
                cargar_campana_gerente()

            dropdown_segmento = ft.Dropdown(
                label="Formato / Segmento de tu Tienda",
                value=segmento_actual,
                options=[
                    ft.dropdown.Option("Todos", "Todos"),
                    ft.dropdown.Option("Formato 6.000/2.0", "Formato 6.000/2.0"),
                    ft.dropdown.Option("Formato Inline 4.0", "Formato Inline 4.0"),
                    ft.dropdown.Option("Formato Inline Skin", "Formato Inline Skin"),
                    ft.dropdown.Option("Formato Inline Boxes", "Formato Inline Boxes"),
                    ft.dropdown.Option("Formato Open Airs (Kioskos)", "Formato Open Airs (Kioskos)"),
                    ft.dropdown.Option("Formato Inline Skin Kiosko", "Formato Inline Skin Kiosko")
                ],
                border_color="#00FFFF",
                width=350
            )
            dropdown_segmento.on_change = cambiar_segmento_gerente

            dropdown_zona = ft.Dropdown(
                label="Zona / Región de tu Tienda",
                value=zona_actual,
                options=[
                    ft.dropdown.Option("Zona Norte", "Zona Norte"),
                    ft.dropdown.Option("Zona Sur", "Zona Sur"),
                    ft.dropdown.Option("Zona Occidente", "Zona Occidente"),
                    ft.dropdown.Option("Zona Centro", "Zona Centro"),
                    ft.dropdown.Option("Palacio de Hierro", "Palacio de Hierro")
                ],
                border_color="#00FFFF",
                width=350
            )
            dropdown_zona.on_change = cambiar_zona_gerente

            def abrir_pdf_campana(id_camp):
                try:
                    db_p = conectar_db()
                    if not db_p:
                        return
                    cursor_p = db_p.cursor(dictionary=True)
                    cursor_p.execute("SELECT Guia_PDF_Nombre, Guia_PDF_Bytes FROM campanas WHERE ID_Campana = %s", (id_camp,))
                    row = cursor_p.fetchone()
                    db_p.close()
                    if row and row["Guia_PDF_Bytes"]:
                        import tempfile
                        ruta_temp = os.path.join(tempfile.gettempdir(), row["Guia_PDF_Nombre"])
                        with open(ruta_temp, "wb") as f_pdf:
                            f_pdf.write(row["Guia_PDF_Bytes"])
                        import os
                        os.startfile(ruta_temp)
                        mostrar_snack(f"Abriendo PDF de la campaña: {row['Guia_PDF_Nombre']}", color="#7CFC00")
                    else:
                        mostrar_snack("No hay archivo PDF cargado para esta campaña.", color="#FF4500")
                except Exception as ex:
                    print("ERROR ABRIR PDF CAMPANA:", ex)
                    mostrar_snack("Error al abrir el archivo PDF.", color="red")
            
            def cargar_campana_gerente():
                gerente_campana_col.controls.clear()
                
                # Verificar que el gerente tenga tienda asignada
                t_nombre = user_info.get("tienda")
                if not t_nombre:
                    gerente_campana_col.controls.append(
                        ft.Text("Advertencia: No tienes una tienda asignada en tu perfil. Contacta al Administrador para poder subir tus fotos de campaña.", color="#FF4500", weight="bold")
                    )
                    page.update()
                    return
                
                # Renderizar selector de formato y zona
                gerente_campana_col.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Text("Región / Zona de tu Tienda:", weight="bold"),
                                dropdown_zona
                            ], alignment="spaceBetween", vertical_alignment="center"),
                            ft.Row([
                                ft.Text("Filtro de Guías por Formato:", weight="bold"),
                                dropdown_segmento
                            ], alignment="spaceBetween", vertical_alignment="center")
                        ], spacing=10),
                        bgcolor="#1e1e1e",
                        padding=15,
                        border_radius=8
                    )
                )

                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        # Buscar campaña activa
                        cursor.execute("SELECT ID_Campana, Nombre, Descripcion, Guia_PDF_Nombre FROM campanas WHERE Estatus = 'Activa'")
                        campana = cursor.fetchone()
                        
                        if not campana:
                            gerente_campana_col.controls.append(
                                ft.Text("No hay ninguna campaña mensual activa en este momento. Vuelve más tarde.", color="#aaaaaa", italic=True)
                            )
                            db.close()
                            page.update()
                            return
                        
                        id_campana = campana["ID_Campana"]
                        
                        # Obtener entrega de esta tienda o crearla
                        cursor.execute("""
                            SELECT ID_Entrega, Estatus FROM campana_entregas_tienda
                            WHERE ID_Campana = %s AND Tienda = %s
                        """, (id_campana, t_nombre))
                        entrega = cursor.fetchone()
                        if not entrega:
                            cursor.execute("""
                                INSERT INTO campana_entregas_tienda (ID_Campana, Tienda, ID_Usuario, Estatus)
                                VALUES (%s, %s, %s, 'Pendiente')
                            """, (id_campana, t_nombre, u_id))
                            db.commit()
                            id_entrega = cursor.lastrowid
                            entrega_status = "Pendiente"
                        else:
                            id_entrega = entrega["ID_Entrega"]
                            entrega_status = entrega["Estatus"]
                            
                        # Obtener fotos guías de la campaña filtrando por el segmento seleccionado o Todos
                        seg_filtro = user_info.get("segmento") or "Todos"
                        cursor.execute("""
                            SELECT ID_Foto_Guia, Nombre_Foto, Instrucciones, Imagen_Bytes, Segmento FROM campana_fotos_guia
                            WHERE ID_Campana = %s AND (Segmento = 'Todos' OR Segmento = %s)
                            ORDER BY ID_Foto_Guia
                        """, (id_campana, seg_filtro))
                        guias = cursor.fetchall()
                        
                        # Obtener fotos subidas por la tienda en esta entrega
                        cursor.execute("""
                            SELECT ID_Foto_Tienda, ID_Foto_Guia, Imagen_Bytes, Estatus_Auditoria, Resultado_IA FROM campana_fotos_tienda
                            WHERE ID_Entrega = %s
                        """, (id_entrega,))
                        fotos_tienda = {f["ID_Foto_Guia"]: f for f in cursor.fetchall()}
                        db.close()
                        
                        # PDF de la guia
                        header_row_widgets = [
                            ft.Text(f"Campaña Activa: {campana['Nombre']}", size=18, color="#00FFFF", weight="bold"),
                        ]
                        if campana.get("Guia_PDF_Nombre"):
                            btn_ver_pdf = ft.ElevatedButton(
                                "Ver Guía de Instalación PDF 📄",
                                icon=ft.Icons.PICTURE_AS_PDF,
                                bgcolor="#9D50BB",
                                color="white",
                                on_click=lambda e, id_c=id_campana: abrir_pdf_campana(id_c)
                            )
                            header_row_widgets.append(btn_ver_pdf)
                            
                        header_row_widgets.append(
                            ft.Container(
                                content=ft.Text(f"ESTATUS: {entrega_status.upper().replace('_', ' ')}", size=10, weight="bold", color="black"),
                                bgcolor="#00FF7F" if entrega_status == "Visto_Bueno" else ("#FFD700" if entrega_status == "Aprobado_IA" else "#FF4500"),
                                padding=5,
                                border_radius=4
                            )
                        )

                        # UI Encabezado
                        gerente_campana_col.controls.append(
                            ft.Row(header_row_widgets, alignment="spaceBetween")
                        )
                        if campana["Descripcion"]:
                            gerente_campana_col.controls.append(ft.Text(campana["Descripcion"], size=13, color="#cccccc"))
                        gerente_campana_col.controls.append(ft.Divider(height=10, color="#333333"))
                        
                        if not guias:
                            gerente_campana_col.controls.append(
                                ft.Text(f"No hay fotos guía configuradas para tu segmento ({seg_filtro}) o para todos.", color="#aaaaaa", italic=True)
                            )
                        else:
                            # Renderizar cada guía
                            for g in guias:
                                id_g = g["ID_Foto_Guia"]
                                nom_foto = g["Nombre_Foto"]
                                instrucciones = g["Instrucciones"]
                                seg_guia = g["Segmento"]
                                
                                import base64
                                img_guia_b64 = base64.b64encode(g["Imagen_Bytes"]).decode("utf-8")
                                
                                subida = fotos_tienda.get(id_g)
                                
                                # Construir interfaz de esta foto
                                tienda_img_widget = None
                                status_txt = "Pendiente de subir"
                                status_color = "#aaaaaa"
                                audit_feedback = ""
                                
                                if subida:
                                    img_tienda_b64 = base64.b64encode(subida["Imagen_Bytes"]).decode("utf-8")
                                    tienda_img_widget = ft.Image(src=f"data:image/jpeg;base64,{img_tienda_b64}", width=180, height=135, fit=ft.ImageFit.CONTAIN)
                                    est = subida["Estatus_Auditoria"]
                                    if est == "Aprobado":
                                        status_txt = "APROBADO POR IA"
                                        status_color = "#00FF7F"
                                    elif est == "Corregir":
                                        status_txt = "CORREGIR (Ver observaciones abajo)"
                                        status_color = "#FF4500"
                                    else:
                                        status_txt = "REVISANDO CON IA..."
                                        status_color = "#FFD700"
                                        
                                    if subida["Resultado_IA"]:
                                        audit_feedback = subida["Resultado_IA"]
                                else:
                                    tienda_img_widget = ft.Container(
                                        content=ft.Icon(ft.Icons.NO_PHOTOGRAPHY, size=40, color="#555555"),
                                        width=180,
                                        height=135,
                                        bgcolor="#1e1e1e",
                                        alignment=ft.alignment.center,
                                        border_radius=5
                                    )
                                    
                                card_border = ft.Border.all(1.5, "#00FF7F" if status_txt.startswith("APROBADO") else ("#FF4500" if status_txt.startswith("CORREGIR") else "#333333"))
                                
                                def make_on_upload(g_id=id_g, ent_id=id_entrega):
                                    return lambda e: seleccionar_archivo_async(
                                        f"Subir Foto para {nom_foto}",
                                        [("Imágenes", "*.png *.jpg *.jpeg")],
                                        lambda path: subir_foto_tienda_gerente(path, g_id, ent_id)
                                    )
                                    
                                gerente_campana_col.controls.append(
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Row([
                                                ft.Row([
                                                    ft.Text(nom_foto, size=15, color="#D8B4FE", weight="bold"),
                                                    ft.Container(
                                                        content=ft.Text(f"Segmento: {seg_guia}", size=8, color="black", weight="bold"),
                                                        bgcolor="#00FFFF",
                                                        padding=2,
                                                        border_radius=2
                                                    )
                                                ], spacing=10),
                                                ft.Container(
                                                    content=ft.Text(status_txt, size=9, weight="bold", color="black"),
                                                    bgcolor=status_color,
                                                    padding=3,
                                                    border_radius=3
                                                )
                                            ], alignment="spaceBetween"),
                                            ft.Text(f"Instrucciones de Montaje: {instrucciones}", size=12, color="#aaaaaa"),
                                            ft.Row([
                                                ft.Column([
                                                    ft.Text("FOTO GUÍA DE MONTAJE", size=9, color="#aaaaaa", weight="bold"),
                                                    ft.Image(src=f"data:image/jpeg;base64,{img_guia_b64}", width=180, height=135, fit=ft.ImageFit.CONTAIN)
                                                ], horizontal_alignment="center"),
                                                ft.Column([
                                                    ft.Text("FOTO REAL DE TU TIENDA", size=9, color="#aaaaaa", weight="bold"),
                                                    tienda_img_widget
                                                ], horizontal_alignment="center")
                                            ], spacing=20, alignment="center"),
                                            ft.Row([
                                                ft.ElevatedButton(
                                                    "Subir Foto" if not subida else "Volver a subir",
                                                    icon=ft.Icons.UPLOAD_FILE,
                                                    bgcolor="#00FFFF",
                                                    color="black",
                                                    on_click=make_on_upload()
                                                )
                                            ], alignment="center"),
                                            ft.Column([
                                                ft.Text("Análisis de IA de Visión:", size=11, color="#aaaaaa", weight="bold"),
                                                ft.Text(audit_feedback, size=11, color="white")
                                            ], spacing=3, visible=bool(audit_feedback))
                                        ], spacing=10),
                                        bgcolor="#222222",
                                        padding=15,
                                        border_radius=8,
                                        border=card_border
                                    )
                                )
                except Exception as ex:
                    print("ERROR CARGANDO VISTA GERENTE CAMPANA:", ex)
                    gerente_campana_col.controls.append(ft.Text("Error al cargar la campaña activa.", color="red"))
                page.update()

            def subir_foto_tienda_gerente(file_path, id_guia, id_entrega):
                try:
                    with open(file_path, "rb") as f:
                        raw_bytes = f.read()
                    
                    # Optimizar imagen
                    img_optimized = optimizar_imagen(raw_bytes)
                    
                    mostrar_snack("Foto subida. Iniciando auditoría con IA...", color="#00FFFF")
                    
                    # Guardar foto en la base de datos con estatus temporal 'Auditando'
                    db = conectar_db()
                    if db:
                        cursor = db.cursor()
                        # Verificar si ya existe un registro para esta foto guia
                        cursor.execute("""
                            SELECT ID_Foto_Tienda FROM campana_fotos_tienda
                            WHERE ID_Entrega = %s AND ID_Foto_Guia = %s
                        """, (id_entrega, id_guia))
                        row = cursor.fetchone()
                        
                        if row:
                            id_foto_tienda = row[0]
                            cursor.execute("""
                                UPDATE campana_fotos_tienda
                                SET Imagen_Bytes = %s, Estatus_Auditoria = 'Auditando', Resultado_IA = 'Revisando imagen con IA de visión...'
                                WHERE ID_Foto_Tienda = %s
                            """, (img_optimized, id_foto_tienda))
                        else:
                            cursor.execute("""
                                INSERT INTO campana_fotos_tienda (ID_Entrega, ID_Foto_Guia, Imagen_Bytes, Estatus_Auditoria, Resultado_IA)
                                VALUES (%s, %s, %s, 'Auditando', 'Revisando imagen con IA de visión...')
                            """, (id_entrega, id_guia, img_optimized))
                        
                        db.commit()
                        db.close()
                        
                        # Notificar al Administrador de la entrega de fotos
                        crear_notificacion_a_rol("Administrador", "Nueva Foto de Campaña 📸", f"La tienda '{t_nombre}' ha subido una foto para revisión.", "campana")
                        
                    # Refrescar UI antes de llamar a Gemini
                    cargar_campana_gerente()
                    
                    # Lanzar auditoría en hilo separado para no bloquear la UI
                    def thread_auditoria():
                        try:
                            # 1. Recuperar fotos guía e instrucciones de la BD
                            db_aud = conectar_db()
                            if db_aud:
                                cursor_aud = db_aud.cursor(dictionary=True)
                                cursor_aud.execute("""
                                    SELECT Imagen_Bytes, Instrucciones, Nombre_Foto FROM campana_fotos_guia
                                    WHERE ID_Foto_Guia = %s
                                """, (id_guia,))
                                guia_row = cursor_aud.fetchone()
                                db_aud.close()
                                
                                if guia_row:
                                    guia_bytes = guia_row["Imagen_Bytes"]
                                    instrucciones = guia_row["Instrucciones"]
                                    nombre_foto = guia_row["Nombre_Foto"]
                                    
                                    # 2. Llamar a la IA
                                    resultado_ia = auditar_foto_con_gemini(guia_bytes, img_optimized, instrucciones)
                                    
                                    # 3. Determinar estatus según la primera palabra
                                    resultado_limpio = resultado_ia.strip()
                                    if resultado_limpio.upper().startswith("APROBADO"):
                                        estatus_final = "Aprobado"
                                    elif resultado_limpio.upper().startswith("CORREGIR"):
                                        estatus_final = "Corregir"
                                    else:
                                        # Buscar palabras clave si no empieza exactamente
                                        if "APROBADO" in resultado_limpio.upper()[:15]:
                                            estatus_final = "Aprobado"
                                        else:
                                            estatus_final = "Corregir"
                                            
                                    # 4. Actualizar en base de datos
                                    db_upd = conectar_db()
                                    if db_upd:
                                        cursor_upd = db_upd.cursor()
                                        cursor_upd.execute("""
                                            UPDATE campana_fotos_tienda
                                            SET Estatus_Auditoria = %s, Resultado_IA = %s, Fecha_Auditoria = CURRENT_TIMESTAMP
                                            WHERE ID_Entrega = %s AND ID_Foto_Guia = %s
                                        """, (estatus_final, resultado_limpio, id_entrega, id_guia))
                                        
                                        # Comprobar si todas las fotos de la entrega están aprobadas para actualizar la entrega a 'Aprobado_IA'
                                        cursor_upd.execute("""
                                            SELECT COUNT(*) FROM campana_fotos_guia g
                                            WHERE g.ID_Campana = (SELECT ID_Campana FROM campana_entregas_tienda WHERE ID_Entrega = %s)
                                        """, (id_entrega,))
                                        total_requeridas = cursor_upd.fetchone()[0]
                                        
                                        cursor_upd.execute("""
                                            SELECT COUNT(*) FROM campana_fotos_tienda
                                            WHERE ID_Entrega = %s AND Estatus_Auditoria = 'Aprobado'
                                        """, (id_entrega,))
                                        total_aprobadas = cursor_upd.fetchone()[0]
                                        
                                        if total_aprobadas >= total_requeridas:
                                            cursor_upd.execute("""
                                                UPDATE campana_entregas_tienda
                                                SET Estatus = 'Aprobado_IA'
                                                WHERE ID_Entrega = %s AND Estatus != 'Visto_Bueno'
                                            """, (id_entrega,))
                                        else:
                                            cursor_upd.execute("""
                                                UPDATE campana_entregas_tienda
                                                SET Estatus = 'Rechazado_IA'
                                                WHERE ID_Entrega = %s AND Estatus != 'Visto_Bueno'
                                            """, (id_entrega,))
                                            
                                        db_upd.commit()
                                        db_upd.close()
                                        
                                        # Notificar al gerente de la sucursal sobre la auditoría IA
                                        crear_notificacion(u_id, "Auditoría IA de Campaña 🤖", f"La sección '{nombre_foto}' ha sido calificada como: {estatus_final.upper()}", "campana")
                                        
                                    # Notificar y refrescar
                                    mostrar_snack("Auditoría de IA completada.", color="#7CFC00" if estatus_final == "Aprobado" else "#FF4500")
                                    cargar_campana_gerente()
                        except Exception as ex_t:
                            print("ERROR EN THREAD AUDITORIA:", ex_t)
                            mostrar_snack("Error en proceso de auditoría con la IA.", color="red")
                            
                    threading.Thread(target=thread_auditoria, daemon=True).start()
                    
                except Exception as ex:
                    print("ERROR SUBIENDO FOTO TIENDA:", ex)
                    mostrar_snack("Error al guardar la foto.", color="red")
                    
            cargar_campana_gerente()
            
            return ft.Column([
                ft.Row([
                    ft.Text("Fotos de Campaña — Tiendas", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Sube las fotos de exhibición de tu tienda y deja que el auditor de IA valide el montaje según las guías.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                gerente_campana_col
            ], expand=True)

        def build_manuals_view():
            manuals_list = ft.Column(spacing=10, scroll=ft.ScrollMode.ALWAYS, expand=True)
            
            def cargar_manuales():
                manuals_list.controls.clear()
                try:
                    db = conectar_db()
                    if db:
                        cursor = db.cursor(dictionary=True)
                        cursor.execute("SELECT ID_Manual, Nombre_Archivo, Titulo, Version FROM manuales ORDER BY Nombre_Archivo")
                        manuales = cursor.fetchall()
                        db.close()

                        manuals_list.controls.append(ft.Text(t("manuals_db_title"), size=14, color="#00FFFF", weight="bold"))
                        if not manuales:
                            manuals_list.controls.append(ft.Text(t("no_manuals"), color="#aaaaaa", size=12))
                        else:
                            for m in manuales:
                                id_m = m["ID_Manual"]
                                nombre = m.get("Nombre_Archivo") or ""
                                version = m.get("Version") or ""
                                titulo = m.get("Titulo") or ""

                                # Botones de Visualizar y Descargar
                                manuals_list.controls.append(
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.INSERT_DRIVE_FILE, color="#00FFFF"),
                                            ft.Column([
                                                ft.Text(nombre, color="white", weight="bold", size=14),
                                                ft.Text(f"{t('version')}: {version} | {titulo}", color="#aaaaaa", size=11)
                                            ], spacing=3, expand=True),
                                            ft.ElevatedButton(
                                                t("view_pdf"),
                                                on_click=lambda e, id_man=id_m: visualizar_pdf(id_man, page),
                                                bgcolor="#6E48AA",
                                                color="white"
                                            ),
                                            ft.ElevatedButton(
                                                t("download_pdf"),
                                                on_click=lambda e, id_man=id_m: descargar_pdf_archivo(id_man, page),
                                                bgcolor="#444444",
                                                color="white"
                                            )
                                        ], alignment="spaceBetween", vertical_alignment="center"),
                                        bgcolor="#222222",
                                        padding=10,
                                        border_radius=8,
                                        border=ft.Border.all(1, "#333333")
                                    )
                                )
                except Exception as ex:
                    print("ERROR MANUALS VIEW LIST:", ex)
                    manuals_list.controls.append(ft.Text("Error", color="red"))
                page.update()
                
            cargar_manuales()
            
            return ft.Column([
                ft.Row([
                    ft.Text(t("manuals_title"), size=24, color="#D8B4FE", weight="bold"),
                    ft.IconButton(ft.Icons.REFRESH, on_click=lambda e: cargar_manuales(), icon_color="#00FFFF")
                ], alignment="spaceBetween", vertical_alignment="center"),
                ft.Text(t("manuals_desc"), color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                manuals_list
            ], expand=True)

        # =================================
        # DISEÑO DEL DASHBOARD (BARRA LATERAL Y CONTENIDO DINÁMICO)
        # =================================

        content_area = ft.Container(
            expand=True,
            padding=20,
            bgcolor="#000000"
        )

        # Avatar de usuario en perfil
        profile_icon = ft.Container(
            content=ft.Image(src=img_usuario, width=40, height=40, fit=ft.controls.box.BoxFit.COVER) if img_usuario else ft.Icon(ft.Icons.PERSON, color="#00FFFF", size=24),
            width=40,
            height=40,
            border_radius=20,
            bgcolor="#333333",
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
            alignment=ft.alignment.Alignment(0, 0),
            border=ft.Border.all(1.5, "#D8B4FE"),
        )

        # --- SISTEMA DE ALERTAS (CAMPANITA) ---
        bell_icon_container = ft.Container()

        def mostrar_notificaciones_dialog(e):
            u_id = user_info.get("id")
            notifs = cargar_notificaciones(u_id)
            marcar_notificaciones_leidas(u_id)
            actualizar_campana_badge() # Limpiar badge
            
            notif_rows = []
            if not notifs:
                notif_rows.append(ft.Text("No tienes notificaciones recientes.", color="#aaaaaa", italic=True))
            else:
                for n in notifs:
                    icon_map = {
                        "tarea": ft.Icons.ASSIGNMENT_ROUNDED,
                        "manual": ft.Icons.BOOK_ROUNDED,
                        "campana": ft.Icons.PHOTO_CAMERA,
                        "sistema": ft.Icons.INFO_ROUNDED
                    }
                    icon_color_map = {
                        "tarea": "#00FFFF",
                        "manual": "#D8B4FE",
                        "campana": "#7CFC00",
                        "sistema": "#FFD700"
                    }
                    tipo = n.get("Tipo") or "sistema"
                    fecha = n.get("Fecha_Hora").strftime("%d/%m %H:%M") if n.get("Fecha_Hora") else ""
                    leida = n.get("Leida")
                    
                    notif_rows.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(icon_map.get(tipo, ft.Icons.INFO_ROUNDED), color=icon_color_map.get(tipo, "#00FFFF"), size=20),
                                ft.Column([
                                    ft.Text(n.get("Titulo") or "", color="white", weight="bold", size=12),
                                    ft.Text(n.get("Mensaje") or "", color="#cccccc", size=11),
                                    ft.Text(fecha, color="#888888", size=9)
                                ], spacing=2, expand=True),
                                ft.Container(
                                    width=8,
                                    height=8,
                                    bgcolor="#00FFFF" if not leida else "transparent",
                                    border_radius=4
                                )
                            ], vertical_alignment="center", spacing=10),
                            bgcolor="#1e1e1e" if not leida else "#111111",
                            padding=10,
                            border_radius=6,
                            border=ft.Border.all(1, "#333333" if leida else "#00FFFF")
                        )
                    )
            
            def cerrar_notif_dialog(e):
                page.pop_dialog()
            
            dlg = ft.AlertDialog(
                title=ft.Text("Campana de Alertas 🔔", color="#00FFFF", weight="bold"),
                content=ft.Container(
                    content=ft.Column(notif_rows, spacing=8, scroll=ft.ScrollMode.ALWAYS),
                    width=420,
                    height=380
                ),
                actions=[
                    ft.TextButton("Cerrar", on_click=cerrar_notif_dialog)
                ],
                actions_alignment="end",
                bgcolor="#111111"
            )
            page.show_dialog(dlg)

        def actualizar_campana_badge():
            u_id = user_info.get("id")
            unread_cnt = obtener_cantidad_notificaciones_sin_leer(u_id)
            
            stack_controls = [
                ft.IconButton(
                    icon=ft.Icons.NOTIFICATIONS,
                    icon_color="#00FFFF",
                    tooltip="Notificaciones 🔔",
                    on_click=mostrar_notificaciones_dialog
                )
            ]
            if unread_cnt > 0:
                stack_controls.append(
                    ft.Container(
                        content=ft.Text(str(unread_cnt), color="white", size=9, weight="bold"),
                        bgcolor="red",
                        width=16,
                        height=16,
                        border_radius=8,
                        alignment=ft.alignment.Alignment(0, 0),
                        margin=ft.Margin(left=22, top=4, right=0, bottom=0)
                    )
                )
            bell_icon_container.content = ft.Stack(stack_controls, width=40, height=40)
            try:
                page.update()
            except Exception:
                pass

        profile_row = ft.Row([
            profile_icon,
            ft.Column([
                ft.Text(user_info["nombre"], color="white", weight="bold", size=14, overflow=ft.TextOverflow.ELLIPSIS),
                ft.Text(user_info["rol"], color="#aaaaaa", size=12),
                ft.Text(f"🌐 IP: {getattr(page, 'client_ip', None) or 'Localhost'}", color="#00FFFF", size=10)
            ], spacing=2, expand=True),
            bell_icon_container
        ], spacing=10)
        def build_presupuesto_view():
            import datetime
            import calendar

            hoy = datetime.date.today()
            current_month = [hoy.month]
            current_year = [hoy.year]

            selected_zona = [user_info.get("zona") or "Zona Centro"]
            selected_tienda = [user_info.get("tienda") or ""]

            tiendas_por_zona = {}
            try:
                db_t = conectar_db()
                if db_t:
                    cur_t = db_t.cursor(dictionary=True)
                    cur_t.execute("SELECT DISTINCT Tienda, Zona FROM usuarios WHERE Tienda IS NOT NULL AND Tienda != '' ORDER BY Tienda ASC")
                    for row in cur_t.fetchall():
                        z = row["Zona"] or "Sin Zona"
                        t_val = row["Tienda"]
                        if z not in tiendas_por_zona:
                            tiendas_por_zona[z] = []
                        if t_val not in tiendas_por_zona[z]:
                            tiendas_por_zona[z].append(t_val)
                    db_t.close()
            except Exception as e_db:
                print("Error loading tiendas list:", e_db)

            # Salvaguarda: Asegurar que tiendas_por_zona nunca esté vacío
            if not tiendas_por_zona:
                tiendas_por_zona["Sin Zona"] = ["Sin Tienda"]

            # Si es admin, determinar selected_zona y selected_tienda
            if es_admin():
                active_z = active_zone_filter[0] if active_zone_filter[0] != "Todas" else "Zona Centro"
                if active_z not in tiendas_por_zona:
                    active_z = list(tiendas_por_zona.keys())[0]
                selected_zona[0] = active_z
                if active_z in tiendas_por_zona and tiendas_por_zona[active_z]:
                    if selected_tienda[0] not in tiendas_por_zona[active_z]:
                        selected_tienda[0] = tiendas_por_zona[active_z][0]
            else:
                # Gerente
                selected_tienda[0] = user_info.get("tienda") or ""
                # Encontrar a qué zona pertenece esta tienda en los datos
                found_zone = "Sin Zona"
                for z, t_list in tiendas_por_zona.items():
                    if selected_tienda[0] in t_list:
                        found_zone = z
                        break
                selected_zona[0] = found_zone

            # Asegurar que el valor inicial exista en las opciones del Dropdown para evitar crash de renderizado de Flet
            if selected_zona[0] not in tiendas_por_zona:
                selected_zona[0] = list(tiendas_por_zona.keys())[0]
            
            zona_tiendas = tiendas_por_zona.get(selected_zona[0], [])
            if not selected_tienda[0] or selected_tienda[0] not in zona_tiendas:
                selected_tienda[0] = zona_tiendas[0] if zona_tiendas else ""

            meta_venta_tf = ft.TextField(
                label="Meta Venta (Sin IVA) 💰",
                value="",
                border_color="#9D50BB",
                focused_border_color="#00FFFF",
                color="white",
                text_size=13,
                height=45,
                expand=True,
                keyboard_type=ft.KeyboardType.NUMBER
            )
            meta_piezas_tf = ft.TextField(
                label="Meta Piezas 📦",
                value="",
                border_color="#9D50BB",
                focused_border_color="#00FFFF",
                color="white",
                text_size=13,
                height=45,
                expand=True,
                keyboard_type=ft.KeyboardType.NUMBER
            )

            progress_bar_venta = ft.ProgressBar(value=0.0, color="#FF4B4B", bgcolor="#222222", height=10, border_radius=5)
            progress_text_venta = ft.Text("Venta: 0% ($0.00 / $0.00 sin IVA)", color="white", size=12)
            
            progress_bar_piezas = ft.ProgressBar(value=0.0, color="#FF4B4B", bgcolor="#222222", height=10, border_radius=5)
            progress_text_piezas = ft.Text("Piezas: 0% (0 / 0 pzs)", color="white", size=12)

            meses_logrados_col = ft.Column(spacing=4, scroll=ft.ScrollMode.ALWAYS, height=180)

            calendar_grid = ft.Column(spacing=10, expand=True)

            tienda_title_txt = ft.Text("", size=18, color="#00FFFF", weight="bold")
            zona_title_txt = ft.Text("", size=12, color="#aaaaaa")
            period_title_txt = ft.Text("", size=16, color="white", weight="bold")

            dd_zona = ft.Dropdown(
                label="Zona",
                value=selected_zona[0],
                border_color="#00FFFF",
                focused_border_color="#00FFFF",
                color="white",
                text_size=12,
                height=45,
                width=160,
                options=[ft.dropdown.Option(z, z) for z in tiendas_por_zona.keys()]
            )

            initial_tiendas = tiendas_por_zona.get(selected_zona[0], [])
            dd_tienda = ft.Dropdown(
                label="Tienda",
                value=selected_tienda[0] if selected_tienda[0] else None,
                border_color="#00FFFF",
                focused_border_color="#00FFFF",
                color="white",
                text_size=12,
                height=45,
                width=200,
                options=[ft.dropdown.Option(t, t) for t in initial_tiendas]
            )

            meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            dd_mes = ft.Dropdown(
                label="Mes",
                value=str(current_month[0]),
                border_color="#9D50BB",
                color="white",
                text_size=12,
                height=45,
                width=120,
                options=[ft.dropdown.Option(str(i+1), meses_nombres[i]) for i in range(12)]
            )

            dd_anio = ft.Dropdown(
                label="Año",
                value=str(current_year[0]),
                border_color="#9D50BB",
                color="white",
                text_size=12,
                height=45,
                width=100,
                options=[ft.dropdown.Option(str(y), str(y)) for y in [2025, 2026, 2027]]
            )

            def cargar_datos_presupuesto():
                tienda_actual = selected_tienda[0]
                mes_actual = current_month[0]
                anio_actual = current_year[0]
                
                if not tienda_actual:
                    return 0.0, 0, []
                
                meta_venta = 0.0
                meta_piezas = 0
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT Meta_Venta, Meta_Piezas 
                            FROM presupuesto_mensual 
                            WHERE Tienda = %s AND Mes = %s AND Anio = %s
                        """, (tienda_actual, mes_actual, anio_actual))
                        row_meta = cur.fetchone()
                        if row_meta:
                            meta_venta = float(row_meta["Meta_Venta"] or 0.0)
                            meta_piezas = int(row_meta["Meta_Piezas"] or 0)
                        
                        cur.execute("""
                            SELECT DAY(Fecha) as Dia, Venta_Con_IVA, Venta_Sin_IVA, Piezas 
                            FROM presupuesto_diario 
                            WHERE Tienda = %s AND MONTH(Fecha) = %s AND YEAR(Fecha) = %s
                        """, (tienda_actual, mes_actual, anio_actual))
                        ventas_diarias = cur.fetchall()
                        db.close()
                        return meta_venta, meta_piezas, ventas_diarias
                except Exception as ex:
                    print("Error loading budget data:", ex)
                
                return 0.0, 0, []

            def open_edit_day_dialog(dia):
                tienda_actual = selected_tienda[0]
                mes_actual = current_month[0]
                anio_actual = current_year[0]
                
                fecha_str = f"{anio_actual:04d}-{mes_actual:02d}-{dia:02d}"
                
                existing_venta_con_iva = 0.0
                existing_piezas = 0
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT Venta_Con_IVA, Piezas 
                            FROM presupuesto_diario 
                            WHERE Tienda = %s AND Fecha = %s
                        """, (tienda_actual, fecha_str))
                        row = cur.fetchone()
                        if row:
                            existing_venta_con_iva = float(row["Venta_Con_IVA"] or 0.0)
                            existing_piezas = int(row["Piezas"] or 0)
                        db.close()
                except Exception as ex:
                    print("Error loading existing day values:", ex)
                
                venta_dia_tf = ft.TextField(
                    label="Venta del día con IVA ($)",
                    value=str(existing_venta_con_iva) if existing_venta_con_iva > 0 else "",
                    border_color="#9D50BB",
                    focused_border_color="#00FFFF",
                    color="white",
                    keyboard_type=ft.KeyboardType.NUMBER
                )
                piezas_dia_tf = ft.TextField(
                    label="Piezas vendidas",
                    value=str(existing_piezas) if existing_piezas > 0 else "",
                    border_color="#9D50BB",
                    focused_border_color="#00FFFF",
                    color="white",
                    keyboard_type=ft.KeyboardType.NUMBER
                )
                
                def guardar_dia_click(e):
                    try:
                        v_con_iva = float(venta_dia_tf.value.strip() or 0.0)
                        p_dia = int(piezas_dia_tf.value.strip() or 0)
                    except ValueError:
                        mostrar_snack("Por favor ingresa números válidos.", color="red")
                        return
                    
                    v_sin_iva = v_con_iva / 1.16
                    
                    try:
                        db = conectar_db()
                        if db:
                            cur = db.cursor()
                            cur.execute("""
                                INSERT INTO presupuesto_diario (Tienda, Fecha, Venta_Con_IVA, Venta_Sin_IVA, Piezas)
                                VALUES (%s, %s, %s, %s, %s)
                                ON DUPLICATE KEY UPDATE 
                                    Venta_Con_IVA = %s,
                                    Venta_Sin_IVA = %s,
                                    Piezas = %s
                            """, (tienda_actual, fecha_str, v_con_iva, v_sin_iva, p_dia, v_con_iva, v_sin_iva, p_dia))
                            db.commit()
                            db.close()
                            
                            page.pop_dialog()
                            mostrar_snack(f"Día {dia} guardado exitosamente.", color="#7CFC00")
                            refresh_data()
                    except Exception as ex:
                        print("Error saving day details:", ex)
                        mostrar_snack("Error al guardar en base de datos.", color="red")
                
                dlg = ft.AlertDialog(
                    title=ft.Text(f"Registrar Venta - Día {dia}", color="#00FFFF", weight="bold"),
                    content=ft.Column([
                        ft.Text(f"Tienda: {tienda_actual}", color="#aaaaaa", size=12),
                        ft.Text(f"Fecha: {fecha_str}", color="#aaaaaa", size=12),
                        ft.Container(height=10),
                        venta_dia_tf,
                        piezas_dia_tf
                    ], tight=True, spacing=10),
                    actions=[
                        ft.TextButton("Cancelar", on_click=lambda e: page.pop_dialog()),
                        ft.ElevatedButton("Guardar 💾", bgcolor="#9D50BB", color="white", on_click=guardar_dia_click)
                    ],
                    actions_alignment="end",
                    bgcolor="#111111"
                )
                page.show_dialog(dlg)

            def render_meses_logrados():
                meses_logrados_col.controls.clear()
                tienda_actual = selected_tienda[0]
                anio_actual = current_year[0]
                if not tienda_actual:
                    meses_logrados_col.controls.append(ft.Text("Selecciona una tienda", color="#aaaaaa", italic=True, size=12))
                    return
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor(dictionary=True)
                        cur.execute("""
                            SELECT 
                                m_list.Mes,
                                m.Meta_Venta,
                                m.Meta_Piezas,
                                COALESCE(SUM(d.Venta_Sin_IVA), 0) as Venta_Lograda,
                                COALESCE(SUM(d.Piezas), 0) as Piezas_Logradas
                            FROM (
                                SELECT 1 as Mes UNION SELECT 2 UNION SELECT 3 UNION SELECT 4 
                                UNION SELECT 5 UNION SELECT 6 UNION SELECT 7 UNION SELECT 8 
                                UNION SELECT 9 UNION SELECT 10 UNION SELECT 11 UNION SELECT 12
                            ) m_list
                            LEFT JOIN presupuesto_mensual m ON m.Mes = m_list.Mes AND m.Tienda = %s AND m.Anio = %s
                            LEFT JOIN presupuesto_diario d ON MONTH(d.Fecha) = m_list.Mes AND YEAR(d.Fecha) = %s AND d.Tienda = %s
                            GROUP BY m_list.Mes, m.Meta_Venta, m.Meta_Piezas
                            ORDER BY m_list.Mes ASC
                        """, (tienda_actual, anio_actual, anio_actual, tienda_actual))
                        rows = cur.fetchall()
                        db.close()
                        
                        has_any = False
                        for row in rows:
                            m_idx = row["Mes"]
                            meta_v = float(row["Meta_Venta"] or 0.0)
                            venta_log = float(row["Venta_Lograda"] or 0.0)
                            
                            if meta_v > 0.0:
                                meta_v_sin = meta_v
                                v_pct = (venta_log / meta_v_sin) * 100 if meta_v_sin > 0 else 0.0
                                lograda = venta_log >= meta_v_sin
                                
                                has_any = True
                                icon_color = "#00FF7F" if lograda else "#FFCC00"
                                icon_name = ft.Icons.CHECK_CIRCLE_OUTLINE_ROUNDED if lograda else ft.Icons.RADIO_BUTTON_UNCHECKED_OUTROUNDED
                                status_txt = "LOGRADO" if lograda else f"{v_pct:.0f}%"
                                
                                meses_logrados_col.controls.append(
                                    ft.Row([
                                        ft.Icon(icon_name, color=icon_color, size=16),
                                        ft.Text(f"{meses_nombres[m_idx-1]} ({status_txt})", color="white" if lograda else "#cccccc", size=12, weight="bold" if lograda else "normal"),
                                    ], spacing=5)
                                )
                        
                        if not has_any:
                            meses_logrados_col.controls.append(ft.Text("Ninguna meta de ventas definida en este año.", color="#aaaaaa", italic=True, size=11))
                except Exception as ex:
                    print("Error in render_meses_logrados:", ex)
                    meses_logrados_col.controls.append(ft.Text("Error al cargar logros.", color="red", size=12))

            def render_calendar(daily_accum_map):
                calendar_grid.controls.clear()
                
                days_headers = ["DOM", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB"]
                header_row = ft.Row(
                    [
                        ft.Container(
                            content=ft.Text(h, color="#D8B4FE", weight="bold", size=10, text_align="center"),
                            expand=True,
                            alignment=ft.alignment.Alignment(0, 0),
                            padding=5
                        ) for h in days_headers
                    ],
                    spacing=5
                )
                calendar_grid.controls.append(header_row)
                
                year = current_year[0]
                month = current_month[0]
                first_weekday_py, num_days = calendar.monthrange(year, month)
                start_offset = (first_weekday_py + 1) % 7
                
                cells = []
                
                for _ in range(start_offset):
                    cells.append(
                        ft.Container(
                            expand=True,
                            height=70,
                            bgcolor="#111111",
                            border_radius=6,
                            opacity=0.3,
                            border=ft.Border.all(1, "#222222")
                        )
                    )
                
                for d in range(1, num_days + 1):
                    d_sin, d_pzs, accum_sin, accum_pzs = daily_accum_map.get(d, (0.0, 0, 0.0, 0))
                    
                    cell_content = None
                    if d == 1:
                        cell_content = ft.Column([
                            ft.Row([
                                ft.Text(str(d), size=12, weight="bold", color="#00FFFF"),
                            ], alignment="spaceBetween"),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(f"${d_sin:,.0f}", size=11, color="white", weight="bold"),
                                    ft.Text(f"{d_pzs} pzs", size=9, color="#aaaaaa")
                                ], spacing=1, alignment="center"),
                                alignment=ft.alignment.Alignment(0, 0),
                                expand=True
                            )
                        ], spacing=2, expand=True)
                    else:
                        cell_content = ft.Column([
                            ft.Row([
                                ft.Text(str(d), size=12, weight="bold", color="#00FFFF"),
                            ], alignment="spaceBetween"),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(f"${d_sin:,.0f} | {d_pzs}p", size=10, color="white"),
                                ], spacing=0, alignment="center"),
                                alignment=ft.alignment.Alignment(0, 0),
                                height=20
                            ),
                            ft.Divider(height=1, color="#333333"),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(f"${accum_sin:,.0f} | {accum_pzs}p", size=10, color="#7CFC00", weight="bold"),
                                ], spacing=0, alignment="center"),
                                alignment=ft.alignment.Alignment(0, 0),
                                height=20
                            )
                        ], spacing=2, expand=True)
                    
                    cell_container = ft.Container(
                        content=cell_content,
                        expand=True,
                        height=75,
                        bgcolor="#152238" if (d_sin > 0 or d_pzs > 0) else "#111111",
                        border_radius=8,
                        padding=ft.Padding(left=6, top=4, right=6, bottom=4),
                        border=ft.Border.all(1, "#3c5c8c" if (d_sin > 0 or d_pzs > 0) else "#222222"),
                        on_click=lambda e, day_num=d: open_edit_day_dialog(day_num)
                    )
                    cells.append(cell_container)
                
                while len(cells) % 7 != 0:
                    cells.append(
                        ft.Container(
                            expand=True,
                            height=70,
                            bgcolor="#111111",
                            border_radius=6,
                            opacity=0.3,
                            border=ft.Border.all(1, "#222222")
                        )
                    )
                
                for i in range(0, len(cells), 7):
                    week_cells = cells[i:i+7]
                    calendar_grid.controls.append(
                        ft.Row(controls=week_cells, spacing=5)
                    )

            def refresh_data():
                meta_v_con_iva, meta_p, sales_diarias = cargar_datos_presupuesto()
                
                meta_venta_tf.value = str(meta_v_con_iva) if meta_v_con_iva > 0 else ""
                meta_piezas_tf.value = str(meta_p) if meta_p > 0 else ""
                
                meta_v_sin_iva = meta_v_con_iva if meta_v_con_iva > 0 else 0.0
                
                sales_map = {row["Dia"]: (float(row["Venta_Con_IVA"]), float(row["Venta_Sin_IVA"]), int(row["Piezas"])) for row in sales_diarias}
                
                accum_venta_sin_iva = 0.0
                accum_piezas = 0
                
                days_in_month = calendar.monthrange(current_year[0], current_month[0])[1]
                daily_accum_map = {}
                
                for d in range(1, days_in_month + 1):
                    d_con, d_sin, d_pzs = sales_map.get(d, (0.0, 0.0, 0))
                    accum_venta_sin_iva += d_sin
                    accum_piezas += d_pzs
                    daily_accum_map[d] = (d_sin, d_pzs, accum_venta_sin_iva, accum_piezas)
                
                if meta_v_sin_iva > 0:
                    v_ratio = accum_venta_sin_iva / meta_v_sin_iva
                    progress_bar_venta.value = min(1.0, v_ratio)
                    progress_bar_venta.color = "#FF4B4B" if v_ratio < 0.5 else ("#FFCC00" if v_ratio < 1.0 else "#00FF7F")
                    progress_text_venta.value = f"Ventas: {v_ratio*100:.1f}% (${accum_venta_sin_iva:,.2f} / ${meta_v_sin_iva:,.2f} sin IVA)"
                else:
                    progress_bar_venta.value = 0.0
                    progress_bar_venta.color = "#FF4B4B"
                    progress_text_venta.value = f"Ventas: Meta no definida (${accum_venta_sin_iva:,.2f} sin IVA)"
                    
                if meta_p > 0:
                    p_ratio = accum_piezas / meta_p
                    progress_bar_piezas.value = min(1.0, p_ratio)
                    progress_bar_piezas.color = "#FF4B4B" if p_ratio < 0.5 else ("#FFCC00" if p_ratio < 1.0 else "#00FF7F")
                    progress_text_piezas.value = f"Piezas: {p_ratio*100:.1f}% ({accum_piezas} / {meta_p} pzs)"
                else:
                    progress_bar_piezas.value = 0.0
                    progress_bar_piezas.color = "#FF4B4B"
                    progress_text_piezas.value = f"Piezas: Meta no definida ({accum_piezas} pzs)"
                
                tienda_title_txt.value = selected_tienda[0].upper() if selected_tienda[0] else "SELECCIONE TIENDA"
                zona_title_txt.value = f"Zona: {selected_zona[0]}"
                period_title_txt.value = f"{meses_nombres[current_month[0]-1].upper()} {current_year[0]}"
                
                render_meses_logrados()
                render_calendar(daily_accum_map)
                try:
                    page.update()
                except Exception:
                    pass

            def guardar_metas_click(e):
                tienda_actual = selected_tienda[0]
                mes_actual = current_month[0]
                anio_actual = current_year[0]
                
                if not tienda_actual:
                    mostrar_snack("Selecciona una tienda primero.", color="red")
                    return
                
                try:
                    m_venta = float(meta_venta_tf.value.strip() or 0.0)
                    m_piezas = int(meta_piezas_tf.value.strip() or 0)
                except ValueError:
                    mostrar_snack("Por favor ingresa números válidos para las metas.", color="red")
                    return
                
                try:
                    db = conectar_db()
                    if db:
                        cur = db.cursor()
                        cur.execute("""
                            INSERT INTO presupuesto_mensual (Tienda, Mes, Anio, Meta_Venta, Meta_Piezas)
                            VALUES (%s, %s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE 
                                Meta_Venta = %s,
                                Meta_Piezas = %s
                        """, (tienda_actual, mes_actual, anio_actual, m_venta, m_piezas, m_venta, m_piezas))
                        db.commit()
                        db.close()
                        mostrar_snack("Metas guardadas exitosamente.", color="#7CFC00")
                        refresh_data()
                except Exception as ex:
                    print("Error saving month goals:", ex)
                    mostrar_snack("Error al guardar metas.", color="red")

            def on_period_changed(e):
                current_month[0] = int(dd_mes.value)
                current_year[0] = int(dd_anio.value)
                refresh_data()

            def on_zona_changed(e):
                selected_zona[0] = dd_zona.value
                tiendas_zona = tiendas_por_zona.get(dd_zona.value, [])
                dd_tienda.options = [ft.dropdown.Option(t, t) for t in tiendas_zona]
                if tiendas_zona:
                    dd_tienda.value = tiendas_zona[0]
                    selected_tienda[0] = tiendas_zona[0]
                else:
                    dd_tienda.value = ""
                    selected_tienda[0] = ""
                refresh_data()

            def on_tienda_changed(e):
                selected_tienda[0] = dd_tienda.value
                refresh_data()

            dd_mes.on_change = on_period_changed
            dd_anio.on_change = on_period_changed
            dd_zona.on_change = on_zona_changed
            dd_tienda.on_change = on_tienda_changed

            if es_admin() and selected_zona[0] in tiendas_por_zona:
                dd_tienda.options = [ft.dropdown.Option(t, t) for t in tiendas_por_zona[selected_zona[0]]]
                dd_tienda.value = selected_tienda[0]

            filters_row = ft.Row([
                dd_mes,
                dd_anio,
            ], spacing=10, wrap=True)

            if es_admin():
                filters_row.controls.insert(0, dd_zona)
                filters_row.controls.insert(1, dd_tienda)

            left_panel = ft.Column([
                ft.Container(
                    content=ft.Column([
                        ft.Text("Definir Metas del Mes", size=14, color="#D8B4FE", weight="bold"),
                        ft.Row([
                            meta_venta_tf,
                            meta_piezas_tf
                        ], spacing=10),
                        ft.ElevatedButton(
                            "Guardar Metas 💾",
                            bgcolor="#9D50BB",
                            color="white",
                            height=35,
                            on_click=guardar_metas_click
                        )
                    ], spacing=10),
                    bgcolor="#111111",
                    padding=15,
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                ),
                
                ft.Container(
                    content=ft.Column([
                        ft.Text("Avance del Período", size=14, color="#D8B4FE", weight="bold"),
                        progress_text_venta,
                        progress_bar_venta,
                        progress_text_piezas,
                        progress_bar_piezas
                    ], spacing=8),
                    bgcolor="#111111",
                    padding=15,
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                ),
                
                ft.Container(
                    content=ft.Column([
                        ft.Text("Meses Logrados", size=14, color="#D8B4FE", weight="bold"),
                        ft.Text("Cumplimiento anual de ventas", size=11, color="#aaaaaa"),
                        ft.Divider(height=1, color="#333333"),
                        meses_logrados_col
                    ], spacing=8),
                    bgcolor="#111111",
                    padding=15,
                    border_radius=8,
                    border=ft.Border.all(1, "#333333")
                )
            ], spacing=15)

            right_panel = ft.Column([
                ft.Container(
                    content=ft.Row([
                        ft.Column([
                            tienda_title_txt,
                            zona_title_txt
                        ], spacing=2),
                        ft.Container(expand=True),
                        period_title_txt
                    ], vertical_alignment="center"),
                    padding=ft.Padding(left=10, top=5, right=10, bottom=5)
                ),
                calendar_grid
            ], spacing=10, expand=True)

            responsive_layout = ft.ResponsiveRow([
                ft.Container(left_panel, col={"xs": 12, "md": 4}),
                ft.Container(right_panel, col={"xs": 12, "md": 8})
            ], spacing=20)

            main_col = ft.Column([
                ft.Row([
                    ft.Text("Presupuesto Operativo 📊", size=24, color="#D8B4FE", weight="bold")
                ]),
                ft.Text("Monitorea las metas mensuales de ventas y piezas, y registra los resultados diarios sin IVA.", color="#aaaaaa", size=13),
                ft.Divider(height=15, color="#333333"),
                filters_row,
                ft.Container(height=10),
                responsive_layout
            ], scroll=ft.ScrollMode.ALWAYS, expand=True)

            refresh_data()
            return main_col

        actualizar_campana_badge()
        active_view = ["chat"]

        # Cambiar vistas con hover y estilos activos
        def cambiar_vista(vista):
            active_view[0] = vista
            for btn, v_name in [(btn_chat, "chat"), (btn_historial, "historial"), (btn_checklists, "checklists"), (btn_manuales, "manuales"), (btn_tareas, "tareas"), (btn_campanas, "campanas"), (btn_presupuesto, "presupuesto")]:
                if btn:
                    btn.style = ft.ButtonStyle(
                        bgcolor="#222222" if v_name == vista else "transparent",
                        shape=ft.RoundedRectangleBorder(radius=8)
                    )
            if btn_dashboard:
                btn_dashboard.style = ft.ButtonStyle(
                    bgcolor="#222222" if vista == "dashboard" else "transparent",
                    shape=ft.RoundedRectangleBorder(radius=8)
                )
            
            if vista == "chat":
                content_area.content = build_chat_view()
            elif vista == "historial":
                content_area.content = build_historial_view()
            elif vista == "checklists":
                content_area.content = build_checklists_view()
            elif vista == "manuales":
                content_area.content = build_manuals_view()
            elif vista == "tareas":
                content_area.content = build_tareas_view()
            elif vista == "campanas":
                content_area.content = build_campanas_view()
            elif vista == "presupuesto":
                content_area.content = build_presupuesto_view()
            elif vista == "dashboard":
                content_area.content = build_dashboard_view()
            page.update()

        btn_chat = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.CHAT_BUBBLE_ROUNDED, color="#00FFFF"), ft.Text(t("chat"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("chat"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_historial = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.HISTORY, color="#00FFFF"), ft.Text(t("history"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("historial"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_checklists = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.CHECKLIST_ROUNDED, color="#00FFFF"), ft.Text(t("checklists"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("checklists"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_manuales = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.BOOK_ROUNDED, color="#00FFFF"), ft.Text(t("manuals_nav"), color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("manuales"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_tareas = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.ASSIGNMENT_ROUNDED, color="#00FFFF"), ft.Text("Tareas", color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("tareas"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_campanas = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.PHOTO_CAMERA, color="#00FFFF"), ft.Text("Campañas 📸", color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("campanas"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_presupuesto = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.BAR_CHART_ROUNDED, color="#00FFFF"), ft.Text("Presupuesto 📊", color="white", weight="bold")], spacing=10),
            on_click=lambda e: cambiar_vista("presupuesto"),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        btn_dashboard = None
        if es_admin():
            btn_dashboard = ft.TextButton(
                content=ft.Row([ft.Icon(ft.Icons.DASHBOARD_ROUNDED, color="#00FFFF"), ft.Text(t("admin_panel"), color="white", weight="bold")], spacing=10),
                on_click=lambda e: cambiar_vista("dashboard"),
                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
            )

        btn_logout = ft.TextButton(
            content=ft.Row([ft.Icon(ft.Icons.LOGOUT, color="#FF4500"), ft.Text(t("logout"), color="#FF4500", weight="bold")], spacing=10),
            on_click=lambda e: cerrar_sesion(),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8))
        )

        # --- RECUADRO DE SUGERENCIAS EN LA BARRA LATERAL (VISIBLE PARA TODOS) ---
        suggestion_input = ft.TextField(
            label=t("suggestion_hint"),
            multiline=True,
            min_lines=1,
            max_lines=3,
            border_color="#9D50BB",
            color="white",
            text_size=12,
            label_style=ft.TextStyle(color="#aaaaaa", size=11),
            focused_border_color="#00FFFF",
        )
        
        def enviar_sugerencia_click(e):
            text = suggestion_input.value.strip()
            if not text:
                mostrar_snack("Error" if selected_lang[0] != "es" else "Por favor escribe algo antes de enviar.", color="red")
                return
            try:
                db_sug = conectar_db()
                if db_sug:
                    cursor_sug = db_sug.cursor()
                    cursor_sug.execute("""
                        CREATE TABLE IF NOT EXISTS sugerencias_luxo (
                            ID_Sugerencia INT AUTO_INCREMENT PRIMARY KEY,
                            ID_Usuario INT NOT NULL,
                            Fecha_Hora DATETIME DEFAULT CURRENT_TIMESTAMP,
                            Sugerencia TEXT NOT NULL,
                            FOREIGN KEY (ID_Usuario) REFERENCES usuarios(ID_Usuario) ON DELETE CASCADE
                        )
                    """)
                    db_sug.commit()
                    
                    cursor_sug.execute("""
                        INSERT INTO sugerencias_luxo (ID_Usuario, Sugerencia)
                        VALUES (%s, %s)
                    """, (user_info["id"], text))
                    db_sug.commit()
                    db_sug.close()
                    
                    suggestion_input.value = ""
                    mostrar_snack("¡Sugerencia enviada! Gracias." if selected_lang[0] == "es" else "Suggestion sent! Thanks.", color="#7CFC00")
                    page.update()
            except Exception as ex:
                print("ERROR AL ENVIAR SUGERENCIA:", ex)
                mostrar_snack("Error", color="red")

        btn_enviar_sug = ft.ElevatedButton(
            t("send"),
            on_click=enviar_sugerencia_click,
            bgcolor="#9D50BB",
            color="white",
            height=30,
        )

        suggestion_box = ft.Container(
            content=ft.Column([
                ft.Text(t("suggestion_title"), color="#D8B4FE", size=11, weight="bold"),
                suggestion_input,
                ft.Row([btn_enviar_sug], alignment="end")
            ], spacing=5),
            bgcolor="#222222",
            padding=10,
            border_radius=8,
            border=ft.Border.all(1, "#333333"),
        )

        # Dropdown de cambio de idioma en la barra lateral
        def language_changed(e):
            selected_lang[0] = lang_dropdown.value
            
            # Actualizar textos del Sidebar
            sidebar.content.controls[0].value = t("login_title")
            btn_chat.content.controls[1].value = t("chat")
            btn_historial.content.controls[1].value = t("history")
            btn_checklists.content.controls[1].value = t("checklists")
            btn_manuales.content.controls[1].value = t("manuals_nav")
            btn_tareas.content.controls[1].value = "Tareas"
            
            # Traducir botón campañas y presupuesto
            lang = selected_lang[0]
            if lang == "es":
                btn_campanas.content.controls[1].value = "Campañas 📸"
                btn_presupuesto.content.controls[1].value = "Presupuesto 📊"
            elif lang == "fr":
                btn_campanas.content.controls[1].value = "Campagnes 📸"
                btn_presupuesto.content.controls[1].value = "Budget 📊"
            elif lang == "it":
                btn_campanas.content.controls[1].value = "Campagne 📸"
                btn_presupuesto.content.controls[1].value = "Budget 📊"
            elif lang == "zh":
                btn_campanas.content.controls[1].value = "活动 📸"
                btn_presupuesto.content.controls[1].value = "预算 📊"
            else:
                btn_campanas.content.controls[1].value = "Campaigns 📸"
                btn_presupuesto.content.controls[1].value = "Budget 📊"

            if btn_dashboard:
                btn_dashboard.content.controls[1].value = t("admin_panel")
            btn_logout.content.controls[1].value = t("logout")
            
            suggestion_box.content.controls[0].value = t("suggestion_title")
            suggestion_input.label = t("suggestion_hint")
            btn_enviar_sug.text = t("send")
            lang_dropdown.label = t("lang_label")
            
            # Recargar la vista actual para que refleje el nuevo idioma
            cambiar_vista(active_view[0])

        lang_dropdown = ft.Dropdown(
            label=t("lang_label"),
            value=selected_lang[0],
            border_color="#9D50BB",
            color="white",
            text_size=12,
            label_style=ft.TextStyle(color="#aaaaaa", size=11),
            focused_border_color="#00FFFF",
            options=[
                ft.dropdown.Option("es", "Español"),
                ft.dropdown.Option("en", "English"),
                ft.dropdown.Option("fr", "Français"),
                ft.dropdown.Option("it", "Italiano"),
                ft.dropdown.Option("zh", "中文")
            ],
            width=200,
            height=45
        )
        lang_dropdown.on_change = language_changed

        # --- FILTRO DE ZONA ACTIVA PARA SUPERVISORES ---
        # (Definido al inicio de main() para persistir el estado entre vistas)

        def cambiar_zona_filtro_admin(e):
            active_zone_filter[0] = e.control.value
            mostrar_snack(f"Filtrando datos de: {active_zone_filter[0]}", color="#00FFFF")
            cambiar_vista(active_view[0])

        dropdown_zona_admin = ft.Dropdown(
            label="Zona de Supervisión Activa",
            value=active_zone_filter[0],
            options=[
                ft.dropdown.Option("Todas", "Todas las Zonas 🌍"),
                ft.dropdown.Option("Zona Norte", "Zona Norte"),
                ft.dropdown.Option("Zona Sur", "Zona Sur"),
                ft.dropdown.Option("Zona Occidente", "Zona Occidente"),
                ft.dropdown.Option("Zona Centro", "Zona Centro"),
                ft.dropdown.Option("Palacio de Hierro", "Palacio de Hierro")
            ],
            border_color="#00FFFF",
            width=200,
            height=45
        )
        dropdown_zona_admin.on_change = cambiar_zona_filtro_admin

        sidebar_items = [
            ft.Text(t("login_title"), size=20, weight="bold", color="#D8B4FE"),
            ft.Divider(height=20, color="#444444"),
            profile_row,
            ft.Divider(height=20, color="#444444"),
            btn_chat,
            btn_historial,
            btn_checklists,
            btn_manuales,
            btn_tareas,
            btn_campanas,
            btn_presupuesto,
        ]

        if btn_dashboard:
            sidebar_items.append(btn_dashboard)

        if es_admin():
            sidebar_items.extend([
                ft.Divider(height=10, color="#444444"),
                dropdown_zona_admin
            ])

        sidebar_items.extend([
            ft.Container(height=10),
            suggestion_box,
            ft.Container(expand=True),
            lang_dropdown,
            btn_logout
        ])

        sidebar = ft.Container(
            content=ft.Column(sidebar_items, spacing=10),
            width=240,
            padding=20,
            bgcolor="#111111",
            border_radius=10,
            border=ft.Border.all(1, "#333333")
        )

        dashboard_layout = ft.Row([
            sidebar,
            ft.VerticalDivider(width=1, color="#333333"),
            content_area
        ], expand=True)

        cambiar_vista("chat")

        if img_fondo:
            page.add(
                ft.Container(
                    expand=True,
                    image=ft.DecorationImage(
                        src=img_fondo,
                        fit=ft.controls.box.BoxFit.COVER
                    ),
                    content=dashboard_layout
                )
            )
        else:
            page.add(dashboard_layout)

        page.update()

    # =====================================
    # LOGIN
    # =====================================

    def login_click(e):

        db = conectar_db()

        if not db:
            mostrar_snack("Error Base de Datos", color="#FF4B4B")
            return

        cursor = db.cursor(dictionary=True)

        cursor.execute(
            """
            SELECT
            ID_Usuario,
            Nombre_Completo,
            Rol,
            Tienda,
            Zona
            FROM usuarios
            WHERE Usuario = %s
            AND BINARY Contrasena = %s
            """,
            (
                txt_user.value,
                txt_pass.value
            )
        )

        res = cursor.fetchone()

        if res:
            login_message.value = ""
            login_error_box.visible = False

            user_info["id"] = res["ID_Usuario"]
            user_info["nombre"] = res["Nombre_Completo"]
            user_info["rol"] = res["Rol"]
            user_info["tienda"] = res["Tienda"] if res["Tienda"] is not None else ""
            user_info["zona"] = res["Zona"] if res["Zona"] is not None else "Zona Centro"

            # --- REGISTRAR INICIO DE SESIÓN ---
            ip_client = getattr(page, "client_ip", None) or "Desconocido"
            
            def registrar_sesion_async(u_id, ip):
                city = "Desconocido"
                country = "Desconocido"
                is_local = False
                
                if not ip or ip == "Desconocido":
                    is_local = True
                else:
                    ip_clean = ip.strip()
                    if ip_clean in ("127.0.0.1", "::1", "localhost") or \
                       ip_clean.startswith("192.168.") or \
                       ip_clean.startswith("10.") or \
                       ip_clean.startswith("172.16.") or \
                       ip_clean.startswith("fe80:"):
                        is_local = True
                
                if is_local:
                    city = "Localhost"
                    country = "Local / Desarrollo"
                else:
                    try:
                        resp = requests.get(f"http://ip-api.com/json/{ip}", timeout=5)
                        if resp.status_code == 200:
                            data = resp.json()
                            if data.get("status") == "success":
                                city = data.get("city", "Desconocido")
                                country = data.get("country", "Desconocido")
                    except Exception as err:
                        print("Error consultando geolocalización de IP:", err)
                        
                try:
                    db_log = conectar_db()
                    if db_log:
                        cursor_log = db_log.cursor()
                        cursor_log.execute(
                            """
                            INSERT INTO sesiones (ID_Usuario, Direccion_IP, Ubicacion_Ciudad, Ubicacion_Pais)
                            VALUES (%s, %s, %s, %s)
                            """,
                            (u_id, ip, city, country)
                        )
                        db_log.commit()
                        db_log.close()
                except Exception as err_db:
                    print("Error al guardar sesión en BD:", err_db)

            threading.Thread(target=registrar_sesion_async, args=(user_info["id"], ip_client), daemon=True).start()

            cargar_chat()

        else:
            cursor.execute(
                """
                SELECT
                ID_Usuario
                FROM usuarios
                WHERE Usuario = %s
                """,
                (txt_user.value,)
            )
            usuario_existe = cursor.fetchone()

            if usuario_existe:
                mensaje = "Contraseña incorrecta"
            else:
                mensaje = "Usuario no registrado"

            login_message.value = mensaje
            login_message.color = "#FF4B4B"
            login_error_box.visible = True
            page.update()

        db.close()

    # =====================================
    # LOGIN UI
    # =====================================

    txt_user = ft.TextField(
        label="Usuario",
        width=300
    )

    txt_pass = ft.TextField(
        label="Contraseña",
        password=True,
        width=300,
        on_submit=login_click
    )

    login_message = ft.Text(
        "",
        size=16,
        weight="bold",
        color="#FF4B4B"
    )

    login_error_box = ft.Container(
        content=login_message,
        bgcolor="#000000",
        padding=10,
        border_radius=10,
        visible=False,
        width=300
    )

    video_avatar = None
    if os.path.exists(video_path):
        video_avatar = ft.Container(
            content=fv.Video(
                playlist=[fv.VideoMedia(video_path)],
                playlist_mode=fv.PlaylistMode.LOOP,
                autoplay=True,
                muted=True,
                controls=None,
                width=120,
                height=120,
            ),
            width=120,
            height=120,
            border_radius=60,
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
            border=ft.Border.all(2, "#00FFFF"),
        )

    login_ui = ft.Container(

        content=ft.Column([

            video_avatar if video_avatar else (
                ft.Image(
                    src=img_avatar,
                    width=120,
                    height=120,
                    fit=ft.controls.box.BoxFit.COVER
                ) if img_avatar else ft.Text(
                    "LUXO",
                    size=30,
                    color="#FFFFFF",
                    weight="bold"
                )
            ),

            login_error_box,

            ft.Text(
                "SISTEMA LUXO",
                size=25,
                weight="bold",
                color="#D8B4FE"
            ),

            txt_user,

            txt_pass,

            ft.ElevatedButton(
                "INGRESAR",
                on_click=login_click,
                width=300,
                bgcolor="#6E48AA",
                color="white"
            )

        ],
        horizontal_alignment="center",
        spacing=20),

        padding=40,
        bgcolor="#000000",
        border_radius=20,
        border=ft.Border.all(2, "#00FFFF"),
        shadow=[
            ft.BoxShadow(
                color="#00FFFF",
                blur_radius=15,
                spread_radius=1,
            )
        ],
        clip_behavior=ft.ClipBehavior.HARD_EDGE
    )

    page.vertical_alignment = "center"

    page.horizontal_alignment = "center"

    page.add(login_ui)

# =========================================
# EJECUTAR APP
# =========================================

if __name__ == "__main__":
    ft.app(target=main, assets_dir="assets")

