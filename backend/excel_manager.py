# =========================================
# excel_manager.py — Gestión de Excels (Web)
# =========================================

import io
from typing import Optional
import openpyxl
import database
import vector_store
from text_cleaner import limpiar_texto_excel

MAX_CHARS_EXCEL = 5_000_000  # Límite muy alto para leer excels completos sin truncar


def extraer_texto_excel_bytes(contenido_bytes: bytes, max_chars: int = MAX_CHARS_EXCEL) -> str:
    """
    Extrae todo el texto de un archivo Excel con formato orientado al contexto.

    Mejoras vs versión anterior:
    - La primera fila de cada hoja se trata como encabezado y se repite
      como prefijo en cada chunk, dando contexto al LLM.
    - Formato: "NombreColumna: valor" en lugar de "[A1]: valor"
      → los embeddings capturan mejor la semántica.
    - Limpieza de texto con text_cleaner.

    Args:
        contenido_bytes: Bytes del archivo Excel
        max_chars: Límite de caracteres (default 500K)

    Returns:
        Texto estructurado listo para indexar en ChromaDB
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(contenido_bytes), read_only=True, data_only=True
        )
        partes = []
        total_chars = 0

        for nombre_hoja in wb.sheetnames:
            if total_chars >= max_chars:
                partes.append(f"\n[TRUNCADO: límite de {max_chars:,} caracteres alcanzado]")
                break

            partes.append(f"\n=== Hoja: {nombre_hoja} ===")
            total_chars += len(partes[-1])

            # Leer todas las filas de la hoja
            filas = list(ws.iter_rows(values_only=True) for ws in [wb[nombre_hoja]])[0]
            filas_lista = [list(fila) for fila in filas]

            if not filas_lista:
                partes.append("(Hoja vacía)")
                continue

            # Detectar encabezados: primera fila no vacía
            encabezados = None
            inicio_datos = 0
            for idx, fila in enumerate(filas_lista):
                valores_fila = [str(v).strip() for v in fila if v is not None and str(v).strip()]
                if valores_fila:
                    encabezados = [str(h).strip() if h is not None else f"Col{i+1}"
                                   for i, h in enumerate(fila)]
                    inicio_datos = idx + 1
                    break

            if encabezados:
                partes.append(f"Columnas: {' | '.join(h for h in encabezados if h)}")

            filas_con_datos = []
            for fila in filas_lista[inicio_datos:]:
                if total_chars >= max_chars:
                    filas_con_datos.append("[... filas omitidas por límite de tamaño]")
                    break

                # Construir "NombreColumna: valor" para mejor semántica
                celdas = []
                for i, valor in enumerate(fila):
                    if valor is None or str(valor).strip() == "":
                        continue
                    nombre_col = encabezados[i] if (encabezados and i < len(encabezados)) else f"Col{i+1}"
                    # Omitir columnas sin nombre o vacías
                    if not nombre_col or nombre_col.startswith("Col") and len(nombre_col) <= 5:
                        celdas.append(str(valor).strip())
                    else:
                        celdas.append(f"{nombre_col}: {str(valor).strip()}")

                if celdas:
                    linea = " | ".join(celdas)
                    filas_con_datos.append(linea)
                    total_chars += len(linea)

            if filas_con_datos:
                partes.append("\n".join(filas_con_datos))
            else:
                partes.append("(No hay datos en esta hoja)")

        wb.close()
        texto_raw = "\n".join(partes)
        return limpiar_texto_excel(texto_raw)

    except Exception as e:
        print("ERROR EXTRAER EXCEL:", e)
        return ""


def normalizar_nombre_excel(nombre_archivo: str) -> str:
    """Convierte el nombre del archivo Excel a MAYÚSCULAS para homologar."""
    return nombre_archivo.upper() if nombre_archivo else nombre_archivo


def cargar_excel(nombre_archivo: str, contenido_bytes: bytes) -> tuple[bool, str]:
    """
    Carga un nuevo Excel al sistema desde bytes (upload HTTP):
    1. Normaliza el nombre a MAYÚSCULAS
    2. Extrae texto con encabezados de contexto
    3. Inserta en MySQL (texto + binario)
    4. Indexa en ChromaDB con chunking especializado para Excel

    Retorna: (exito: bool, mensaje: str)
    """
    try:
        nombre_archivo = normalizar_nombre_excel(nombre_archivo)

        texto = extraer_texto_excel_bytes(contenido_bytes)
        if not texto.strip():
            return False, f"El Excel '{nombre_archivo}' no contiene texto extraíble."

        id_manual = database.insertar_manual_excel(nombre_archivo, contenido_bytes, texto)
        if not id_manual:
            return False, "Error al guardar en la base de datos."

        # Usar chunking especializado para Excel (chunks más grandes, por sección)
        vector_store.indexar_manual(id_manual, nombre_archivo, texto, tipo="excel")

        return True, f"Excel '{nombre_archivo}' cargado ({len(texto):,} chars extraídos)."

    except Exception as e:
        print("ERROR CARGAR EXCEL:", e)
        return False, f"Error al cargar Excel: {e}"


def actualizar_excel(nombre_archivo: str, contenido_bytes: bytes) -> tuple[bool, str]:
    """
    Actualiza un manual Excel existente o carga uno nuevo si no existe.

    Retorna: (exito: bool, mensaje: str)
    """
    try:
        nombre_archivo = normalizar_nombre_excel(nombre_archivo)

        texto = extraer_texto_excel_bytes(contenido_bytes)

        existente = database.buscar_manual_por_nombre(nombre_archivo)

        if existente:
            version_actual = existente.get("Version") or "1.0"
            try:
                nueva_version = str(round(float(version_actual) + 0.1, 1))
            except Exception:
                nueva_version = version_actual

            database.actualizar_manual_excel(
                existente["ID_Manual"], contenido_bytes, texto, nueva_version
            )
            vector_store.indexar_manual(existente["ID_Manual"], nombre_archivo, texto, tipo="excel")

            return True, f"Excel '{nombre_archivo}' actualizado a versión {nueva_version}."
        else:
            return cargar_excel(nombre_archivo, contenido_bytes)

    except Exception as e:
        print("ERROR ACTUALIZAR EXCEL:", e)
        return False, f"Error al actualizar el Excel: {e}"


def obtener_excel_para_descarga(id_manual: int) -> Optional[dict]:
    """
    Obtiene el binario y nombre del Excel para enviar como respuesta HTTP.

    Retorna: dict {nombre, contenido_bytes} o None
    """
    try:
        manual = database.obtener_excel_binario(id_manual)
        if not manual or not manual.get("Archivo_Excel"):
            return None

        return {
            "nombre": manual["Nombre_Archivo"],
            "contenido_bytes": manual["Archivo_Excel"],
        }
    except Exception as e:
        print("ERROR OBTENER EXCEL:", e)
        return None
